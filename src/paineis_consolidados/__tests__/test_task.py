import io
import os
import re
import tempfile
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from faker import Faker
from freezegun import freeze_time
from openpyxl import load_workbook

from src.dados_comuns.models import LogSolicitacoesUsuario
from src.paineis_consolidados.api import constants
from src.paineis_consolidados.api.serializers import SolicitacoesExportXLSXSerializer
from src.paineis_consolidados.models import SolicitacoesCODAE
from src.paineis_consolidados.tasks import (
    aplica_fundo_amarelo_canceladas,
    aplica_fundo_amarelo_tipo1,
    aplica_fundo_amarelo_tipo2,
    build_pdf,
    build_subtitulo,
    build_xlsx,
    gera_pdf_relatorio_solicitacoes_alimentacao_async,
    gera_xls_relatorio_solicitacoes_alimentacao_async,
    get_formats,
    nomes_colunas,
    novas_linhas_inc_continua_e_kit_lanche,
)
from src.relatorios.utils import extrair_texto_de_pdf

pytestmark = pytest.mark.django_db
f = Faker("pt-br")


def test_xls_relatorio_status(users_diretor_escola):
    usuario = users_diretor_escola[5]
    uuids = ["7fa6e609-db33-48e1-94ea-6d5a0c07935c"]
    lotes = []
    tipos_solicitacao = []
    tipos_unidade = []
    unidades_educacionais = []
    request_data = {
        "status": "CANCELADOS",
        "tipos_solicitacao": ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    resultado = gera_xls_relatorio_solicitacoes_alimentacao_async.delay(
        user=usuario.username,
        nome_arquivo="relatorio_solicitacoes_alimentacao.xlsx",
        data=request_data,
        uuids=uuids,
        lotes=lotes,
        tipos_solicitacao=tipos_solicitacao,
        tipos_unidade=tipos_unidade,
        unidades_educacionais=unidades_educacionais,
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


def test_pdf_relatorio_status(users_diretor_escola):
    usuario = users_diretor_escola[5]
    uuids = ["7fa6e609-db33-48e1-94ea-6d5a0c07935c"]
    request_data = {
        "status": "CANCELADOS",
        "tipos_solicitacao": ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    resultado = gera_pdf_relatorio_solicitacoes_alimentacao_async.delay(
        user=usuario.username,
        nome_arquivo="relatorio_solicitacoes_alimentacao.pdf",
        data=request_data,
        uuids=uuids,
        status=request_data.get("status", None),
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


@freeze_time("2025-01-22")
def test_build_xlsx(dados_para_geracao_excel_e_pdf):
    instituicao, sem_uuids_repetido, serializer, data = dados_para_geracao_excel_e_pdf
    output = io.BytesIO()
    lotes = []
    tipos_solicitacao = ["SUSP_ALIMENTACAO", "INC_ALIMENTA"]
    tipos_unidade = []
    unidades_educacionais = []
    build_xlsx(
        output,
        serializer,
        sem_uuids_repetido,
        data,
        lotes,
        tipos_solicitacao,
        tipos_unidade,
        unidades_educacionais,
        instituicao,
    )
    workbook = load_workbook(output)
    sheet = workbook[f"Relatório - Canceladas"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Solicitações de Alimentação Canceladas",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "Total de Solicitações Canceladas: 11 | Tipo(s) de solicitação(ões): Suspensão de Alimentação, Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "N",
        "Lote",
        "Unidade Educacional",
        "Tipo de Solicitação",
        "ID da Solicitação",
        "Data do Evento",
        "Encerrado a partir de",
        "Dia da semana",
        "Período",
        "Tipo de Alimentação",
        "Tipo de Alteração",
        "Nª de Alunos",
        "N° total de Kits",
        "Observações",
        "Data de Cancelamento",
    )
    for row in rows:
        if row[4] == "Inclusão de Alimentação":
            assert row[5] == "14/01/2025"
            assert row[6] == "-"
            assert row[13] == "cancelado"


def test_novas_linhas_inc_continua_e_kit_lanche(dados_para_geracao_excel_e_pdf):
    instituicao, queryset, serializer, _ = dados_para_geracao_excel_e_pdf

    df = pd.DataFrame(serializer.data)
    novas_colunas = [
        "encerrado_a_partir_de",
        "dia_semana",
        "periodo_inclusao",
        "tipo_alimentacao",
    ]
    for i, nova_coluna in enumerate(novas_colunas):
        df.insert(constants.COL_IDX_DATA_EVENTO + i, nova_coluna, "-")
    df.insert(constants.COL_IDX_NUMERO_DE_ALUNOS, "quantidade_alimentacoes", "-")
    novas_linhas, lista_uuids = novas_linhas_inc_continua_e_kit_lanche(
        df, queryset, instituicao
    )

    assert isinstance(novas_linhas, list)
    for linha in novas_linhas:
        assert isinstance(linha, pd.Series)

    assert isinstance(lista_uuids, list)
    for lu in lista_uuids:
        assert isinstance(lu, SolicitacoesCODAE)


def test_get_formats():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_file = tmp.name
    xlwriter = pd.ExcelWriter(output_file, engine="xlsxwriter")
    workbook = xlwriter.book

    merge_format, cell_format, single_cell_format = get_formats(workbook)
    assert merge_format.bold == True

    assert cell_format.bold == True
    assert cell_format.text_wrap == True

    workbook.close()
    xlwriter.close()
    if os.path.exists(output_file):
        os.remove(output_file)
    Path(output_file).unlink(missing_ok=True)


@freeze_time("2025-01-22")
def test_build_subtitulo_autorizadas(dados_para_geracao_excel_e_pdf):
    _, queryset, _, _ = dados_para_geracao_excel_e_pdf

    data = {
        "status": "AUTORIZADAS",
        "tipos_solicitacao": ["INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    subtitulo_autorizadas = build_subtitulo(
        data,
        "Autorizadas",
        queryset,
        [],
        ["INC_ALIMENTA"],
        [],
        [],
    )
    assert (
        subtitulo_autorizadas
        == "Total de Solicitações Autorizadas: 11 | Tipo(s) de solicitação(ões): Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025"
    )


@freeze_time("2025-01-22")
def test_build_subtitulo_canceladas(dados_para_geracao_excel_e_pdf):
    _, queryset, _, data = dados_para_geracao_excel_e_pdf

    subtitulo_canceladas = build_subtitulo(
        data,
        "Canceladas",
        queryset,
        [],
        ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        [],
        [],
    )
    assert (
        subtitulo_canceladas
        == "Total de Solicitações Canceladas: 11 | Tipo(s) de solicitação(ões): Suspensão de Alimentação, Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025"
    )


def test_nomes_colunas(dados_para_montar_excel):
    (
        linhas,
        colunas,
        _,
        _,
        _,
        worksheet,
        single_cell_format,
        nome_aba,
    ) = dados_para_montar_excel

    nomes_colunas(worksheet, "Canceladas", linhas, colunas, single_cell_format)

    assert 15 == len(colunas)
    assert len(worksheet.row_sizes) == 2
    assert worksheet.name == nome_aba


def test_build_xlsx_inclusao_continua_exibe_coluna_encerrado_e_traceja_data_final(
    escola_factory,
    motivo_inclusao_continua_factory,
    inclusao_alimentacao_continua_factory,
    quantidade_por_periodo_factory,
    tipo_alimentacao_factory,
    periodo_escolar_factory,
    usuario_factory,
    log_solicitacoes_usuario_factory,
):
    escola = escola_factory.create(nome="EMEF PERICLES EUGENIO DA SILVA RAMOS")
    motivo = motivo_inclusao_continua_factory.create(
        nome="Programas/Projetos Contínuos"
    )
    periodo_integral = periodo_escolar_factory.create(nome="INTEGRAL")
    periodo_tarde = periodo_escolar_factory.create(nome="TARDE")
    tipo_lanche = tipo_alimentacao_factory.create(nome="Lanche")
    tipo_refeicao = tipo_alimentacao_factory.create(nome="Refeição")
    usuario = usuario_factory.create()

    inclusao = inclusao_alimentacao_continua_factory.create(
        escola=escola,
        motivo=motivo,
        rastro_escola=escola,
        rastro_dre=escola.diretoria_regional,
        rastro_lote=escola.lote,
        rastro_terceirizada=escola.lote.terceirizada,
        data_inicial=datetime(2026, 5, 12).date(),
        data_final=datetime(2026, 12, 31).date(),
        status="CODAE_AUTORIZADO",
    )

    quantidade_por_periodo_factory.create(
        inclusao_alimentacao_continua=inclusao,
        grupo_inclusao_normal=None,
        periodo_escolar=periodo_integral,
        numero_alunos=40,
        tipos_alimentacao=[tipo_lanche],
        dias_semana=[0, 1, 2, 3],
        observacao="Observação integral",
        encerrado_a_partir_de=datetime(2026, 5, 21).date(),
    )
    quantidade_por_periodo_factory.create(
        inclusao_alimentacao_continua=inclusao,
        grupo_inclusao_normal=None,
        periodo_escolar=periodo_tarde,
        numero_alunos=21,
        tipos_alimentacao=[tipo_refeicao],
        dias_semana=[1, 2, 3],
        observacao="Observação tarde",
    )

    log_solicitacoes_usuario_factory.create(
        uuid_original=inclusao.uuid,
        usuario=usuario,
        solicitacao_tipo=LogSolicitacoesUsuario.INCLUSAO_ALIMENTACAO_CONTINUA,
        status_evento=LogSolicitacoesUsuario.CODAE_AUTORIZOU,
    )

    queryset = list(SolicitacoesCODAE.objects.filter(uuid=inclusao.uuid))
    serializer = SolicitacoesExportXLSXSerializer(
        queryset,
        context={
            "instituicao": escola,
            "status": "AUTORIZADOS",
        },
        many=True,
    )

    output = io.BytesIO()
    build_xlsx(
        output,
        serializer,
        queryset,
        {"status": "AUTORIZADOS"},
        [],
        [],
        [],
        [],
        escola,
    )

    conteudo = output.getvalue()
    workbook = load_workbook(io.BytesIO(conteudo))
    sheet = workbook["Relatório - Autorizadas"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[3][6] == "Encerrado a partir de"
    assert rows[4][5] == "12/05/2026 - 31/12/2026"
    assert rows[4][6] == "21/05/2026"
    assert rows[5][6] == "-"

    with zipfile.ZipFile(io.BytesIO(conteudo)) as arquivo_xlsx:
        shared_strings = arquivo_xlsx.read("xl/sharedStrings.xml").decode("utf-8")

    assert re.search(
        r"12/05/2026\s*-\s*</t>.*?<strike/>.*?<t>31/12/2026</t>",
        shared_strings,
        re.S,
    )


def test_aplica_fundo_amarelo_canceladas(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    queryset = [
        solicitacao
        for solicitacao in queryset
        if solicitacao.tipo_doc in ["INC_ALIMENTA", "SUSP_ALIMENTACAO"]
    ]
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)

    aplica_fundo_amarelo_canceladas(df, worksheet, workbook, queryset, linhas, colunas)

    xlwriter.close()
    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.coordinate == "F5":
                assert cell.fill.start_color.rgb == "FFFFFF00"
    workbook_openpyxl.close()


def test_aplica_fundo_amarelo_tipo1(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    queryset = [
        solicitacao
        for solicitacao in queryset
        if solicitacao.tipo_doc in ["INC_ALIMENTA", "SUSP_ALIMENTACAO"]
    ]
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)
    indexes_cancelados = []

    for index, solicitacao in enumerate(queryset):
        model_obj = solicitacao.get_raw_model.objects.get(uuid=solicitacao.uuid)
        if model_obj.status == "ESCOLA_CANCELOU":
            indexes_cancelados.append(index)
            aplica_fundo_amarelo_tipo1(
                df, worksheet, workbook, solicitacao, model_obj, linhas, colunas, index
            )

    xlwriter.close()
    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for index, row in enumerate(sheet.iter_rows()):
        for cell in row:
            if cell.coordinate == "F5" and index in indexes_cancelados:
                assert cell.fill.start_color.rgb == "FFFFFF00"

    workbook_openpyxl.close()


def test_aplica_fundo_amarelo_tipo2(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    queryset = [
        solicitacao
        for solicitacao in queryset
        if solicitacao.tipo_doc in ["INC_ALIMENTA", "SUSP_ALIMENTACAO"]
    ]
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)

    solicitacao = queryset[0]
    model_obj = solicitacao.get_raw_model.objects.get(uuid=solicitacao.uuid)
    idx = aplica_fundo_amarelo_tipo2(
        df, worksheet, workbook, queryset[0], model_obj, linhas, colunas, 0, 0
    )
    xlwriter.close()
    assert idx == 0

    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for row in sheet.iter_rows():
        for cell in row:
            assert cell.fill.start_color.rgb == "00000000"
    workbook_openpyxl.close()


def test_build_pdf():
    pdf_cancelados = build_pdf([], "CANCELADOS")
    assert isinstance(pdf_cancelados, bytes)

    texto = extrair_texto_de_pdf(pdf_cancelados).lower()
    assert "Total de Solicitações Cancelados: 0".lower() in texto
    assert "SIGPAE - RELATÓRIO DE solicitações de alimentação".lower() in texto
    assert datetime.now().strftime("%d/%m/%Y") in texto
    assert texto.count("cancelados") == 2

    pdf_autorizados = build_pdf([], "AUTORIZADOS")
    assert isinstance(pdf_autorizados, bytes)

    texto = extrair_texto_de_pdf(pdf_autorizados).lower()
    assert "Total de Solicitações autorizados: 0".lower() in texto
    assert "SIGPAE - RELATÓRIO DE solicitações de alimentação".lower() in texto
    assert datetime.now().strftime("%d/%m/%Y") in texto
    assert texto.count("autorizados") == 2


def test_build_pdf_inclusao_continua_exibe_coluna_encerramento_historicos_e_nao_traceja_valor(
    escola_factory,
    motivo_inclusao_continua_factory,
    inclusao_alimentacao_continua_factory,
    quantidade_por_periodo_factory,
    tipo_alimentacao_factory,
    periodo_escolar_factory,
    usuario_factory,
    log_solicitacoes_usuario_factory,
    monkeypatch,
):
    monkeypatch.setattr(
        "src.paineis_consolidados.tasks.html_to_pdf_file",
        lambda html_string, *_args, **_kwargs: html_string,
    )

    escola = escola_factory.create(nome="EMEF PERICLES EUGENIO DA SILVA RAMOS")
    motivo = motivo_inclusao_continua_factory.create(
        nome="Programas/Projetos Contínuos"
    )
    periodo_integral = periodo_escolar_factory.create(nome="INTEGRAL")
    periodo_tarde = periodo_escolar_factory.create(nome="TARDE")
    tipo_lanche = tipo_alimentacao_factory.create(nome="Lanche")
    tipo_refeicao = tipo_alimentacao_factory.create(nome="Refeição")
    tipo_lanche_4h = tipo_alimentacao_factory.create(nome="Lanche 4h")
    usuario = usuario_factory.create()

    inclusao = inclusao_alimentacao_continua_factory.create(
        escola=escola,
        motivo=motivo,
        rastro_escola=escola,
        rastro_dre=escola.diretoria_regional,
        rastro_lote=escola.lote,
        rastro_terceirizada=escola.lote.terceirizada,
        data_inicial=datetime(2026, 5, 12).date(),
        data_final=datetime(2026, 12, 31).date(),
        status="CODAE_AUTORIZADO",
    )

    quantidade_por_periodo_factory.create(
        inclusao_alimentacao_continua=inclusao,
        grupo_inclusao_normal=None,
        periodo_escolar=periodo_integral,
        numero_alunos=40,
        tipos_alimentacao=[tipo_lanche],
        dias_semana=[0, 1, 2, 3],
        encerrado_a_partir_de=datetime(2026, 5, 21).date(),
        cancelado_justificativa="Encerramento do projeto no período integral.",
    )
    quantidade_por_periodo_factory.create(
        inclusao_alimentacao_continua=inclusao,
        grupo_inclusao_normal=None,
        periodo_escolar=periodo_integral,
        numero_alunos=21,
        tipos_alimentacao=[tipo_refeicao],
        dias_semana=[1, 2, 3],
        cancelado=True,
        cancelado_justificativa="",
    )
    quantidade_por_periodo_factory.create(
        inclusao_alimentacao_continua=inclusao,
        grupo_inclusao_normal=None,
        periodo_escolar=periodo_tarde,
        numero_alunos=2,
        tipos_alimentacao=[tipo_lanche_4h],
        dias_semana=[2, 3],
        encerrado_a_partir_de=datetime(2026, 5, 30).date(),
        cancelado_justificativa="Encerramento do projeto no período da tarde.",
    )

    log_alteracao_1 = log_solicitacoes_usuario_factory.create(
        uuid_original=inclusao.uuid,
        usuario=usuario,
        solicitacao_tipo=LogSolicitacoesUsuario.INCLUSAO_ALIMENTACAO_CONTINUA,
        status_evento=LogSolicitacoesUsuario.ESCOLA_ALTEROU_ENCERRAMENTO_INCLUSAO_CONTINUA,
        justificativa="Encerramento do projeto no período integral.",
    )
    log_alteracao_1.criado_em = datetime(2026, 5, 6, 17, 22, 51)
    log_alteracao_1.save(update_fields=["criado_em"])

    log_alteracao_2 = log_solicitacoes_usuario_factory.create(
        uuid_original=inclusao.uuid,
        usuario=usuario,
        solicitacao_tipo=LogSolicitacoesUsuario.INCLUSAO_ALIMENTACAO_CONTINUA,
        status_evento=LogSolicitacoesUsuario.ESCOLA_ALTEROU_ENCERRAMENTO_INCLUSAO_CONTINUA,
        justificativa="Encerramento do projeto no período da tarde.",
    )
    log_alteracao_2.criado_em = datetime(2026, 5, 8, 10, 9, 56)
    log_alteracao_2.save(update_fields=["criado_em"])

    solicitacao_dict = inclusao.solicitacao_dict_para_relatorio(
        "de Autorização", "08/05/2026", escola
    )
    html_string = build_pdf([solicitacao_dict], "AUTORIZADOS")

    assert "Encerrado a partir de:" in html_string
    assert "Histórico de alteração:" in html_string
    assert "Histórico de cancelamento:" in html_string
    assert (
        "INTEGRAL - Lanche - 40 - Encerramento previsto para:\n                21/05/2026"
        in html_string
    )
    assert "INTEGRAL - Refeição - 21 - justificativa:" in html_string
    assert re.search(
        r'class="col-2 encerrado-a-partir-relatorio-sol-alim">\s*<b>21/05/2026</b>\s*</div>',
        html_string,
    )
    assert "encerrado-a-partir-relatorio-sol-alim dia-cancelado" not in html_string
    assert "#c0392b" not in html_string


def test_build_pdf_inclusao_continua_destaca_data_final_quando_todas_quantidades_encerram_mesma_data(
    escola_factory,
    motivo_inclusao_continua_factory,
    inclusao_alimentacao_continua_factory,
    quantidade_por_periodo_factory,
    tipo_alimentacao_factory,
    periodo_escolar_factory,
    monkeypatch,
):
    monkeypatch.setattr(
        "src.paineis_consolidados.tasks.html_to_pdf_file",
        lambda html_string, *_args, **_kwargs: html_string,
    )

    escola = escola_factory.create(nome="EMEF PERICLES EUGENIO DA SILVA RAMOS")
    motivo = motivo_inclusao_continua_factory.create(
        nome="Programas/Projetos Contínuos"
    )
    periodo_integral = periodo_escolar_factory.create(nome="INTEGRAL")
    periodo_tarde = periodo_escolar_factory.create(nome="TARDE")
    tipo_lanche = tipo_alimentacao_factory.create(nome="Lanche")

    inclusao = inclusao_alimentacao_continua_factory.create(
        escola=escola,
        motivo=motivo,
        rastro_escola=escola,
        rastro_dre=escola.diretoria_regional,
        rastro_lote=escola.lote,
        rastro_terceirizada=escola.lote.terceirizada,
        data_inicial=datetime(2026, 5, 12).date(),
        data_final=datetime(2026, 12, 31).date(),
        status="CODAE_AUTORIZADO",
    )

    for periodo in [periodo_integral, periodo_tarde]:
        quantidade_por_periodo_factory.create(
            inclusao_alimentacao_continua=inclusao,
            grupo_inclusao_normal=None,
            periodo_escolar=periodo,
            numero_alunos=10,
            tipos_alimentacao=[tipo_lanche],
            dias_semana=[1, 2, 3],
            encerrado_a_partir_de=datetime(2026, 5, 21).date(),
            cancelado_justificativa="Encerramento geral",
        )

    solicitacao_dict = inclusao.solicitacao_dict_para_relatorio(
        "de Autorização", "08/05/2026", escola
    )
    html_string = build_pdf([solicitacao_dict], "AUTORIZADOS")

    assert "Encerrado a partir de:" not in html_string
    assert (
        '<span style="text-decoration: line-through;">31/12/2026</span>' in html_string
    )
    assert "<span>21/05/2026</span>" in html_string
    assert "#c0392b" not in html_string

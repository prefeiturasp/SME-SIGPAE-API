from io import BytesIO

import openpyxl
import pandas as pd
import pytest
from model_bakery import baker
from openpyxl import load_workbook

from src.dados_comuns.constants import (
    DIETA_ESPECIAL_TIPO_A,
    DIETA_ESPECIAL_TIPO_B,
    GRUPO_PROGRAMAS_E_PROJETOS,
    TIPO_UNIDADE_CEI_DIRET,
    TIPOS_ALIMENTACAO,
    TIPOS_UNIDADE_ESCOLAR,
)
from src.medicao_inicial.services.relatorio_consolidado_excel import (
    _formata_filtros,
    _formata_total_geral,
    _formata_unidades_sem_lancamento,
    _preenche_linha_dos_filtros_selecionados,
    _preenche_titulo,
    gera_relatorio_consolidado_xlsx,
)

pytestmark = pytest.mark.django_db


def test_gera_relatorio_consolidado_xlsx_emef(
    relatorio_consolidado_xlsx_emef, mock_query_params_excel_emef
):
    solicitacoes = [relatorio_consolidado_xlsx_emef.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emef, contem_recreio=False
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emef.mes}-{ relatorio_consolidado_xlsx_emef.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEF.value,
        "123456",
        "EMEF TESTE",
        10,
        10,
        125,
        125,
        125,
        125,
        125,
        125,
        20,
        20,
        10,
        10,
        10,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        10,
        10,
        125,
        125,
        125,
        125,
        125,
        125,
        20,
        20,
        10,
        10,
        10,
    )


def test_gera_relatorio_consolidado_xlsx_emef_com_filtro_de_datas(
    relatorio_consolidado_xlsx_emef, mock_query_params_excel_emef
):
    solicitacoes = [relatorio_consolidado_xlsx_emef.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    query_params = {
        **mock_query_params_excel_emef,
        "data_inicial": "2025-04-03",
        "data_final": "2025-04-05",
    }

    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, query_params, contem_recreio=False
    )
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emef.mes}-{ relatorio_consolidado_xlsx_emef.ano}"
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF - 03/04/2025 A 05/04/2025",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEF.value,
        "123456",
        "EMEF TESTE",
        10,
        10,
        75,
        75,
        75,
        75,
        75,
        75,
        12,
        12,
        6,
        6,
        6,
    )


def test_gera_relatorio_consolidado_xlsx_emei(
    relatorio_consolidado_xlsx_emei, mock_query_params_excel_emei
):
    solicitacoes = [relatorio_consolidado_xlsx_emei.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEI.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emei, contem_recreio=False
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emei.mes}-{ relatorio_consolidado_xlsx_emei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEI.value,
        "987654",
        "EMEI TESTE",
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        40,
        40,
        20,
        20,
        20,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        40,
        40,
        20,
        20,
        20,
    )


def test_gera_relatorio_consolidado_xlsx_cei(
    relatorio_consolidado_xlsx_cei, mock_query_params_excel_cei
):
    solicitacoes = [relatorio_consolidado_xlsx_cei.uuid]
    tipos_unidade = [TIPO_UNIDADE_CEI_DIRET]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_cei, contem_recreio=False
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_cei.mes}-{ relatorio_consolidado_xlsx_cei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEI DIRET",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "PARCIAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        "TARDE",
        None,
        DIETA_ESPECIAL_TIPO_A,
        DIETA_ESPECIAL_TIPO_B,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "04 a 05 meses",
        "08 a 11 meses",
        "06 a 07 meses",
        "02 anos a 03 anos e 11 meses",
        "04 a 05 meses",
        "06 a 07 meses",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPO_UNIDADE_CEI_DIRET,
        "765432",
        "CEI DIRET TESTE",
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        60,
        60,
        80,
        8,
        4,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        60,
        60,
        80,
        8,
        4,
    )


def test_gera_relatorio_consolidado_xlsx_cemei(
    relatorio_consolidado_xlsx_cemei, mock_query_params_excel_cemei
):
    solicitacoes = [relatorio_consolidado_xlsx_cemei.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.CEMEI.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_cemei, contem_recreio=False
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_cemei.mes}-{ relatorio_consolidado_xlsx_cemei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE - LOTE 2 - CEMEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "PARCIAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "INFANTIL INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "INFANTIL MANHA",
        None,
        None,
        None,
        None,
        None,
        "INFANTIL TARDE",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.CEMEI.value,
        "543210",
        "CEMEI TESTE",
        5,
        5,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        30,
        30,
        15,
        15,
        15,
        1,
        1,
        1,
        1,
        1,
        1,
        1,
        1,
        1,
        1,
    )


def test_gera_relatorio_consolidado_xlsx_cemei_unifica_dieta_enteral_programas_e_projetos(
    relatorio_consolidado_xlsx_cemei,
    mock_query_params_excel_cemei,
    categoria_medicao_dieta_a_enteral_aminoacidos,
):
    medicao_programas_e_projetos = relatorio_consolidado_xlsx_cemei.medicoes.get(
        grupo__nome=GRUPO_PROGRAMAS_E_PROJETOS
    )
    baker.make(
        "ValorMedicao",
        dia="05",
        nome_campo="refeicao",
        medicao=medicao_programas_e_projetos,
        categoria_medicao=categoria_medicao_dieta_a_enteral_aminoacidos,
        valor=1,
    )

    arquivo = gera_relatorio_consolidado_xlsx(
        [relatorio_consolidado_xlsx_cemei.uuid],
        [TIPOS_UNIDADE_ESCOLAR.CEMEI.value],
        mock_query_params_excel_cemei,
        contem_recreio=False,
    )

    workbook = load_workbook(filename=BytesIO(arquivo))
    nome_aba = (
        f"Relatório Consolidado {relatorio_consolidado_xlsx_cemei.mes}-"
        f"{relatorio_consolidado_xlsx_cemei.ano}"
    )
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert not any(
        value == "DIETA ESPECIAL - TIPO A - ENTERAL / RESTRIÇÃO DE AMINOÁCIDOS"
        for value in rows[2]
    )
    assert rows[2].count(DIETA_ESPECIAL_TIPO_A) == 4


def test_gera_relatorio_consolidado_xlsx_emebs(
    relatorio_consolidado_xlsx_emebs, mock_query_params_excel_emebs
):
    solicitacoes = [relatorio_consolidado_xlsx_emebs.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEBS.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emebs, contem_recreio=False
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emebs.mes}-{ relatorio_consolidado_xlsx_emebs.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "INFANTIL (4 a 6 anos)",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "FUNDAMENTAL (acima de 6 anos)",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "TARDE",
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "TARDE",
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "NOITE",
        None,
        None,
        None,
        None,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
    )
    assert rows[4] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Lanche Emerg.",
        "Kit Lanche",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
    )
    assert rows[5] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[6] == (
        TIPOS_UNIDADE_ESCOLAR.EMEBS.value,
        "000329",
        "EMEBS TESTE",
        5,
        5,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        40,
        40,
        20,
        20,
        20,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        50,
        50,
        25,
        25,
        25,
    )
    assert rows[7] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        40,
        40,
        20,
        20,
        20,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        50,
        50,
        25,
        25,
        25,
    )


def test_gera_relatorio_consolidado_xlsx_cieja_cmct(
    relatorio_consolidado_xlsx_cieja, mock_query_params_excel_cieja_cmct
):
    solicitacoes = [relatorio_consolidado_xlsx_cieja.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.CIEJA.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_cieja_cmct,
        contem_recreio=False,
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_cieja.mes}-{ relatorio_consolidado_xlsx_cieja.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CIEJA",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "TARDE",
        None,
        None,
        None,
        None,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        DIETA_ESPECIAL_TIPO_B,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        "Total de Refeições para Pagamento",
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.LANCHE.value,
        TIPOS_ALIMENTACAO.LANCHE_4H.value,
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.CIEJA.value,
        "111329",
        "CIEJA TESTE",
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        20,
        0,
        0,
        80,
        80,
        40,
        40,
        40,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        20,
        0,
        0,
        80,
        80,
        40,
        40,
        40,
    )


def test_preenche_titulo(informacoes_excel_writer_emef):
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _preenche_titulo(workbook, worksheet, df.columns)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A1:P1"}
    assert {str(r) for r in merged_ranges} == esperados

    assert (
        sheet["A1"].value
        == "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar"
    )
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A1"].alignment.vertical == "center"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.color.rgb == "FF42474A"
    assert sheet["A1"].fill.fgColor.rgb == "FFD6F2E7"
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emef(
    mock_query_params_excel_emef, informacoes_excel_writer_emef
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_emef,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A2:P2"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emei(
    mock_query_params_excel_emei, informacoes_excel_writer_emei
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEI.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emei
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_emei,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A2:P2"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_cei(
    mock_query_params_excel_cei, informacoes_excel_writer_cei
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_cei
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_cei,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 10

    assert "A2:BE2" in str(merged_ranges)
    assert "A3:C3" in str(merged_ranges)
    assert "D3:K3" in str(merged_ranges)
    assert "L3:S3" in str(merged_ranges)
    assert "T3:AA3" in str(merged_ranges)
    assert "AB3:AI3" in str(merged_ranges)
    assert "AJ3:AQ3" in str(merged_ranges)
    assert "AR3:AY3" in str(merged_ranges)
    assert "AZ3:BA3" in str(merged_ranges)
    assert "BB3:BC3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_cemei(
    mock_query_params_excel_cemei, informacoes_excel_writer_cemei
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CEMEI.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_cemei
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_cemei,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 13

    assert "A2:BX2" in str(merged_ranges)
    assert "A3:E3" in str(merged_ranges)
    assert "F3:M3" in str(merged_ranges)
    assert "N3:U3" in str(merged_ranges)
    assert "V3:AC3" in str(merged_ranges)
    assert "AD3:AK3" in str(merged_ranges)
    assert "AL3:AS3" in str(merged_ranges)
    assert "AT3:BA3" in str(merged_ranges)
    assert "BB3:BG3" in str(merged_ranges)
    assert "BH3:BM3" in str(merged_ranges)
    assert "BN3:BS3" in str(merged_ranges)
    assert "BT3:BV3" in str(merged_ranges)
    assert "BW3:BX3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE - LOTE 2 - CEMEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emebs(
    mock_query_params_excel_emebs, informacoes_excel_writer_emebs
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEBS.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emebs
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_emebs,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 18

    assert "A2:BO2" in str(merged_ranges)
    assert "AF4:AG4" in str(merged_ranges)
    assert "AH4:AM4" in str(merged_ranges)
    assert "A4:E4" in str(merged_ranges)
    assert "BF4:BJ4" in str(merged_ranges)
    assert "BN4:BO4" in str(merged_ranges)
    assert "L4:Q4" in str(merged_ranges)
    assert "R4:W4" in str(merged_ranges)
    assert "AZ4:BE4" in str(merged_ranges)
    assert "F4:K4" in str(merged_ranges)
    assert "BK4:BM4" in str(merged_ranges)
    assert "AC4:AE4" in str(merged_ranges)
    assert "AN4:AS4" in str(merged_ranges)
    assert "AT4:AY4" in str(merged_ranges)
    assert "F3:AG3" in str(merged_ranges)
    assert "X4:AB4" in str(merged_ranges)
    assert "A3:E3" in str(merged_ranges)
    assert "AH3:BO3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_cieja_cmct(
    mock_query_params_excel_cieja_cmct, informacoes_excel_writer_cieja_cmct
):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CIEJA.value]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_cieja_cmct
    _preenche_linha_dos_filtros_selecionados(
        workbook,
        worksheet,
        mock_query_params_excel_cieja_cmct,
        df.columns,
        tipos_unidades,
        contem_recreio=False,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 7

    assert "A2:Y2" in str(merged_ranges)
    assert "R3:T3" in str(merged_ranges)
    assert "L3:Q3" in str(merged_ranges)
    assert "U3:W3" in str(merged_ranges)
    assert "X3:Y3" in str(merged_ranges)
    assert "F3:K3" in str(merged_ranges)
    assert "A3:E3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CIEJA"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_formata_total_geral(informacoes_excel_writer_emef):
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _formata_total_geral(workbook, worksheet, df)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]
    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A7:C7"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A7"].value == "TOTAL"
    assert sheet["A7"].alignment.horizontal == "center"
    assert sheet["A7"].alignment.vertical == "center"
    assert sheet["A7"].font.bold is True
    workbook_openpyxl.close()


def test_formata_filtros_unidade_emef(mock_query_params_excel_emef):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    filtros = _formata_filtros(
        mock_query_params_excel_emef, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF"


def test_formata_filtros_unidade_emei(mock_query_params_excel_emei):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEI.value]
    filtros = _formata_filtros(
        mock_query_params_excel_emei, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEI"


def test_formata_filtros_unidade_cei(mock_query_params_excel_cei):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    filtros = _formata_filtros(
        mock_query_params_excel_cei, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - CEI"


def test_formata_filtros_unidade_cemei(mock_query_params_excel_cemei):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CEMEI.value]
    filtros = _formata_filtros(
        mock_query_params_excel_cemei, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE - LOTE 2 - CEMEI"


def test_formata_filtros_unidade_emebs(mock_query_params_excel_emebs):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEBS.value]
    filtros = _formata_filtros(
        mock_query_params_excel_emebs, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS"


def test_formata_filtros_unidade_cieja_cmct(mock_query_params_excel_cieja_cmct):
    tipos_unidades = [
        TIPOS_UNIDADE_ESCOLAR.CIEJA.value,
        TIPOS_UNIDADE_ESCOLAR.CMCT.value,
    ]
    filtros = _formata_filtros(
        mock_query_params_excel_cieja_cmct, tipos_unidades, contem_recreio=False
    )
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - CIEJA, CMCT"


def test_gera_relatorio_consolidado_xlsx_tipo_unidade_invalida():
    tipos_de_unidade = ["aaa", "bbb"]
    with pytest.raises(ValueError, match=f"Unidades inválidas"):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {}, contem_recreio=False)


def test_gera_relatorio_consolidado_xlsx_retorna_exception():
    tipos_de_unidade = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    with pytest.raises(Exception):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {}, contem_recreio=False)


def test_gera_relatorio_consolidado_xlsx_recreio_emei(
    solicitacao_recreio_emei, mock_query_params_excel_recreio_emei
):
    solicitacoes = [solicitacao_recreio_emei.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEI.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_recreio_emei,
        contem_recreio=True,
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { solicitacao_recreio_emei.mes}-{ solicitacao_recreio_emei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "RECREIO NAS FÉRIAS - DEZEMBRO/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        "ALIMENTAÇÕES ALUNOS PARTICIPANTES",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        "COLABORADORES",
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEI.value,
        "987654",
        "EMEI TESTE",
        1260,
        1260,
        1260,
        1260,
        1260,
        1260,
        14,
        280,
        280,
        560,
        280,
        280,
        560,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        1260,
        1260,
        1260,
        1260,
        1260,
        1260,
        14,
        280,
        280,
        560,
        280,
        280,
        560,
    )


def test_formata_filtros_unidade_recreio_emei(mock_query_params_excel_recreio_emei):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEI.value]
    filtros = _formata_filtros(
        mock_query_params_excel_recreio_emei, tipos_unidades, contem_recreio=True
    )
    assert isinstance(filtros, str)
    assert (
        filtros
        == "RECREIO NAS FÉRIAS - Dezembro/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEI"
    )


def test_gera_relatorio_consolidado_recreio_xlsx_tipo_unidade_invalida():
    tipos_de_unidade = ["aaa", "bbb"]
    with pytest.raises(ValueError, match=f"Unidades inválidas"):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {}, contem_recreio=True)


def test_gera_relatorio_consolidado_recreio_xlsx_retorna_exception():
    tipos_de_unidade = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    with pytest.raises(Exception):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {}, contem_recreio=True)


def test_gera_relatorio_consolidado_xlsx_recreio_cei(
    solicitacao_recreio_cei, mock_query_params_excel_recreio_cei
):
    solicitacoes = [solicitacao_recreio_cei.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_recreio_cei,
        contem_recreio=True,
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { solicitacao_recreio_cei.mes}-{ solicitacao_recreio_cei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "RECREIO NAS FÉRIAS - DEZEMBRO/2025 - DIRETORIA REGIONAL TESTE -  - CEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        "ALIMENTAÇÕES ALUNOS PARTICIPANTES",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "COLABORADORES",
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPO_UNIDADE_CEI_DIRET,
        "765432",
        "CEI DIRET TESTE",
        168,
        168,
        168,
        168,
        168,
        168,
        168,
        168,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        280,
        280,
        560,
        280,
        280,
        560,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        168,
        168,
        168,
        168,
        168,
        168,
        168,
        168,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        280,
        280,
        560,
        280,
        280,
        560,
    )


def test_formata_filtros_unidade_recreio_cei(mock_query_params_excel_recreio_cei):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.CEI.value]
    filtros = _formata_filtros(
        mock_query_params_excel_recreio_cei, tipos_unidades, contem_recreio=True
    )
    assert isinstance(filtros, str)
    assert (
        filtros
        == "RECREIO NAS FÉRIAS - Dezembro/2025 - DIRETORIA REGIONAL TESTE -  - CEI"
    )


def test_gera_relatorio_consolidado_xlsx_recreio_emef(
    solicitacao_recreio_emef, mock_query_params_excel_recreio_emef
):
    solicitacoes = [solicitacao_recreio_emef.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_recreio_emef,
        contem_recreio=True,
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { solicitacao_recreio_emef.mes}-{ solicitacao_recreio_emef.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "RECREIO NAS FÉRIAS - DEZEMBRO/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        "ALIMENTAÇÕES ALUNOS PARTICIPANTES",
        None,
        None,
        None,
        None,
        None,
        DIETA_ESPECIAL_TIPO_A,
        "COLABORADORES",
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEF.value,
        "123456",
        "EMEF TESTE",
        1260,
        1260,
        1400,
        1260,
        1260,
        1400,
        14,
        280,
        280,
        560,
        280,
        280,
        560,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        1260,
        1260,
        1400,
        1260,
        1260,
        1400,
        14,
        280,
        280,
        560,
        280,
        280,
        560,
    )


def test_formata_filtros_unidade_recreio_emef(mock_query_params_excel_recreio_emef):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    filtros = _formata_filtros(
        mock_query_params_excel_recreio_emef, tipos_unidades, contem_recreio=True
    )
    assert isinstance(filtros, str)
    assert (
        filtros
        == "RECREIO NAS FÉRIAS - Dezembro/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF"
    )


def test_gera_relatorio_consolidado_xlsx_recreio_cemei(
    solicitacao_recreio_cemei,
    mock_query_params_excel_recreio_cemei,
):
    solicitacoes = [solicitacao_recreio_cemei.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.CEMEI.value]

    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_recreio_cemei,
        contem_recreio=True,
    )

    assert isinstance(arquivo, bytes)

    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)

    nome_aba = (
        f"Relatório Consolidado "
        f"{solicitacao_recreio_cemei.mes}-{solicitacao_recreio_cemei.ano}"
    )

    assert nome_aba in workbook.sheetnames

    sheet = workbook[nome_aba]

    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        *([None] * 31),
    )

    assert rows[1] == (
        "RECREIO NAS FÉRIAS - DEZEMBRO/2025 - DIRETORIA REGIONAL TESTE - LOTE 2 - CEMEI",
        *([None] * 31),
    )

    assert rows[2] == (
        None,
        None,
        None,
        "ALIMENTAÇÕES ALUNOS PARTICIPANTES",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA TIPO A",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "ALIMENTAÇÕES TURMA INFANTIL",
        None,
        None,
        None,
        None,
        None,
        "DIETA TIPO A - INFANTIL",
        "COLABORADORES",
        None,
        None,
        None,
        None,
        None,
    )

    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "0 a 1 mes",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
        TIPOS_ALIMENTACAO.REFEICAO.value,
        TIPOS_ALIMENTACAO.REFEICAO.value,
        "Repetição de Refeição",
        "Total de Refeições para Pagamento",
        TIPOS_ALIMENTACAO.SOBREMESA.value,
        "Repetição de Sobremesa",
        "Total de Sobremesas para Pagamento",
    )

    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )

    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.CEMEI.value,
        "543210",
        "CEMEI TESTE",
        154,
        154,
        154,
        154,
        154,
        154,
        154,
        154,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        1260,
        1260,
        1260,
        1260,
        1260,
        1260,
        14,
        210,
        210,
        420,
        210,
        210,
        420,
    )

    assert rows[6] == (
        "TOTAL",
        None,
        None,
        154,
        154,
        154,
        154,
        154,
        154,
        154,
        154,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        28,
        1260,
        1260,
        1260,
        1260,
        1260,
        1260,
        14,
        210,
        210,
        420,
        210,
        210,
        420,
    )


def test_formata_unidades_sem_lancamento(informacoes_excel_writer_sem_lancamentos):
    tipos_unidades = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    aba, writer, workbook, worksheet, df, arquivo = (
        informacoes_excel_writer_sem_lancamentos
    )
    _formata_unidades_sem_lancamento(workbook, worksheet, df, tipos_unidades)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 3

    esperados = {"A3:C3", "D3:E3", "D6:E6"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["D3"].value == "MANHA"
    assert sheet["D4"].value == "Total de Refeições para Pagamento"
    assert sheet["D6"].value == "UNIDADE SEM LANÇAMENTOS"
    assert sheet["E4"].value == "Total de Sobremesas para Pagamento"
    workbook_openpyxl.close()


def test_gera_relatorio_consolidado_xlsx_unidades_sem_lançamento(
    solicitacao_sem_lancamento, mock_query_params_excel_recreio_emei
):
    mock_query_params_excel_recreio_emei["mes"] = solicitacao_sem_lancamento.mes
    mock_query_params_excel_recreio_emei["ano"] = solicitacao_sem_lancamento.ano

    solicitacoes = [solicitacao_sem_lancamento.uuid]
    tipos_unidade = [TIPOS_UNIDADE_ESCOLAR.EMEF.value]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes,
        tipos_unidade,
        mock_query_params_excel_recreio_emei,
        contem_recreio=False,
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { solicitacao_sem_lancamento.mes}-{ solicitacao_sem_lancamento.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE - LOTE 1 - EMEF",
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (None, None, None, "MANHA", None)
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Total de Refeições para Pagamento",
        "Total de Sobremesas para Pagamento",
    )
    assert rows[4] == (None, None, None, None, None)
    assert rows[5] == (
        TIPOS_UNIDADE_ESCOLAR.EMEF.value,
        "123456",
        "EMEF TESTE",
        "UNIDADE SEM LANÇAMENTOS",
        None,
    )
    assert rows[6] == ("TOTAL", None, None, 0, 0)

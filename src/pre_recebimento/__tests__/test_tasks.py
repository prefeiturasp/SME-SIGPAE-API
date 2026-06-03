import io
from datetime import date
from unittest.mock import MagicMock, call, patch

import pandas as pd
import pytest
from freezegun import freeze_time
from openpyxl import load_workbook

from src.pre_recebimento.cronograma_entrega.models import (
    InterrupcaoProgramadaEntrega,
)
from src.pre_recebimento.tasks import (
    _aplicar_estilos_leve_leite,
    _criar_linha_base_excel,
    _deve_mostrar_linha_a_receber,
    _processar_fichas_recebimento,
    importa_feriados_para_interrupcoes_programadas,
    avisa_empresa_sobre_etapa_programada_proxima,
    avisa_empresa_sobre_documentos_pagamento,
)


def test_criar_linha_base_excel():
    cronograma_mock = {
        "numero": "CR-001",
        "produto": "Arroz",
        "empresa": "Empresa X",
        "marca": "Marca Y",
        "qtd_total_programada": 1000,
        "custo_unitario_produto": "5,50",
        "armazem": "Armazém Central",
        "status": "ATIVO",
    }

    etapa_mock = {
        "etapa": "1",
        "parte": "1",
        "data_programada": "15/01/2024",
        "quantidade": "500,00 kg",
        "total_embalagens": "50",
        "unidade_medida": "kg",
    }

    result = _criar_linha_base_excel(cronograma_mock, etapa_mock)

    assert result["cronograma_numero"] == "CR-001"
    assert result["etapa"] == "1"
    assert result["parte"] == "1"
    assert "quantidade" in result


def test_processar_fichas_recebimento():
    linha_base = {"etapa": "1", "parte": "1", "situacao": ""}
    etapa = {"etapa": "1", "parte": "1"}
    fichas = [
        {"situacao": "Recebido", "houve_reposicao": False},
        {"situacao": "Reposição", "houve_reposicao": True},
    ]

    result = _processar_fichas_recebimento(linha_base, etapa, fichas)

    assert len(result) == 2
    assert result[0]["situacao"] == "Recebido"
    assert result[1]["situacao"] == "Reposição"
    assert "Reposição / Pagamento" in result[1]["etapa"]


def test_deve_mostrar_linha_a_receber(parametros_deve_mostrar_linha_a_receber):
    params = parametros_deve_mostrar_linha_a_receber
    etapa = {"foi_recebida": params["foi_recebida"]}
    filtros = (
        {"situacao": params["filtros_situacao"]} if params["filtros_situacao"] else {}
    )

    result = _deve_mostrar_linha_a_receber(etapa, params["fichas_recebimento"], filtros)
    assert result == params["expected"]


def test_aplicar_estilos_leve_leite_escreve_na_linha_correta():
    worksheet = MagicMock()
    workbook = MagicMock()
    workbook.add_format.return_value = "formato-azul"

    dados = [
        {"produto_nome": "ABACAXI"},
        {"produto_nome": "BANANA"},
        {"produto_nome": "LEITE"},
    ]

    _aplicar_estilos_leve_leite(worksheet, workbook, [0, 2], dados)

    worksheet.write.assert_has_calls(
        [
            call(1, 1, "ABACAXI", "formato-azul"),
            call(3, 1, "LEITE", "formato-azul"),
        ]
    )


def test_aplicar_estilos_leve_leite_nao_sobrescreve_coluna_produto_no_xlsx():
    produtos = [
        "ABACAXI DESIDRATADO PICADO 30G",
        "DOCE DE BANANA ORGANICO SEM ACUCAR",
        "FORMULA INFANTIL DE SEGUIMENTO 6 AO 12 MES",
        "FORMULA INFANTIL DE PARTIDA 0 AO 6 MES",
        "MACA DESIDRATADA CROCANTE",
    ]

    df = pd.DataFrame(
        {
            "cronograma_numero": [
                "001/2026A",
                "002/2026A",
                "003/2026A",
                "004/2026A",
                "005/2026A",
            ],
            "produto_nome": produtos,
        }
    )

    dados = [{"produto_nome": produto} for produto in produtos]
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(
            writer,
            "Relatório de Cronogramas",
            index=False,
            header=False,
            startrow=1,
        )
        workbook = writer.book
        worksheet = writer.sheets["Relatório de Cronogramas"]
        _aplicar_estilos_leve_leite(worksheet, workbook, [2, 3], dados)

    output.seek(0)
    workbook_lido = load_workbook(output)
    worksheet_lida = workbook_lido["Relatório de Cronogramas"]

    for indice, produto_esperado in enumerate(produtos):
        # startrow=1 => primeira linha de dados no Excel e a linha 2.
        assert worksheet_lida.cell(row=indice + 2, column=2).value == produto_esperado


def test_criar_linha_base_excel():
    cronograma_mock = {
        "numero": "CR-001",
        "produto": "Arroz",
        "empresa": "Empresa X",
        "marca": "Marca Y",
        "qtd_total_programada": 1000,
        "custo_unitario_produto": "5,50",
        "armazem": "Armazém Central",
        "status": "ATIVO",
        "modalidade": "Pregão Eletrônico",
    }

    etapa_mock = {
        "etapa": "1",
        "parte": "1",
        "data_programada": "15/01/2024",
        "quantidade": "500,00 kg",
        "total_embalagens": "50",
        "unidade_medida": "kg",
    }

    result = _criar_linha_base_excel(cronograma_mock, etapa_mock)

    assert result["cronograma_numero"] == "CR-001"
    assert result["etapa"] == "1"
    assert result["parte"] == "1"
    assert "quantidade" in result
    assert result["modalidade"] == "Pregão Eletrônico"


@pytest.mark.django_db
class TestImportaFeriadosTask:

    @freeze_time("2024-05-20")
    @patch("src.pre_recebimento.tasks.BrazilSaoPauloCity")
    def test_importa_feriados_cria_registros_corretamente(self, mock_calendario_class):
        mock_calendario = MagicMock()
        mock_calendario_class.return_value = mock_calendario

        mock_calendario.holidays.side_effect = [
            [(date(2024, 1, 1), "New year")],
            [(date(2025, 1, 1), "New year")],
        ]

        importa_feriados_para_interrupcoes_programadas()

        assert InterrupcaoProgramadaEntrega.objects.count() == 4

        qs_2024 = InterrupcaoProgramadaEntrega.objects.filter(data=date(2024, 1, 1))
        assert qs_2024.count() == 2

        assert qs_2024.filter(descricao_motivo="Ano novo").count() == 2

        assert qs_2024.filter(
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL
        ).exists()
        assert qs_2024.filter(
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_PONTO_A_PONTO
        ).exists()

    @freeze_time("2024-05-20")
    @patch("src.pre_recebimento.tasks.BrazilSaoPauloCity")
    def test_importa_feriados_nao_cria_duplicados(self, mock_calendario_class):
        mock_calendario = MagicMock()
        mock_calendario_class.return_value = mock_calendario
        mock_calendario.holidays.side_effect = [
            [(date(2024, 1, 1), "New year")],
            [(date(2025, 1, 1), "New year")],
        ]

        InterrupcaoProgramadaEntrega.objects.create(
            data=date(2024, 1, 1),
            motivo=InterrupcaoProgramadaEntrega.MOTIVO_OUTROS,
            descricao_motivo="Outro motivo",
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL,
        )

        importa_feriados_para_interrupcoes_programadas()

        assert InterrupcaoProgramadaEntrega.objects.count() == 4

        existente = InterrupcaoProgramadaEntrega.objects.get(
            data=date(2024, 1, 1),
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL,
        )
        assert existente.motivo == InterrupcaoProgramadaEntrega.MOTIVO_OUTROS

        novos_feriados = InterrupcaoProgramadaEntrega.objects.filter(
            motivo=InterrupcaoProgramadaEntrega.MOTIVO_FERIADO
        )
        assert novos_feriados.count() == 3


@pytest.mark.django_db
class TestAvisaEmpresaEtapaProgramadaProxima:

    @freeze_time("2025-01-10")
    @patch("src.pre_recebimento.tasks.envia_email_em_massa_task")
    @patch("src.pre_recebimento.tasks.render_to_string", return_value="<html>teste</html>")
    @patch("src.pre_recebimento.tasks.PartesInteressadasService")
    def test_envia_email_para_cada_etapa(
        self, mock_partes_service, mock_render, mock_envia, cronograma_assinado_perfil_dilog
    ):
        from model_bakery import baker
        from src.pre_recebimento.cronograma_entrega.models import EtapasDoCronograma

        data_alvo = date(2025, 1, 13)
        data_incorreta = date(2025, 1, 14)
        etapa1 = baker.make(EtapasDoCronograma, cronograma=cronograma_assinado_perfil_dilog, data_programada=data_alvo, etapa=1)
        etapa2 = baker.make(EtapasDoCronograma, cronograma=cronograma_assinado_perfil_dilog, data_programada=data_alvo, etapa=2)

        mock_partes_service.usuarios_vinculados_a_empresa_do_objeto.return_value = ["empresa@teste.com"]

        avisa_empresa_sobre_etapa_programada_proxima()

        assert mock_envia.delay.call_count == 2

        assuntos = [c.kwargs["assunto"] for c in mock_envia.delay.call_args_list]
        assert any("Etapa 1" in a for a in assuntos)
        assert any("Etapa 2" in a for a in assuntos)
        assert all(cronograma_assinado_perfil_dilog.numero in a for a in assuntos)

    @freeze_time("2025-01-10")
    @patch("src.pre_recebimento.tasks.envia_email_em_massa_task")
    @patch("src.pre_recebimento.tasks.render_to_string", return_value="<html>teste</html>")
    @patch("src.pre_recebimento.tasks.PartesInteressadasService")
    def test_nao_envia_para_cronogramas_nao_assinados(
        self, mock_partes_service, mock_render, mock_envia
    ):
        from model_bakery import baker
        from src.pre_recebimento.cronograma_entrega.models import EtapasDoCronograma

        cronograma_rascunho = baker.make("Cronograma", status="RASCUNHO")
        baker.make(EtapasDoCronograma, cronograma=cronograma_rascunho, data_programada=date(2025, 1, 13), etapa=1)

        avisa_empresa_sobre_etapa_programada_proxima()

        mock_envia.delay.assert_not_called()

    @freeze_time("2025-01-10")
    @patch("src.pre_recebimento.tasks.envia_email_em_massa_task")
    @patch("src.pre_recebimento.tasks.render_to_string", return_value="<html>teste</html>")
    @patch("src.pre_recebimento.tasks.PartesInteressadasService")
    def test_nao_envia_para_etapas_de_outra_data(
        self, mock_partes_service, mock_render, mock_envia, cronograma_assinado_perfil_dilog
    ):
        from model_bakery import baker
        from src.pre_recebimento.cronograma_entrega.models import EtapasDoCronograma

        baker.make(EtapasDoCronograma, cronograma=cronograma_assinado_perfil_dilog, data_programada=date(2025, 1, 20), etapa=1)

        avisa_empresa_sobre_etapa_programada_proxima()

        mock_envia.delay.assert_not_called()


@pytest.mark.django_db
class TestAvisaEmpresaDocumentosPagamento:

    @freeze_time("2025-01-10")
    @patch("src.pre_recebimento.tasks.envia_email_em_massa_task")
    @patch("src.pre_recebimento.tasks.render_to_string", return_value="<html>teste</html>")
    @patch("src.pre_recebimento.tasks.PartesInteressadasService")
    def test_envia_email_para_etapas_do_dia_anterior(
        self, mock_partes_service, mock_render, mock_envia, cronograma_assinado_perfil_dilog
    ):
        from model_bakery import baker
        from src.pre_recebimento.cronograma_entrega.models import EtapasDoCronograma

        data_alvo = date(2025, 1, 9)
        baker.make(EtapasDoCronograma, cronograma=cronograma_assinado_perfil_dilog, data_programada=data_alvo, etapa=1)

        mock_partes_service.usuarios_vinculados_a_empresa_do_objeto.return_value = ["empresa@teste.com"]

        avisa_empresa_sobre_documentos_pagamento()

        assert mock_envia.delay.call_count == 1
        assunto = mock_envia.delay.call_args.kwargs["assunto"]
        assert "Documentos para Pagamento" in assunto
        assert cronograma_assinado_perfil_dilog.numero in assunto

    @freeze_time("2025-01-10")
    @patch("src.pre_recebimento.tasks.envia_email_em_massa_task")
    @patch("src.pre_recebimento.tasks.render_to_string", return_value="<html>teste</html>")
    @patch("src.pre_recebimento.tasks.PartesInteressadasService")
    def test_nao_envia_para_etapas_de_outra_data(
        self, mock_partes_service, mock_render, mock_envia, cronograma_assinado_perfil_dilog
    ):
        from model_bakery import baker
        from src.pre_recebimento.cronograma_entrega.models import EtapasDoCronograma

        baker.make(EtapasDoCronograma, cronograma=cronograma_assinado_perfil_dilog, data_programada=date(2025, 1, 5), etapa=1)

        avisa_empresa_sobre_documentos_pagamento()

        mock_envia.delay.assert_not_called()

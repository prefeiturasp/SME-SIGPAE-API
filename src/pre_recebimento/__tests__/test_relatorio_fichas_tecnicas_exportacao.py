import datetime
import re
from io import BytesIO

import openpyxl
import pytest
from model_bakery import baker
from rest_framework import status

from src.dados_comuns.constants import (
    ADMINISTRADOR_CODAE_GABINETE,
    ADMINISTRADOR_GESTAO_PRODUTO,
    COORDENADOR_GESTAO_PRODUTO,
    DJANGO_ADMIN_PASSWORD,
)
from src.dados_comuns.fluxo_status import FichaTecnicaDoProdutoWorkflow
from src.dados_comuns.models import CentralDeDownload
from src.pre_recebimento.base.models import UnidadeMedida
from src.pre_recebimento.ficha_tecnica.api.relatorio_fichas_tecnicas_excel import (
    COLUNAS,
    MENSAGEM_SEM_REGISTROS,
    TITULO_RELATORIO,
    gera_relatorio_fichas_tecnicas_xlsx,
)
from src.pre_recebimento.ficha_tecnica.models import FichaTecnicaDoProduto
from src.pre_recebimento.tasks import exporta_relatorio_fichas_tecnicas_xlsx
from src.produto.models import NomeDeProdutoEdital

URL_EXPORTAR_EXCEL = "/ficha-tecnica/exportar-excel/"
URL_LISTAGEM_RELATORIO = "/ficha-tecnica/listagem-relatorio/"
DETAIL_SUCESSO = "Solicitação de geração de arquivo recebida com sucesso."
REGEX_NOME_ARQUIVO = (
    r"Relatorio_Fichas_Tecnicas_\d{4}-\d{2}-\d{2}_\d{6}\.xlsx"
)


def _cria_usuario_com_perfil_codae(django_user_model, email, nome_perfil):
    """Cria usuário autenticado com vínculo ativo à CODAE para o perfil informado."""
    password = DJANGO_ADMIN_PASSWORD
    user = django_user_model.objects.create_user(
        username=email, password=password, email=email, registro_funcional="8888888"
    )
    perfil = baker.make("Perfil", nome=nome_perfil, ativo=True)
    codae = baker.make("Codae")
    baker.make(
        "Vinculo",
        usuario=user,
        instituicao=codae,
        perfil=perfil,
        data_inicial=datetime.date.today(),
        ativo=True,
    )
    return user, password


@pytest.fixture
def client_autenticado_coordenador_gestao_produto(client, django_user_model):
    email = "coordenador_gestao_produto@test.com"
    _, password = _cria_usuario_com_perfil_codae(
        django_user_model, email, COORDENADOR_GESTAO_PRODUTO
    )
    client.login(username=email, password=password)
    return client


@pytest.fixture
def client_autenticado_administrador_gestao_produto(client, django_user_model):
    email = "administrador_gestao_produto@test.com"
    _, password = _cria_usuario_com_perfil_codae(
        django_user_model, email, ADMINISTRADOR_GESTAO_PRODUTO
    )
    client.login(username=email, password=password)
    return client


@pytest.fixture
def client_autenticado_administrador_codae_gabinete(client, django_user_model):
    email = "administrador_codae_gabinete@test.com"
    _, password = _cria_usuario_com_perfil_codae(
        django_user_model, email, ADMINISTRADOR_CODAE_GABINETE
    )
    client.login(username=email, password=password)
    return client


@pytest.mark.django_db
def test_exportar_excel_retorna_200_e_enfileira_task(
    client_autenticado_qualidade, monkeypatch
):
    chamadas = {}

    def fake_delay(**kwargs):
        chamadas.update(kwargs)

    monkeypatch.setattr(
        "src.pre_recebimento.ficha_tecnica.api.viewsets"
        ".exporta_relatorio_fichas_tecnicas_xlsx.delay",
        fake_delay,
    )

    response = client_autenticado_qualidade.get(URL_EXPORTAR_EXCEL)

    assert response.status_code == status.HTTP_200_OK
    assert response.json() == {"detail": DETAIL_SUCESSO}
    assert chamadas["user"] == "qualidade@test.com"
    assert re.fullmatch(REGEX_NOME_ARQUIVO, chamadas["nome_arquivo"])
    assert chamadas["fichas_ids"] == []


@pytest.mark.django_db
def test_exportar_excel_aplica_filtros(
    client_autenticado_qualidade, ficha_tecnica_factory, monkeypatch
):
    ficha_leve = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_PERECIVEIS,
        programa=FichaTecnicaDoProduto.LEVE_LEITE,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )
    ficha_escolar = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_NAO_PERECIVEIS,
        programa=FichaTecnicaDoProduto.ALIMENTACAO_ESCOLAR,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )

    chamadas = {}

    def fake_delay(**kwargs):
        chamadas.update(kwargs)

    monkeypatch.setattr(
        "src.pre_recebimento.ficha_tecnica.api.viewsets"
        ".exporta_relatorio_fichas_tecnicas_xlsx.delay",
        fake_delay,
    )

    response = client_autenticado_qualidade.get(
        URL_EXPORTAR_EXCEL, {"programa": "LEVE_LEITE"}
    )

    assert response.status_code == status.HTTP_200_OK
    assert ficha_leve.id in chamadas["fichas_ids"]
    assert ficha_escolar.id not in chamadas["fichas_ids"]
    assert set(chamadas["fichas_ids"]) == {ficha_leve.id}


@pytest.mark.django_db
def test_exportar_excel_exclui_flv(
    client_autenticado_qualidade, ficha_tecnica_factory, monkeypatch
):
    ficha_flv = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_FLV,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )
    ficha_perecivel = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_PERECIVEIS,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )
    ficha_nao_perecivel = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_NAO_PERECIVEIS,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )

    chamadas = {}

    def fake_delay(**kwargs):
        chamadas.update(kwargs)

    monkeypatch.setattr(
        "src.pre_recebimento.ficha_tecnica.api.viewsets"
        ".exporta_relatorio_fichas_tecnicas_xlsx.delay",
        fake_delay,
    )

    response = client_autenticado_qualidade.get(URL_EXPORTAR_EXCEL)

    assert response.status_code == status.HTTP_200_OK
    assert ficha_flv.id not in chamadas["fichas_ids"]
    assert set(chamadas["fichas_ids"]) == {
        ficha_perecivel.id,
        ficha_nao_perecivel.id,
    }


@pytest.mark.django_db
def test_exportar_excel_sem_permissao(client_autenticado):
    response = client_autenticado.get(URL_EXPORTAR_EXCEL)

    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
def test_exportar_excel_perfil_nao_permitido_403(client_autenticado_dilog_diretoria):
    response = client_autenticado_dilog_diretoria.get(URL_EXPORTAR_EXCEL)

    assert response.status_code == status.HTTP_403_FORBIDDEN


FIXTURES_PERFIS_PERMITIDOS = [
    "client_autenticado_qualidade",
    "client_autenticado_dilog_cronograma",
    "client_autenticado_coordenador_codae_dilog",
    "client_autenticado_coordenador_gestao_produto",
    "client_autenticado_administrador_gestao_produto",
]


@pytest.mark.django_db
@pytest.mark.parametrize("client_fixture", FIXTURES_PERFIS_PERMITIDOS)
def test_exportar_excel_perfis_permitidos_200(
    request, client_fixture, monkeypatch
):
    client = request.getfixturevalue(client_fixture)
    if isinstance(client, tuple):
        client = client[0]

    chamadas = {}

    def fake_delay(**kwargs):
        chamadas.update(kwargs)

    monkeypatch.setattr(
        "src.pre_recebimento.ficha_tecnica.api.viewsets"
        ".exporta_relatorio_fichas_tecnicas_xlsx.delay",
        fake_delay,
    )

    response = client.get(URL_EXPORTAR_EXCEL)

    assert response.status_code == status.HTTP_200_OK
    assert response.json() == {"detail": DETAIL_SUCESSO}
    assert re.fullmatch(REGEX_NOME_ARQUIVO, chamadas["nome_arquivo"])
    assert chamadas["fichas_ids"] == []


@pytest.mark.django_db
@pytest.mark.parametrize("client_fixture", FIXTURES_PERFIS_PERMITIDOS)
def test_listagem_relatorio_perfis_permitidos_200(request, client_fixture):
    client = request.getfixturevalue(client_fixture)
    if isinstance(client, tuple):
        client = client[0]

    response = client.get(URL_LISTAGEM_RELATORIO)

    assert response.status_code == status.HTTP_200_OK
    assert "results" in response.json()


@pytest.mark.django_db
def test_listagem_relatorio_perfil_nao_permitido_403(
    client_autenticado_dilog_diretoria
):
    response = client_autenticado_dilog_diretoria.get(URL_LISTAGEM_RELATORIO)

    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
@pytest.mark.parametrize(
    "client_fixture",
    [
        "client_autenticado_administrador_codae_gabinete",
        "client_autenticado_dilog_abastecimento",
    ],
)
def test_listagem_relatorio_gabinete_abastecimento_403(request, client_fixture):
    client = request.getfixturevalue(client_fixture)
    if isinstance(client, tuple):
        client = client[0]

    response = client.get(URL_LISTAGEM_RELATORIO)

    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
def test_listagem_relatorio_fornecedor_403(client_autenticado_fornecedor):
    response = client_autenticado_fornecedor.get(URL_LISTAGEM_RELATORIO)

    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
def test_exporta_relatorio_fichas_tecnicas_xlsx_task(
    usuario, ficha_tecnica_factory
):
    ficha = ficha_tecnica_factory(
        categoria=FichaTecnicaDoProduto.CATEGORIA_PERECIVEIS,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )
    nome_arquivo = "Relatorio_Fichas_Tecnicas_2026-01-03_120000.xlsx"

    exporta_relatorio_fichas_tecnicas_xlsx(
        user=usuario.username,
        nome_arquivo=nome_arquivo,
        fichas_ids=[ficha.id],
    )

    assert CentralDeDownload.objects.count() == 1
    arquivo = CentralDeDownload.objects.first()
    assert arquivo.identificador == nome_arquivo
    assert arquivo.status == CentralDeDownload.STATUS_CONCLUIDO
    assert arquivo.arquivo
    assert arquivo.arquivo.size > 0


@pytest.mark.django_db
def test_gera_relatorio_fichas_tecnicas_xlsx_conteudo(
    empresa, ficha_tecnica_factory
):
    unidade_primaria = baker.make(
        UnidadeMedida, nome="Quilograma", abreviacao="kg"
    )
    unidade_secundaria = baker.make(
        UnidadeMedida, nome="Litro", abreviacao="l"
    )
    fabricante = baker.make("Fabricante", nome="FABRICANTE A")
    envasador = baker.make("Fabricante", nome="ENVASADOR B")
    fabricante_fft = baker.make(
        "FabricanteFichaTecnica", fabricante=fabricante
    )
    envasador_fft = baker.make(
        "FabricanteFichaTecnica", fabricante=envasador
    )
    produto = baker.make(
        "NomeDeProdutoEdital",
        nome="ARROZ",
        tipo_produto=NomeDeProdutoEdital.LOGISTICA,
    )

    ficha = ficha_tecnica_factory(
        produto=produto,
        empresa=empresa,
        categoria=FichaTecnicaDoProduto.CATEGORIA_PERECIVEIS,
        programa=FichaTecnicaDoProduto.LEVE_LEITE,
        pregao_chamada_publica="PE-2025-001",
        fabricante=fabricante_fft,
        envasador_distribuidor=envasador_fft,
        prazo_validade="30 dias",
        organico=True,
        mecanismo_controle=FichaTecnicaDoProduto.MECANISMO_CERTIFICACAO,
        alergenicos=False,
        gluten=None,
        lactose=True,
        produto_eh_liquido=False,
        peso_liquido_embalagem_primaria=1.5,
        unidade_medida_primaria=unidade_primaria,
        peso_liquido_embalagem_secundaria=20.0,
        unidade_medida_secundaria=unidade_secundaria,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )

    arquivo = gera_relatorio_fichas_tecnicas_xlsx([ficha.id])
    assert isinstance(arquivo, bytes)

    workbook = openpyxl.load_workbook(filename=BytesIO(arquivo))
    sheet = workbook["Fichas Técnicas"]

    # Linha 0 (row 1): título
    assert sheet.cell(row=1, column=1).value == TITULO_RELATORIO
    assert sheet.cell(row=1, column=1).fill.start_color.rgb.endswith(
        "A9D18E"
    )

    # Linha 1 (row 2): subtítulo com data de extração
    assert sheet.cell(row=2, column=1).value.startswith("Data de extração: ")

    # Linha 2 (row 3): cabeçalho com as 18 colunas na ordem exata
    assert tuple(
        sheet.iter_rows(min_row=3, max_row=3, values_only=True)
    )[0] == tuple(COLUNAS)

    # Linha 3 (row 4): primeira ficha
    linha_dados = [sheet.cell(row=4, column=col).value for col in range(1, 19)]
    assert linha_dados[0] == "ARROZ"
    assert linha_dados[1] == "Alimentos SA"
    assert linha_dados[2] == "Perecíveis"
    assert linha_dados[3] == "Leve Leite"
    assert linha_dados[4] == "PE-2025-001"
    assert linha_dados[5] == "FABRICANTE A / ENVASADOR B"
    assert linha_dados[6] == "30 dias"
    assert linha_dados[7] == "Sim"
    assert linha_dados[8] == "Certificação"
    assert linha_dados[9] == "Não"
    assert linha_dados[10] == "Não"  # gluten=None
    assert linha_dados[11] == "Sim"
    assert linha_dados[12] == "Não"
    assert linha_dados[13] == 1.5
    assert linha_dados[14] == "kg"
    assert linha_dados[15] == 20.0
    assert linha_dados[16] == "l"
    assert linha_dados[17] == "Aprovada"

    # Células de peso (colunas 14 e 16): number_format + valores numéricos
    assert sheet.cell(row=4, column=14).number_format == "#,##0.00"
    assert sheet.cell(row=4, column=16).number_format == "#,##0.00"


@pytest.mark.django_db
def test_gera_relatorio_fichas_tecnicas_xlsx_sem_dados():
    arquivo = gera_relatorio_fichas_tecnicas_xlsx([])
    assert isinstance(arquivo, bytes)

    workbook = openpyxl.load_workbook(filename=BytesIO(arquivo))
    sheet = workbook["Fichas Técnicas"]

    # Cabeçalho presente + linha única "Nenhum registro encontrado"
    assert tuple(
        sheet.iter_rows(min_row=3, max_row=3, values_only=True)
    )[0] == tuple(COLUNAS)
    assert sheet.cell(row=4, column=1).value == MENSAGEM_SEM_REGISTROS


@pytest.mark.django_db
def test_gera_relatorio_fichas_tecnicas_xlsx_sanitiza_injecao_formula(
    ficha_tecnica_factory,
):
    produto = baker.make(
        "NomeDeProdutoEdital",
        nome="=1+1",
        tipo_produto=NomeDeProdutoEdital.LOGISTICA,
    )
    empresa = baker.make(
        "Terceirizada",
        nome_fantasia="@cmd",
        razao_social="Empresa Teste",
        tipo_servico="FORNECEDOR",
    )

    ficha = ficha_tecnica_factory(
        produto=produto,
        empresa=empresa,
        categoria=FichaTecnicaDoProduto.CATEGORIA_PERECIVEIS,
        status=FichaTecnicaDoProdutoWorkflow.APROVADA,
    )

    arquivo = gera_relatorio_fichas_tecnicas_xlsx([ficha.id])
    assert isinstance(arquivo, bytes)

    workbook = openpyxl.load_workbook(filename=BytesIO(arquivo))
    sheet = workbook["Fichas Técnicas"]

    # Linha 3 (row 4): primeira ficha; colunas 1 (produto) e 2 (empresa)
    # sanitizadas contra injeção de fórmula.
    assert sheet.cell(row=4, column=1).value == "'=1+1"
    assert sheet.cell(row=4, column=2).value == "'@cmd"
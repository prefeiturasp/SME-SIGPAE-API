import pytest
from model_bakery import baker
from rest_framework import status

from src.dados_comuns.fluxo_status import DocumentoDeRecebimentoWorkflow
from src.pre_recebimento.documento_recebimento.api.serializers.serializers import (
    CronogramaRelatorioDocumentosSerializer,
    DocumentoDeRecebimentoDetalheRelatorioSerializer,
)

pytestmark = pytest.mark.django_db


class TestDocumentoDeRecebimentoDetalheRelatorioSerializer:
    def test_campos_retornados(
        self, documento_recebimento_leve_leite, laboratorio, unidade_medida_logistica
    ):
        doc = documento_recebimento_leve_leite
        doc.laboratorio = laboratorio
        doc.unidade_medida = unidade_medida_logistica
        doc.save()

        data = DocumentoDeRecebimentoDetalheRelatorioSerializer(doc).data

        assert "uuid" in data
        assert "numero_laudo" in data
        assert "nome_laboratorio" in data
        assert "numero_lote_laudo" in data
        assert "data_final_lote" in data
        assert "unidade_medida" in data
        assert "quantidade_laudo" in data
        assert "saldo_inicial_laudo" in data
        assert "saldo_atual" in data
        assert "datas_fabricacao_e_prazos" in data
        assert "status_documento" in data

    def test_valores_do_documento(self, documento_recebimento_leve_leite, laboratorio):
        doc = documento_recebimento_leve_leite
        doc.quantidade_laudo = 500
        doc.laboratorio = laboratorio
        doc.save()

        data = DocumentoDeRecebimentoDetalheRelatorioSerializer(doc).data
        assert str(data["saldo_inicial_laudo"]) == str(data["quantidade_laudo"])
        assert data["nome_laboratorio"] == laboratorio.nome
        assert data["status_documento"] == doc.get_status_display()


class TestCronogramaRelatorioDocumentosSerializer:
    def test_campos_retornados(
        self, cronograma_leve_leite, documento_recebimento_leve_leite
    ):
        data = CronogramaRelatorioDocumentosSerializer(cronograma_leve_leite).data

        assert "uuid" in data
        assert "numero_cronograma" in data
        assert "produto" in data
        assert "empresa" in data
        assert "numero_pregao_chamada_publica" in data
        assert "numero_processo_sei" in data
        assert "documentos" in data

    def test_documentos_aninhados(
        self, cronograma_leve_leite, documento_recebimento_leve_leite
    ):
        data = CronogramaRelatorioDocumentosSerializer(cronograma_leve_leite).data

        assert len(data["documentos"]) >= 1
        for doc in data["documentos"]:
            assert "numero_laudo" in doc
            assert "status_documento" in doc

    def test_dados_do_cronograma(
        self, cronograma_leve_leite, documento_recebimento_leve_leite
    ):
        data = CronogramaRelatorioDocumentosSerializer(cronograma_leve_leite).data

        assert data["numero_cronograma"] == cronograma_leve_leite.numero
        assert data["produto"] == cronograma_leve_leite.ficha_tecnica.produto.nome
        assert data["empresa"] == cronograma_leve_leite.empresa.razao_social
        assert data["numero_processo_sei"] == cronograma_leve_leite.contrato.processo


class TestListaRelatorioAction:
    url = "/documentos-de-recebimento/listagem-relatorio/"

    def test_retorna_somente_cronogramas_com_documentos(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
        cronograma,
    ):
        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK
        uuids = [item["uuid"] for item in response.data["results"]]
        assert str(cronograma_leve_leite.uuid) in uuids
        assert str(cronograma.uuid) not in uuids

    def test_sem_duplicatas(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        baker.make(
            "DocumentoDeRecebimento",
            cronograma=cronograma_leve_leite,
            status=DocumentoDeRecebimentoWorkflow.APROVADO,
        )

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK
        uuids = [item["uuid"] for item in response.data["results"]]
        assert len(uuids) == len(set(uuids))

    def test_totalizadores_presentes(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        documento_recebimento_leve_leite,
    ):
        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK
        totalizadores = response.data["totalizadores"]
        assert "Total de Documentos Recebidos" in totalizadores
        assert "Total de Pendentes de Aprovação" in totalizadores
        assert "Total de Enviados para Correção" in totalizadores
        assert "Total de Aprovados" in totalizadores

    def test_totalizadores_valores(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        baker.make(
            "DocumentoDeRecebimento",
            cronograma=cronograma_leve_leite,
            status=DocumentoDeRecebimentoWorkflow.APROVADO,
        )

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        totalizadores = response.data["totalizadores"]
        assert totalizadores["Total de Documentos Recebidos"] == 2
        assert totalizadores["Total de Aprovados"] == 1
        assert totalizadores["Total de Pendentes de Aprovação"] == 1

    def test_filtro_por_status_aprovado(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        baker.make(
            "DocumentoDeRecebimento",
            cronograma=cronograma_leve_leite,
            status=DocumentoDeRecebimentoWorkflow.APROVADO,
        )

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(
            self.url, {"status_documento": DocumentoDeRecebimentoWorkflow.APROVADO}
        )

        assert response.status_code == status.HTTP_200_OK
        for item in response.data["results"]:
            for doc in item["documentos"]:
                assert doc["status_documento"] == "Aprovado"

    def test_filtro_status_exclui_cronogramas_sem_documentos_correspondentes(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        documento_recebimento_leve_leite,
        cronograma,
    ):
        baker.make(
            "DocumentoDeRecebimento",
            cronograma=cronograma,
            status=DocumentoDeRecebimentoWorkflow.ENVIADO_PARA_ANALISE,
        )

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(
            self.url, {"status_documento": DocumentoDeRecebimentoWorkflow.APROVADO}
        )

        assert response.status_code == status.HTTP_200_OK
        uuids = [item["uuid"] for item in response.data["results"]]
        assert str(cronograma.uuid) not in uuids

    def test_sem_permissao_fornecedor_retorna_403(
        self,
        client_user_autenticado_fornecedor,
        documento_recebimento_leve_leite,
    ):
        client, _ = client_user_autenticado_fornecedor
        response = client.get(self.url)

        assert response.status_code == status.HTTP_403_FORBIDDEN


class TestExportarExcelAction:
    url = "/documentos-de-recebimento/exportar-excel/"

    def test_export_retorna_200_e_content_type_correto(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "Content-Disposition" in response
        assert "relatorio_documentos_recebimento" in response["Content-Disposition"]

    def test_export_com_filtro_status_aprovado(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        baker.make(
            "DocumentoDeRecebimento",
            cronograma=cronograma_leve_leite,
            status=DocumentoDeRecebimentoWorkflow.APROVADO,
        )

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(
            self.url, {"status_documento": DocumentoDeRecebimentoWorkflow.APROVADO}
        )

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )

    def test_export_sem_permissao_fornecedor_retorna_403(
        self,
        client_user_autenticado_fornecedor,
        documento_recebimento_leve_leite,
    ):
        client, _ = client_user_autenticado_fornecedor
        response = client.get(self.url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_arquivo_com_colunas_corretas(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        import openpyxl
        import io

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        # Linha 2 (0-indexed) = cabeçalho das colunas
        colunas_esperadas = [
            "Nº do Cronograma",
            "Produto",
            "Empresa",
            "Nº do Pregão ou Chamada Pública",
            "Nº do Processo SEI",
            "Nº do Laudo",
            "Status",
            "Nome do Laboratório",
            "Data de Fabricação",
            "Data de Validade",
            "Nº do(s) Lote(s) do Laudo",
            "Data de Fabricação Até",
            "Data Máxima de Recebimento",
            "Data de Conclusão do Laudo",
            "Unidade",
            "Quantidade Total do Laudo",
            "Saldo Inicial do Laudo",
            "Saldo Atual",
            "Obs",
            "Solicitação de Correção",
        ]

        for col_idx, coluna_esperada in enumerate(colunas_esperadas):
            assert (
                sheet.cell(row=3, column=col_idx + 1).value == coluna_esperada
            ), f"Coluna {col_idx + 1}: esperado '{coluna_esperada}', obtido '{sheet.cell(row=3, column=col_idx + 1).value}'"

    def test_export_arquivo_com_totalizadores(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
    ):
        import openpyxl
        import io

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        # Linha 2 (1-indexed) = subtítulo com totalizadores
        subtitulo = sheet.cell(row=2, column=1).value or ""
        assert "Total de Documentos Recebidos" in subtitulo
        assert "Pendentes de Aprovação" in subtitulo
        assert "Enviados para Correção" in subtitulo
        assert "Aprovados" in subtitulo
        assert "Data/Hora da extração" in subtitulo

    def test_export_arquivo_com_dados_dos_documentos(
        self,
        client_autenticado_vinculo_dilog_qualidade,
        cronograma_leve_leite,
        documento_recebimento_leve_leite,
        laboratorio,
        unidade_medida_logistica,
    ):
        import openpyxl
        import io

        doc = documento_recebimento_leve_leite
        doc.laboratorio = laboratorio
        doc.unidade_medida = unidade_medida_logistica
        doc.quantidade_laudo = 500
        doc.save()

        client, _ = client_autenticado_vinculo_dilog_qualidade
        response = client.get(self.url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        # Linha 4 (1-indexed) = primeira linha de dados
        numero_cronograma = sheet.cell(row=4, column=1).value
        assert numero_cronograma == cronograma_leve_leite.numero

        produto = sheet.cell(row=4, column=2).value
        assert produto == cronograma_leve_leite.ficha_tecnica.produto.nome

        numero_laudo = sheet.cell(row=4, column=6).value
        assert numero_laudo == doc.numero_laudo

        laboratorio_nome = sheet.cell(row=4, column=8).value
        assert laboratorio_nome == laboratorio.nome

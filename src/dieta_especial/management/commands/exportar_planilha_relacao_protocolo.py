import re
from difflib import SequenceMatcher

from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font

from src.dados_comuns.constants import (
    CHAVE_COD_EOL_ALUNO,
    CHAVE_COD_EOL_ESCOLA,
    CHAVE_LOTE,
    CHAVE_NOME_ALUNO,
    CHAVE_NOME_ESCOLA,
    CHAVE_NOME_PROTOCOLO_DB,
    CHAVE_NOME_PROTOCOLO_IMPORTADO_DIETA,
    CHAVE_UUID,
)
from src.dieta_especial.protocolo_padrao.models import (
    ProtocoloPadraoDietaEspecial,
)
from src.dieta_especial.solicitacao_dieta_especial.models import (
    SolicitacaoDietaEspecial,
)
from src.terceirizada.models import Edital


class Command(BaseCommand):
    help = """
    Verifica se os protocolos padrões da planilha de importação estão cadastrados no sistema
    para determinado Edital e exporta uma planilha com o resultado
    """

    def formatar_tamanho_celulas(self, ws):
        for column in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[column].width = 25 if column in ["C", "E", "F"] else 50

    def exportar_planilha(self, cabecalho, dicts_para_planilha, nome):
        wb = Workbook()
        ws = wb.active
        self.formatar_tamanho_celulas(ws)
        ws.title = "Relação Protocolos"

        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = str(title)
            celula.font = Font(size="13", bold=True)

        for ind, dict_solicitacao in enumerate(dicts_para_planilha, 2):
            ws.cell(row=ind, column=1, value=str(dict_solicitacao[CHAVE_UUID]))
            ws.cell(row=ind, column=2, value=dict_solicitacao[CHAVE_NOME_ESCOLA])
            ws.cell(row=ind, column=3, value=dict_solicitacao[CHAVE_COD_EOL_ESCOLA])
            ws.cell(row=ind, column=4, value=dict_solicitacao[CHAVE_NOME_ALUNO])
            ws.cell(row=ind, column=5, value=dict_solicitacao[CHAVE_COD_EOL_ALUNO])
            ws.cell(row=ind, column=6, value=dict_solicitacao[CHAVE_LOTE])
            ws.cell(
                row=ind,
                column=7,
                value=dict_solicitacao[CHAVE_NOME_PROTOCOLO_IMPORTADO_DIETA],
            )
            ws.cell(row=ind, column=8, value=dict_solicitacao[CHAVE_NOME_PROTOCOLO_DB])

        wb.save(f"relacao-protocolos-{nome}.xlsx")

    def fix_nome_protocolo(self, queryset):
        for instance in queryset.filter(nome_protocolo__icontains="–"):
            instance.nome_protocolo = instance.nome_protocolo.replace("–", "-")
            instance.save()

    def similar(self, a, b):
        return SequenceMatcher(None, a, b).ratio()

    def compara_nome_protocolo(
        self, nome_protocolo, nomes_protocolos, dict_solicitacao
    ):
        resultado = False
        nomes_protocolos_formatados = [
            re.sub("[^A-Za-z]+", "", nome.upper()) for nome in nomes_protocolos
        ]
        nome_protocolo = re.sub("[^A-Za-z]+", "", nome_protocolo.upper())
        for ind, nome in enumerate(nomes_protocolos_formatados):
            porcentagem_similar = self.similar(nome, nome_protocolo)
            if (
                float(porcentagem_similar) > 0.95
            ):  # verifica se a similaridade é maior que 95%
                dict_solicitacao[CHAVE_NOME_PROTOCOLO_DB] = nomes_protocolos[ind]
                resultado = True
                break
        return resultado

    def handle(self, *args, **options):
        protocolos = ProtocoloPadraoDietaEspecial.objects.all()
        solicitacoes = SolicitacaoDietaEspecial.objects.filter(eh_importado=True)
        self.fix_nome_protocolo(protocolos)
        self.fix_nome_protocolo(solicitacoes)
        dicts_protocolo_invalidos = []
        dicts_protocolo_validos = []
        for solicitacao in solicitacoes:
            nome_protocolo = solicitacao.nome_protocolo
            editais_uuids = solicitacao.escola.lote.contratos_do_lote.values_list(
                "edital__uuid", flat=True
            )
            editais = Edital.objects.filter(uuid__in=editais_uuids)
            nomes_protocolos = editais.values_list(
                "protocolos_padroes_dieta_especial__nome_protocolo", flat=True
            )
            dict_solicitacao = {
                CHAVE_UUID: solicitacao.uuid,
                CHAVE_NOME_ESCOLA: solicitacao.escola.nome,
                CHAVE_COD_EOL_ESCOLA: solicitacao.escola.codigo_eol,
                CHAVE_NOME_ALUNO: solicitacao.aluno.nome,
                CHAVE_COD_EOL_ALUNO: solicitacao.aluno.codigo_eol,
                CHAVE_LOTE: solicitacao.escola.lote.nome,
                CHAVE_NOME_PROTOCOLO_IMPORTADO_DIETA: nome_protocolo,
                CHAVE_NOME_PROTOCOLO_DB: "",
            }
            if self.compara_nome_protocolo(
                nome_protocolo, nomes_protocolos, dict_solicitacao
            ):
                dicts_protocolo_validos.append(dict_solicitacao)
            else:
                dicts_protocolo_invalidos.append(dict_solicitacao)
        cabecalho = [
            CHAVE_UUID,
            CHAVE_NOME_ESCOLA,
            CHAVE_COD_EOL_ESCOLA,
            CHAVE_NOME_ALUNO,
            CHAVE_COD_EOL_ALUNO,
            CHAVE_LOTE,
            CHAVE_NOME_PROTOCOLO_IMPORTADO_DIETA,
            CHAVE_NOME_PROTOCOLO_DB,
        ]
        self.exportar_planilha(cabecalho, dicts_protocolo_validos, "validos")
        self.exportar_planilha(cabecalho, dicts_protocolo_invalidos, "invalidos")

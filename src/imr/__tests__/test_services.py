import io
import os
import tempfile
from datetime import date
from types import SimpleNamespace

import openpyxl
import pytest
from django.template.loader import render_to_string
from pypdf import PdfReader
from rest_framework import status

from src.imr.api.services import RelatorioNotificacaoService
from src.relatorios.relatorios import exportar_relatorio_notificacao

pytestmark = pytest.mark.django_db


def test_modelo_excel_tipos_penalidade(client_admin_django):
    response = client_admin_django.get(
        "/admin/imr/importacaoplanilhatipopenalidade/exportar_planilha_importacao_tipos_penalidade",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        follow=True,
    )
    assert response.status_code == status.HTTP_200_OK

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_file.write(response.content)
        temp_file_path = temp_file.name

    assert os.path.exists(temp_file_path), "Falha ao criar o arquivo"

    workbook = openpyxl.load_workbook(temp_file_path)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert first_row == (
        "Edital",
        "Número da Cláusula/Item",
        "Gravidade",
        "Obrigações (separadas por ;)",
        "Descrição da Cláusula/Item",
        "Status",
    )

    os.remove(temp_file_path)


def test_modelo_excel_tipos_ocorrencia(client_admin_django):
    response = client_admin_django.get(
        "/admin/imr/importacaoplanilhatipoocorrencia/exportar_planilha_importacao_tipos_ocorrencia",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        follow=True,
    )
    assert response.status_code == status.HTTP_200_OK

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_file.write(response.content)
        temp_file_path = temp_file.name

    assert os.path.exists(temp_file_path), "Falha ao criar o arquivo"

    workbook = openpyxl.load_workbook(temp_file_path)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert first_row == (
        "Posição",
        "Perfis",
        "Edital",
        "Categoria da Ocorrência",
        "Título",
        "Descrição",
        "Penalidade",
        "É IMR?",
        "Pontuação (IMR)",
        "Tolerância",
        "% de Desconto",
        "Status",
        "Aceita múltiplas respostas?",
    )

    os.remove(temp_file_path)


def test_formata_quantidades_da_notificacao_como_inteiros(
    categoria_ocorrencia_factory,
    formulario_supervisao_factory,
    parametrizacao_ocorrencia_factory,
    resposta_campo_numerico_factory,
    resposta_campo_texto_simples_factory,
    tipo_ocorrencia_factory,
    tipo_pergunta_parametrizacao_ocorrencia_factory,
    tipo_resposta_modelo_factory,
):
    formulario = formulario_supervisao_factory.create()
    categoria = categoria_ocorrencia_factory.create(
        nome=RelatorioNotificacaoService.CATEGORIA_QUANTIDADE_QUALIDADE
    )
    tipo_ocorrencia = tipo_ocorrencia_factory.create(categoria=categoria)
    tipo_resposta_texto = tipo_resposta_modelo_factory.create(
        nome="RespostaCampoTextoSimples"
    )
    tipo_pergunta_texto = tipo_pergunta_parametrizacao_ocorrencia_factory.create(
        tipo_resposta=tipo_resposta_texto
    )
    tipo_resposta_numerico = tipo_resposta_modelo_factory.create(
        nome="RespostaCampoNumerico"
    )
    tipo_pergunta_numerica = tipo_pergunta_parametrizacao_ocorrencia_factory.create(
        tipo_resposta=tipo_resposta_numerico
    )
    parametrizacoes = [
        parametrizacao_ocorrencia_factory.create(
            tipo_ocorrencia=tipo_ocorrencia,
            tipo_pergunta=tipo_pergunta_texto,
            posicao=posicao,
        )
        for posicao in range(1, 3)
    ]
    parametrizacoes += [
        parametrizacao_ocorrencia_factory.create(
            tipo_ocorrencia=tipo_ocorrencia,
            tipo_pergunta=tipo_pergunta_numerica,
            posicao=posicao,
        )
        for posicao in range(3, 6)
    ]

    for grupo in range(1, 3):
        resposta_campo_texto_simples_factory.create(
            formulario_base=formulario.formulario_base,
            parametrizacao=parametrizacoes[0],
            grupo=grupo,
            resposta=f"Equipamento {grupo}",
        )
        resposta_campo_texto_simples_factory.create(
            formulario_base=formulario.formulario_base,
            parametrizacao=parametrizacoes[1],
            grupo=grupo,
            resposta=f"Motivo {grupo}",
        )
        for indice, quantidade in enumerate((10.0, 20.0, 30.0), start=2):
            resposta_campo_numerico_factory.create(
                formulario_base=formulario.formulario_base,
                parametrizacao=parametrizacoes[indice],
                grupo=grupo,
                resposta=quantidade,
            )

    respostas = RelatorioNotificacaoService(formulario).formatar_respostas(
        categoria.nome
    )

    assert len(respostas) == 2
    assert respostas[0][0].resposta == "Equipamento 1"
    assert respostas[0][1].resposta == "Motivo 1"
    assert respostas[0][2:] == [10, 20, 30]
    assert respostas[1][0].resposta == "Equipamento 2"
    assert respostas[1][1].resposta == "Motivo 2"
    assert respostas[1][2:] == [10, 20, 30]


def test_retorna_dados_do_cabecalho_sem_horario(
    contrato_factory,
    escola_factory,
    formulario_supervisao_factory,
    lote_factory,
):
    lote = lote_factory.create(nome="LOTE 01")
    escola = escola_factory.create(
        codigo_eol="019432",
        lote=lote,
        nome="CEMEI PARQUE DO LAGO",
    )
    contrato = contrato_factory.create(numero="EDITAL IMR")
    contrato.lotes.add(lote)
    formulario = formulario_supervisao_factory.create(
        escola=escola,
        formulario_base__data=date(2026, 8, 5),
    )

    dados = RelatorioNotificacaoService(formulario).retornar_dados_formatados(
        RelatorioNotificacaoService.CATEGORIA_QUANTIDADE_QUALIDADE
    )

    assert dados["unidade"] == "019432 - CEMEI PARQUE DO LAGO"
    assert dados["data_visita"] == "05/08/2026"
    assert dados["edital"] == "Edital nº EDITAL IMR"
    assert "hora_geracao" not in dados


@pytest.fixture
def dados_relatorio_notificacao():
    return {
        "data_geracao": "27/08/2026",
        "data_visita": "05/08/2026",
        "diretoria_regional": "DRE TESTE",
        "edital": "Edital nº EDITAL IMR",
        "lote": "LOTE 01",
        "maior_frequencia_no_periodo": 100,
        "respostas": [["Equipamento", "Motivo", 10, 20, 30]],
        "terceirizada": "EMPRESA TESTE",
        "total_matriculados_por_data": 120,
        "unidade": "019432 - CEMEI PARQUE DO LAGO",
        "usuario": SimpleNamespace(
            nome="NUTRICIONISTA SUPERVISORA",
            registro_funcional="1234567",
        ),
    }


def test_renderiza_conteudo_da_notificacao_quantidade_qualidade(
    dados_relatorio_notificacao,
):
    html = render_to_string(
        "IMR/relatorio_de_notificacao/quantidade_qualidade/pdf.html",
        {"dados": dados_relatorio_notificacao},
    )

    assert "logo-color.svg" in html
    assert (
        "NOTIFICAÇÃO PARA REPOSIÇÃO DE UTENSÍLIOS DE MESA, MOBILIÁRIOS E "
        "EQUIPAMENTOS"
    ) in html
    assert "Data da visita:" in html
    assert "05/08/2026" in html
    assert 'class="prazo-atendimento"' in html
    assert "A empresa deve atender a solicitação no prazo de até" in html
    assert "<u>TRÊS</u>" in html
    assert "Quantidade em boas condições para uso na unidade" in html
    assert "Quantidade para reposição" in html
    assert "Quantidade para retirada" in html
    assert "Data da conferência:" in html
    assert "Responsável pela U.E." in html
    assert "Entrega da presente notificação à Empresa:" in html
    assert "Assinatura Nutricionista RT da Empresa" in html
    assert "Documento gerado eletronicamente em 27/08/2026" in html
    assert "em 27/08/2026." in html
    assert "27/08/2026, às" not in html
    assert "10.0" not in html
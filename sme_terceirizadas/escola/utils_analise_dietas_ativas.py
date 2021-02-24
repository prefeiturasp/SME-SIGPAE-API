import json
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from utility.carga_dados.helper import excel_to_list_with_openpyxl, progressbar

from sme_terceirizadas.escola.models import Escola
from sme_terceirizadas.produto.models import ProtocoloDeDietaEspecial

DATA = date.today().isoformat().replace('-', '_')
home = str(Path.home())
dict_codigos_escolas = {}
dict_codigo_aluno_por_codigo_escola = {}


def get_codigo_eol_escola(valor):
    return valor.strip().zfill(6)


def get_codigo_eol_aluno(valor):
    return str(valor).strip().zfill(7)


def gera_dict_codigos_escolas(items_codigos_escolas):
    for item in items_codigos_escolas:
        dict_codigos_escolas[str(item['CÓDIGO UNIDADE'])] = str(item['CODIGO EOL'])


def gera_dict_codigo_aluno_por_codigo_escola(items):
    for item in items:
        try:
            codigo_eol_escola = dict_codigos_escolas[item['CodEscola']]
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            with open(f'{home}/codescola_nao_existentes.txt', 'a') as f:
                f.write(f"{item['CodEscola']}\n")
            raise e

        cod_eol_aluno = get_codigo_eol_aluno(item['CodEOLAluno'])
        # chave: cod_eol_aluno, valor: codigo_eol_escola
        dict_codigo_aluno_por_codigo_escola[cod_eol_aluno] = get_codigo_eol_escola(codigo_eol_escola)


def get_escolas_unicas(items):
    """A partir da planilha, pegar todas as escolas únicas "escolas_da_planilha".

    Retorna escolas únicas.
    """
    escolas = []
    for item in items:
        escolas.append(item['CodEscola'])
    return set(escolas)


def escreve_xlsx(codigo_eol_escola_nao_existentes):
    wb = openpyxl.Workbook()
    with NamedTemporaryFile() as tmp:
        ws = wb.create_sheet('Código EOL das Escolas não identificadas no SIGPAE')
        ws['A1'] = 'codigo_eol_escola'
        for i, item in enumerate(progressbar(list(codigo_eol_escola_nao_existentes), 'Escrevendo...')):
            ws[f'A{i+2}'] = str(item)
        nome_excel_tempfile = f'{tmp.name}.xlsx'
        wb.save(nome_excel_tempfile)
        output = BytesIO(save_virtual_workbook(wb))
    return output, nome_excel_tempfile


def escreve_xlsx_primeira_aba(resultado, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.worksheets[0]
    ws['A1'] = 'Este arquivo contém as planilhas:'
    ws['A2'] = 'Código EOL das Escolas não identificadas no SIGPAE'
    ws['A3'] = 'Código EOL dos Alunos não matriculados na escola'
    ws['A4'] = 'CodEscola não existentes em unidades_da_rede...'
    ws['A5'] = 'Alunos com nome diferente do EOL'
    ws['A6'] = 'Alunos com data nascimento diferente do EOL'
    ws['A7'] = 'Dados do SIGPAE para as escolas da planilha'
    ws['A8'] = 'CodDiagnostico inexistentes'
    ws['A9'] = 'ProtocoloDieta inexistentes'
    wb.save(arquivo_saida)
    return resultado, arquivo_saida


def escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('Código EOL dos Alunos não matriculados na escola')
    ws['A1'] = 'codigo_eol_aluno'
    ws['B1'] = 'nome_aluno'
    ws['C1'] = 'codigo_eol_escola'
    for i, item in enumerate(progressbar(list(alunos_nao_matriculados_na_escola_lista), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
        ws[f'C{i+2}'] = str(item[2])
    wb.save(arquivo_saida)
    return arquivo_saida


def escreve_xlsx_codescola_nao_existentes(codescola_nao_existentes, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('CodEscola não existentes em unidades_da_rede...')
    ws['A1'] = 'CodEscola'
    for i, item in enumerate(progressbar(list(codescola_nao_existentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    wb.save(arquivo_saida)


def retorna_codigo_eol_escolas_nao_identificadas(items):  # noqa C901
    aux = []
    codescola_nao_existentes = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            codescola_nao_existentes.append(item['CodEscola'])
            raise e

    cod_escola_unicos = set(aux)
    codigo_eol_sigpae_lista = Escola.objects.values_list('codigo_eol', flat=True)
    codigo_eol_escola_nao_existentes = cod_escola_unicos - set(codigo_eol_sigpae_lista)

    arquivo_saida, nome_excel_tempfile = escreve_xlsx(codigo_eol_escola_nao_existentes)

    if set(codescola_nao_existentes):
        escreve_xlsx_codescola_nao_existentes(set(codescola_nao_existentes), arquivo_saida)

    return arquivo_saida, nome_excel_tempfile


def get_escolas_json(arquivo):
    # Lê os dados de 'escolas.json' e retorna um json.
    with open(arquivo, 'r') as f:
        data = json.load(f)
    return data


def retorna_alunos_nao_matriculados_na_escola(items, escolas, arquivo_saida):  # noqa C901
    data = escolas
    alunos_nao_matriculados_na_escola_lista = []

    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if data.get(escola):
            pertence = any(
                get_codigo_eol_aluno(_aluno['cd_aluno']) == str(aluno['CodEOLAluno'])
                for _aluno in data.get(escola)
            )
        else:
            pertence = False

        if not pertence:
            tupla = (
                str(aluno['CodEOLAluno']),
                aluno['NomeAluno'],
                escola
            )
            alunos_nao_matriculados_na_escola_lista.append(tupla)

    if alunos_nao_matriculados_na_escola_lista:
        # Criar lista com o nome e cod_eol_aluno não matriculado na escola informada.
        return escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida)


def escreve_xlsx_dados_sigpae(items, arquivo_saida):  # noqa C901
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('Dados do SIGPAE para as escolas da planilha')
    ws['A1'] = 'codigo_eol_escola'
    ws['B1'] = 'nome_da_escola'
    ws['C1'] = 'nome_dre'
    ws['D1'] = 'lote'
    ws['E1'] = 'tipo_gestao'
    ws['F1'] = 'contato_email'
    ws['G1'] = 'contato_telefone'
    ws['H1'] = 'contato_telefone2'
    ws['I1'] = 'contato_celular'
    i = 0  # indice criado manualmente pra não inserir linhas em branco na planilha.
    for item in progressbar(items, 'Escrevendo...'):
        escola = Escola.objects.filter(codigo_eol=item).first()
        if escola:
            ws[f'A{i+2}'] = escola.codigo_eol
            ws[f'B{i+2}'] = escola.nome
            if escola.diretoria_regional:
                ws[f'C{i+2}'] = escola.diretoria_regional.nome
            if escola.lote:
                ws[f'D{i+2}'] = escola.lote.nome
            if escola.tipo_gestao:
                ws[f'E{i+2}'] = escola.tipo_gestao.nome
            if escola.contato:
                ws[f'F{i+2}'] = escola.contato.email
                ws[f'G{i+2}'] = escola.contato.telefone
                ws[f'H{i+2}'] = escola.contato.telefone2
                ws[f'I{i+2}'] = escola.contato.celular
            i += 1
    wb.save(arquivo_saida)


def retorna_dados_sigpae(items, arquivo_saida):
    aux = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            raise e

    cod_escola_unicos = set(aux)
    escreve_xlsx_dados_sigpae(list(cod_escola_unicos), arquivo_saida)


def escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('CodDiagnostico inexistentes')
    ws['A1'] = 'cod_diagnostico'
    for i, item in enumerate(progressbar(list(cod_diagnostico_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    wb.save(arquivo_saida)


def escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('ProtocoloDieta inexistentes')
    ws['A1'] = 'protocolo_dieta'
    for i, item in enumerate(progressbar(list(protocolo_dieta_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    wb.save(arquivo_saida)


def retorna_cod_diagnostico_inexistentes(items, arquivo_saida):
    cod_diagnostico_unicos = [item['CodDiagnostico'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    cod_diagnostico_inexistentes = set(cod_diagnostico_unicos) - set(protocolo_de_dieta_especial)
    if cod_diagnostico_inexistentes:
        escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida)


def retorna_protocolo_dieta_inexistentes(items, arquivo_saida):
    protocolo_dieta_unicos = [item['ProtocoloDieta'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    protocolo_dieta_inexistentes = set(protocolo_dieta_unicos) - set(protocolo_de_dieta_especial)
    if protocolo_dieta_inexistentes:
        escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida)


def string_to_date(texto, formato):
    # Transforma string em data.
    return datetime.strptime(texto, formato).date()


def escreve_xlsx_alunos_com_nome_diferente(lista, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('Alunos com nome diferente do EOL')
    ws['A1'] = 'nome_aluno_planilha'
    ws['B1'] = 'nome_aluno_eol'
    for i, item in enumerate(progressbar(lista, 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
    wb.save(arquivo_saida)


def escreve_xlsx_alunos_com_nascimento_diferente(lista, arquivo_saida):
    wb = openpyxl.load_workbook(arquivo_saida)
    ws = wb.create_sheet('Alunos com nascimento diferente do EOL')
    ws['A1'] = 'nascimento_planilha'
    ws['B1'] = 'nascimento_eol'
    for i, item in enumerate(progressbar(lista, 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
    wb.save(arquivo_saida)


def retorna_alunos_com_nome_diferente(items, escolas, arquivo_saida):  # noqa C901
    aux = []
    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if escolas.get(escola):
            aluno_localizado = list(filter(lambda x: x['cd_aluno'] == aluno['CodEOLAluno'], escolas.get(escola)))
            if aluno_localizado:
                nome_aluno_planilha = aluno['NomeAluno']
                nome_aluno_eol = aluno_localizado[0]['nm_aluno']
                if nome_aluno_planilha != nome_aluno_eol:
                    tupla = (nome_aluno_planilha, nome_aluno_eol)
                    aux.append(tupla)
    if aux:
        escreve_xlsx_alunos_com_nome_diferente(aux, arquivo_saida)


def retorna_alunos_com_nascimento_diferente(items, escolas, arquivo_saida):  # noqa C901
    aux = []
    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if escolas.get(escola):
            aluno_localizado = list(filter(lambda x: x['cd_aluno'] == aluno['CodEOLAluno'], escolas.get(escola)))
            if aluno_localizado:
                nascimento_planilha = string_to_date(aluno['DataNascimento'], '%d/%m/%Y')
                nascimento_eol = string_to_date(aluno_localizado[0]['dt_nascimento_aluno'], '%Y-%m-%dT%H:%M:%S')
                if nascimento_planilha != nascimento_eol:
                    tupla = (nascimento_planilha, nascimento_eol)
                    aux.append(tupla)
    if aux:
        escreve_xlsx_alunos_com_nascimento_diferente(aux, arquivo_saida)


def main(arquivo, arquivo_codigos_escolas, tempfile):
    items = excel_to_list_with_openpyxl(arquivo, in_memory=False)
    escolas = get_escolas_json(tempfile)

    # 1
    items_codigos_escolas = excel_to_list_with_openpyxl(arquivo_codigos_escolas, in_memory=False)
    gera_dict_codigos_escolas(items_codigos_escolas)
    resultado, nome_excel_tempfile = retorna_codigo_eol_escolas_nao_identificadas(items)

    # 2
    # Usa items_codigos_escolas
    # Usa gera_dict_codigos_escolas
    gera_dict_codigo_aluno_por_codigo_escola(items)
    retorna_alunos_nao_matriculados_na_escola(items, escolas, nome_excel_tempfile)

    # 3 - Retorna nome e data de nascimento que forem diferentes entre a planilha e o EOL.
    retorna_alunos_com_nome_diferente(items, escolas, nome_excel_tempfile)
    retorna_alunos_com_nascimento_diferente(items, escolas, nome_excel_tempfile)

    # 5
    # Usa items_codigos_escolas
    # Usa gera_dict_codigos_escolas
    retorna_dados_sigpae(items, nome_excel_tempfile)

    # 6 Verificar os campos "CodDiagnostico" e "ProtocoloDieta"
    retorna_cod_diagnostico_inexistentes(items, nome_excel_tempfile)
    retorna_protocolo_dieta_inexistentes(items, nome_excel_tempfile)
    res, arquivo_final = escreve_xlsx_primeira_aba(resultado, nome_excel_tempfile)
    return res, arquivo_final

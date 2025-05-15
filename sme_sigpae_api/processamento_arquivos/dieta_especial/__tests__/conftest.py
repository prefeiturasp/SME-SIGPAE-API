import uuid
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from model_mommy import mommy
from openpyxl import Workbook
from openpyxl.cell import Cell

from sme_sigpae_api.dieta_especial.models import (
    AlergiaIntolerancia,
    Alimento,
    ArquivoCargaAlimentosSubstitutos,
    ArquivoCargaDietaEspecial,
    ArquivoCargaUsuariosEscola,
    ClassificacaoDieta,
    SolicitacaoDietaEspecial,
)
from sme_sigpae_api.processamento_arquivos.dieta_especial.schemas import (
    ArquivoCargaDietaEspecialSchema,
)


@pytest.fixture
def arquivo_carga_dieta_especial():
    return mommy.make(ArquivoCargaDietaEspecial)


@pytest.fixture
def arquivo_carga_alimentos_e_substitutos():
    return mommy.make(ArquivoCargaAlimentosSubstitutos)


@pytest.fixture
def arquivo_carga_usuarios_escola():
    return mommy.make(ArquivoCargaUsuariosEscola)


@pytest.fixture
def aluno():
    return mommy.make("escola.Aluno", codigo_eol="1234567", nome="TESTE ALUNO DIETA")


@pytest.fixture
def perfil():
    return mommy.make("Perfil", nome="DIRETOR_UE")


@pytest.fixture
def escola():
    return mommy.make(
        "escola.Escola",
        codigo_codae="12345678",
        codigo_eol="654321",
        lote=mommy.make("Lote", terceirizada=mommy.make("Terceirizada")),
    )


@pytest.fixture
def classificacao_dieta():
    return mommy.make(ClassificacaoDieta, nome="Tipo A")


@pytest.fixture
def edital(escola):
    edital = mommy.make(
        "Edital",
        numero="Edital MAIO/2025",
        uuid="12288b47-9d27-4089-8c2e-48a6061d83ea",
    )
    contrato = mommy.make(
        "Contrato",
        terceirizada=mommy.make("Terceirizada"),
        edital=edital,
        make_m2m=True,
        uuid="44d51e10-8999-48bb-889a-1540c9e8c895",
    )
    contrato.lotes.set([escola.lote])
    return edital


@pytest.fixture
def protocolo_padrao_dieta_especial(edital):
    protocolo = mommy.make(
        "ProtocoloPadraoDietaEspecial",
        nome_protocolo="ALERGIA - OVO",
        uuid="5d7f80b8-7b62-441b-89da-4d5dd5c1e7e8",
    )
    protocolo.editais.add(edital)
    return protocolo


@pytest.fixture
def substituicao_alimento(protocolo_padrao_dieta_especial):
    return mommy.make(
        "SubstituicaoAlimentoProtocoloPadrao",
        protocolo_padrao=protocolo_padrao_dieta_especial,
    )


@pytest.fixture
def alergia_intolerancia():
    return [
        mommy.make(AlergiaIntolerancia, descricao="Pão"),
        mommy.make(AlergiaIntolerancia, descricao="Bolo"),
        mommy.make(AlergiaIntolerancia, descricao="Biscoito"),
    ]


@pytest.fixture
def solicitacao_dieta_especial(aluno, escola, classificacao_dieta):
    return mommy.make(
        SolicitacaoDietaEspecial,
        aluno=aluno,
        ativo=True,
        classificacao=classificacao_dieta,
        escola_destino=escola,
        nome_protocolo="Alérgico",
        eh_importado=True,
    )


@pytest.fixture
def dieta_especial_ativa(solicitacao_dieta_especial):
    alergias_set = mommy.make(
        AlergiaIntolerancia, descricao="Aluno Alérgico", _quantity=1
    )
    solicitacao_dieta_especial.alergias_intolerancias.set(alergias_set)
    solicitacao_dieta_especial.save()
    return solicitacao_dieta_especial


@pytest.fixture
def usuario():
    return mommy.make("perfil.Usuario")


@pytest.fixture
def usuario_diretor():
    return mommy.make("perfil.Usuario")


@pytest.fixture
def perfil_diretor():
    return mommy.make("perfil.perfil", nome="DIRETOR")


@pytest.fixture
def mock_cabecalho_e_informacoes_excel(
    aluno,
    escola,
    protocolo_padrao_dieta_especial,
    classificacao_dieta,
    alergia_intolerancia,
):
    alergia = ";".join([alergia.descricao for alergia in alergia_intolerancia])
    cabecalhos = [
        "dre",
        "tipo_gestao",
        "tipo_unidade",
        "codigo_escola",
        "nome_unidade",
        "codigo_eol_aluno",
        "nome_aluno",
        "data_nascimento",
        "data_ocorrencia",
        "codigo_diagnostico",
        "protocolo_dieta",
        "codigo_categoria_dieta",
    ]
    informacoes = [
        "DRE XYZ",  # dre
        "TERCEIRIZADA",  # tipo_gestao
        "EMEF",  # tipo_unidade
        escola.codigo_codae,  # codigo_escola
        "Escola Teste",  # nome_unidade
        aluno.codigo_eol,  # codigo_eol_aluno (obrigatório)
        aluno.nome,  # nome_aluno
        "01/01/2010",  # data_nascimento
        "15/05/2023",  # data_ocorrencia
        f"{alergia};Maionese",  # codigo_diagnostico (obrigatório)
        protocolo_padrao_dieta_especial.nome_protocolo,  # protocolo_dieta (obrigatório)
        classificacao_dieta.nome[-1],  # codigo_categoria_dieta (obrigatório)
    ]
    return cabecalhos, informacoes


@pytest.fixture
def arquivo_carga_dieta_especial_com_informacoes(
    arquivo_carga_dieta_especial, mock_cabecalho_e_informacoes_excel
):
    cabecalhos, informacoes = mock_cabecalho_e_informacoes_excel
    wb = Workbook()
    ws = wb.active
    ws.append(cabecalhos)
    ws.append(informacoes)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    arquivo = SimpleUploadedFile(
        name=f"{uuid.uuid4()}.xlsx",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_dieta_especial.conteudo = arquivo
    arquivo_carga_dieta_especial.save()
    return arquivo_carga_dieta_especial


@pytest.fixture
def arquivo_extensao_incorreta(arquivo_carga_dieta_especial):
    wb = Workbook()
    ws = wb.active
    ws.append(["A"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    arquivo = SimpleUploadedFile(
        name=f"{uuid.uuid4()}.xlsm",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_dieta_especial.conteudo = arquivo
    arquivo_carga_dieta_especial.save()
    return arquivo_carga_dieta_especial


@pytest.fixture
def arquivo_colunas_incorreta(arquivo_carga_dieta_especial):
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    arquivo = SimpleUploadedFile(
        name=f"{uuid.uuid4()}.xlsx",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_dieta_especial.conteudo = arquivo
    arquivo_carga_dieta_especial.save()
    return arquivo_carga_dieta_especial


@pytest.fixture
def solicitacao_dieta_schema(mock_cabecalho_e_informacoes_excel):
    cabecalhos, informacoes = mock_cabecalho_e_informacoes_excel
    dicionario_dados = dict(zip(cabecalhos, informacoes))
    return ArquivoCargaDietaEspecialSchema(**dicionario_dados)


@pytest.fixture
def alimentos():
    return [
        mommy.make(
            "dieta_especial.Alimento",
            nome="ARROZ",
            tipo_listagem_protocolo=Alimento.SO_ALIMENTOS,
        ),
        mommy.make(
            "dieta_especial.Alimento",
            nome="FEIJAO",
            tipo_listagem_protocolo=Alimento.SO_ALIMENTOS,
        ),
        mommy.make(
            "dieta_especial.Alimento",
            nome="BATATA",
            tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS,
        ),
    ]


@pytest.fixture
def alimentos_substitutivos():
    return [
        mommy.make(
            "dieta_especial.Alimento",
            nome="AVEIA",
            tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS,
        ),
        mommy.make(
            "dieta_especial.Alimento",
            nome="LENTILHA",
            tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS,
        ),
        mommy.make(
            "dieta_especial.Alimento",
            nome="ERVILHA",
            tipo_listagem_protocolo=Alimento.SO_ALIMENTOS,
        ),
    ]


@pytest.fixture
def arquivo_carga_alimentos_com_informacoes(
    arquivo_carga_alimentos_e_substitutos, alimentos, alimentos_substitutivos
):
    cabecalho = ["nome"]
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Aba1"
    ws1.append(cabecalho)
    for alimento in alimentos:
        ws1.append([alimento.nome])
    ws1.append(["FARINHA DE TRIGO"])

    ws2 = wb.create_sheet("Aba2")
    ws2.append(cabecalho)
    for alimento in alimentos_substitutivos:
        ws2.append([alimento.nome])
    ws2.append(["FARINHA DE MILHO"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    arquivo = SimpleUploadedFile(
        name=f"{uuid.uuid4()}.xlsx",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_alimentos_e_substitutos.conteudo = arquivo
    arquivo_carga_alimentos_e_substitutos.save()
    return arquivo_carga_alimentos_e_substitutos


@pytest.fixture
def arquivo_alimentos_abas_incorreta(arquivo_carga_alimentos_e_substitutos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Aba1"
    ws.append(["A", "B"])
    ws.append(["Arroz", "Feijão"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    arquivo = SimpleUploadedFile(
        name=f"{uuid.uuid4()}.xlsx",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_alimentos_e_substitutos.conteudo = arquivo
    arquivo_carga_alimentos_e_substitutos.save()
    return arquivo_carga_alimentos_e_substitutos


@pytest.fixture
def dados_planilha_valida(escola):
    return {
        "dre": "DRE Teste",
        "unidade_escola": "Escola Teste",
        "codigo_eol_escola": escola.codigo_eol,
        "nome_diretor": "Diretor Teste",
        "rg_diretor": "12345678",
        "rf_diretor": "1234567",
        "cpf_diretor": "12345678901",
        "email_diretor": "diretor@teste.com",
        "telefone_diretor": "11999999999",
        "nome_assistente": "Assistente Teste",
        "rg_assistente": "87654321",
        "rf_assistente": "4710987",
        "cpf_assistente": "10987654321",
        "email_assistente": "assistenter@email",
        "telefone_assistente": "11888888888",
    }


@pytest.fixture
def dados_planilha_invalida(escola, usuario_diretor):
    return {
        "dre": "DRE Teste",
        "unidade_escola": "Escola Teste",
        "codigo_eol_escola": escola.codigo_codae,
        "nome_diretor": "Diretor Teste",
        "rg_diretor": "12345678",
        "rf_diretor": "1234567",
        "cpf_diretor": "12345678901",
        "email_diretor": "diretor@teste.com",
        "telefone_diretor": "11999999999",
        "nome_assistente": "Assistente Teste",
        "rg_assistente": "87654321",
        "rf_assistente": usuario_diretor.registro_funcional,
        "cpf_assistente": "10987654321",
        "email_assistente": usuario_diretor.email,
        "telefone_assistente": "11888888888",
    }


@pytest.fixture
def linha_planilha_valida(dados_planilha_valida):
    colunas = [
        "dre",
        "unidade_escola",
        "codigo_eol_escola",
        "nome_diretor",
        "rg_diretor",
        "rf_diretor",
        "cpf_diretor",
        "email_diretor",
        "telefone_diretor",
        "nome_assistente",
        "rg_assistente",
        "rf_assistente",
        "cpf_assistente",
        "email_assistente",
        "telefone_assistente",
    ]
    valores = [dados_planilha_valida[col] for col in colunas]
    return tuple(Cell(worksheet=None, value=valor) for valor in valores)


@pytest.fixture
def arquivo_carga_usuario_escola_com_informacoes(
    arquivo_carga_usuarios_escola, dados_planilha_valida
):
    wb = Workbook()
    ws = wb.active
    cabecalho = [
        "dre",
        "unidade_escola",
        "codigo_eol_escola",
        "nome_diretor",
        "rg_diretor",
        "rf_diretor",
        "cpf_diretor",
        "email_diretor",
        "telefone_diretor",
        "nome_assistente",
        "rg_assistente",
        "rf_assistente",
        "cpf_assistente",
        "email_assistente",
        "telefone_assistente",
    ]
    ws.append(cabecalho)

    # Dados na mesma ordem
    dados = [dados_planilha_valida[col] for col in cabecalho]
    ws.append(dados)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    arquivo = SimpleUploadedFile(
        name="teste_valido.xlsx",
        content=buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    arquivo_carga_usuarios_escola.conteudo = arquivo
    arquivo_carga_usuarios_escola.save()
    return arquivo_carga_usuarios_escola

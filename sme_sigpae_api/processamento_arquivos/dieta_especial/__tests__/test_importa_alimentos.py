import pytest
from openpyxl import load_workbook
from openpyxl.cell import Cell

from sme_sigpae_api.dados_comuns.constants import StatusProcessamentoArquivo
from sme_sigpae_api.dieta_especial.models import (
    Alimento,
    ArquivoCargaAlimentosSubstitutos,
)
from sme_sigpae_api.processamento_arquivos.dieta_especial import (
    importa_alimentos as imp_alimentos,
)
from sme_sigpae_api.processamento_arquivos.dieta_especial.importa_alimentos import (
    ProcessadorPlanilha,
)

pytestmark = pytest.mark.django_db


def test_importa_dietas_especiais(arquivo_carga_alimentos_com_informacoes):
    importacao = imp_alimentos(arquivo_carga_alimentos_com_informacoes)
    assert importacao is None


def test_existe_conteudo_retorna_true(arquivo_carga_alimentos_com_informacoes):
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    tem_conteudo = processador.existe_conteudo()
    assert tem_conteudo is True


def test_existe_conteudo_retorna_false(arquivo_carga_alimentos_e_substitutos):
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_e_substitutos)
    tem_conteudo = processador.existe_conteudo()
    assert tem_conteudo is False


def test_extensao_do_arquivo_esta_correta_retorna_true(
    arquivo_carga_alimentos_com_informacoes,
):
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    tem_extensao_correta = processador.extensao_do_arquivo_esta_correta()
    assert tem_extensao_correta is True


def test_extensao_do_arquivo_esta_correta_retorna_false(arquivo_extensao_incorreta):
    processador = ProcessadorPlanilha(arquivo_extensao_incorreta)
    tem_extensao_correta = processador.extensao_do_arquivo_esta_correta()
    assert tem_extensao_correta is False


def test_validacao_inicial(arquivo_carga_alimentos_com_informacoes):
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    validado = processador.validacao_inicial()
    assert validado is True


def test_monta_dicionario_de_dados(
    arquivo_carga_alimentos_com_informacoes,
    alimentos,
):
    cabecalhos = ["nome"]
    informacoes = [alimento.nome for alimento in alimentos]
    linha = tuple(Cell(worksheet=None, value=valor) for valor in informacoes)
    dicionario = dict(zip(cabecalhos, informacoes))
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    dicionario_dados = processador.monta_dicionario_de_dados(linha)
    assert dicionario_dados == dicionario


def test_processa_alimentos_so_alimentos(arquivo_carga_alimentos_com_informacoes):
    assert Alimento.objects.all().count() == 6
    assert Alimento.objects.filter(tipo_listagem_protocolo=Alimento.AMBOS).count() == 0
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_ALIMENTOS).count()
        == 3
    )
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS).count()
        == 3
    )

    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    workbook = load_workbook(processador.path)
    worksheet_alimentos = workbook.worksheets[0]
    validado = processador.processa_alimentos(
        worksheet=worksheet_alimentos, tipo_listagem=Alimento.SO_ALIMENTOS
    )
    assert validado is None

    assert Alimento.objects.all().count() == 7
    assert Alimento.objects.filter(tipo_listagem_protocolo=Alimento.AMBOS).count() == 1
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_ALIMENTOS).count()
        == 4
    )
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS).count()
        == 2
    )


def test_processa_alimentos_so_substitutivos(arquivo_carga_alimentos_com_informacoes):
    assert Alimento.objects.all().count() == 6
    assert Alimento.objects.filter(tipo_listagem_protocolo=Alimento.AMBOS).count() == 0
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_ALIMENTOS).count()
        == 3
    )
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS).count()
        == 3
    )

    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    workbook = load_workbook(processador.path)
    worksheet_alimentos = workbook.worksheets[1]
    validado = processador.processa_alimentos(
        worksheet=worksheet_alimentos, tipo_listagem=Alimento.SO_SUBSTITUTOS
    )
    assert validado is None

    assert Alimento.objects.all().count() == 7
    assert Alimento.objects.filter(tipo_listagem_protocolo=Alimento.AMBOS).count() == 1
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_ALIMENTOS).count()
        == 2
    )
    assert (
        Alimento.objects.filter(tipo_listagem_protocolo=Alimento.SO_SUBSTITUTOS).count()
        == 4
    )


def test_finaliza_processamento(arquivo_carga_alimentos_com_informacoes):
    assert (
        ArquivoCargaAlimentosSubstitutos.objects.last().status
        == StatusProcessamentoArquivo.PENDENTE.value
    )
    processador = ProcessadorPlanilha(arquivo_carga_alimentos_com_informacoes)
    processador.processamento()
    processador.finaliza_processamento()
    assert (
        ArquivoCargaAlimentosSubstitutos.objects.last().status
        == StatusProcessamentoArquivo.SUCESSO.value
    )


def test_processamento_erro_validacao_inicial(arquivo_extensao_incorreta):
    processador = ProcessadorPlanilha(arquivo_extensao_incorreta)
    assert len(processador.erros) == 0
    erro_validacao_inicial = processador.processamento()
    assert erro_validacao_inicial is None
    assert len(processador.erros) == 0


def test_processamento_erro_abas(arquivo_alimentos_abas_incorreta):
    assert (
        ArquivoCargaAlimentosSubstitutos.objects.last().status
        == StatusProcessamentoArquivo.PENDENTE.value
    )
    processador = ProcessadorPlanilha(arquivo_alimentos_abas_incorreta)
    assert len(processador.erros) == 0
    erro_abas = processador.processamento()
    assert erro_abas is None
    assert (
        ArquivoCargaAlimentosSubstitutos.objects.last().status
        == StatusProcessamentoArquivo.ERRO.value
    )
    assert (
        ArquivoCargaAlimentosSubstitutos.objects.last().log
        == "Erro: Número de abas na planilha é diferente de 2"
    )

import io
import os
import uuid
from datetime import datetime

import pandas as pd
import pytest
from faker import Faker
from freezegun import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.paineis_consolidados.api import constants
from sme_sigpae_api.paineis_consolidados.models import SolicitacoesCODAE
from sme_sigpae_api.paineis_consolidados.tasks import (
    aplica_fundo_amarelo_canceladas,
    aplica_fundo_amarelo_tipo1,
    aplica_fundo_amarelo_tipo2,
    build_pdf,
    build_subtitulo,
    build_xlsx,
    gera_pdf_relatorio_solicitacoes_alimentacao_async,
    gera_xls_relatorio_solicitacoes_alimentacao_async,
    get_formats,
    nomes_colunas,
    novas_linhas_inc_continua_e_kit_lanche,
)
from sme_sigpae_api.relatorios.utils import extrair_texto_de_pdf

pytestmark = pytest.mark.django_db
f = Faker("pt-br")


def test_xls_relatorio_status(users_diretor_escola):
    usuario = users_diretor_escola[5]
    uuids = ["7fa6e609-db33-48e1-94ea-6d5a0c07935c"]
    lotes = []
    tipos_solicitacao = []
    tipos_unidade = []
    unidades_educacionais = []
    request_data = {
        "status": "CANCELADOS",
        "tipos_solicitacao": ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    resultado = gera_xls_relatorio_solicitacoes_alimentacao_async.delay(
        user=usuario.username,
        nome_arquivo="relatorio_solicitacoes_alimentacao.xlsx",
        data=request_data,
        uuids=uuids,
        lotes=lotes,
        tipos_solicitacao=tipos_solicitacao,
        tipos_unidade=tipos_unidade,
        unidades_educacionais=unidades_educacionais,
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


def test_pdf_relatorio_status(users_diretor_escola):
    usuario = users_diretor_escola[5]
    uuids = ["7fa6e609-db33-48e1-94ea-6d5a0c07935c"]
    request_data = {
        "status": "CANCELADOS",
        "tipos_solicitacao": ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    resultado = gera_pdf_relatorio_solicitacoes_alimentacao_async.delay(
        user=usuario.username,
        nome_arquivo="relatorio_solicitacoes_alimentacao.pdf",
        data=request_data,
        uuids=uuids,
        status=request_data.get("status", None),
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


@freeze_time("2025-01-22")
def test_build_xlsx(dados_para_geracao_excel_e_pdf):
    instituicao, sem_uuids_repetido, serializer, data = dados_para_geracao_excel_e_pdf
    output = io.BytesIO()
    lotes = []
    tipos_solicitacao = ["SUSP_ALIMENTACAO", "INC_ALIMENTA"]
    tipos_unidade = []
    unidades_educacionais = []
    build_xlsx(
        output,
        serializer,
        sem_uuids_repetido,
        data,
        lotes,
        tipos_solicitacao,
        tipos_unidade,
        unidades_educacionais,
        instituicao,
    )
    workbook = load_workbook(output)
    sheet = workbook[f"Relatório - Canceladas"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Solicitações de Alimentação Canceladas",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "Total de Solicitações Canceladas: 11 | Tipo(s) de solicitação(ões): Suspensão de Alimentação, Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "N",
        "Lote",
        "Unidade Educacional",
        "Tipo de Solicitação",
        "ID da Solicitação",
        "Data do Evento",
        "Dia da semana",
        "Período",
        "Tipo de Alimentação",
        "Tipo de Alteração",
        "Nª de Alunos",
        "N° total de Kits",
        "Observações",
        "Data de Cancelamento",
    )
    assert rows[4][3] == "Inclusão de Alimentação"
    assert rows[4][5] == "14/01/2025"
    assert rows[4][12] == "cancelado"


def test_novas_linhas_inc_continua_e_kit_lanche(dados_para_geracao_excel_e_pdf):
    instituicao, queryset, serializer, _ = dados_para_geracao_excel_e_pdf

    df = pd.DataFrame(serializer.data)
    novas_colunas = ["dia_semana", "periodo_inclusao", "tipo_alimentacao"]
    for i, nova_coluna in enumerate(novas_colunas):
        df.insert(constants.COL_IDX_DATA_EVENTO + i, nova_coluna, "-")
    df.insert(constants.COL_IDX_NUMERO_DE_ALUNOS, "quantidade_alimentacoes", "-")
    novas_linhas, lista_uuids = novas_linhas_inc_continua_e_kit_lanche(
        df, queryset, instituicao
    )

    assert isinstance(novas_linhas, list)
    for linha in novas_linhas:
        assert isinstance(linha, pd.Series)

    assert isinstance(lista_uuids, list)
    for lu in lista_uuids:
        assert isinstance(lu, SolicitacoesCODAE)


def test_get_formats():
    output_file = "/tmp/test.xlsx"
    xlwriter = pd.ExcelWriter(output_file, engine="xlsxwriter")
    workbook = xlwriter.book
    bg_color = "#a9d18e"
    border_color = "#198459"

    merge_format, cell_format, single_cell_format = get_formats(workbook)

    assert merge_format.right_color == border_color
    assert merge_format.bottom_color == border_color
    assert merge_format.left_color == border_color
    assert merge_format.top_color == border_color
    assert merge_format.bg_color == bg_color
    assert merge_format.bold == True

    assert cell_format.bold == True
    assert cell_format.text_wrap == True

    assert single_cell_format.bg_color == bg_color

    workbook.close()
    xlwriter.close()
    if os.path.exists(output_file):
        os.remove(output_file)


@freeze_time("2025-01-22")
def test_build_subtitulo_autorizadas(dados_para_geracao_excel_e_pdf):
    _, queryset, _, _ = dados_para_geracao_excel_e_pdf

    data = {
        "status": "AUTORIZADAS",
        "tipos_solicitacao": ["INC_ALIMENTA"],
        "de": "01/01/2025",
        "ate": "28/02/2025",
    }
    subtitulo_autorizadas = build_subtitulo(
        data,
        "Autorizadas",
        queryset,
        [],
        ["INC_ALIMENTA"],
        [],
        [],
    )
    assert (
        subtitulo_autorizadas
        == "Total de Solicitações Autorizadas: 11 | Tipo(s) de solicitação(ões): Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025"
    )


@freeze_time("2025-01-22")
def test_build_subtitulo_canceladas(dados_para_geracao_excel_e_pdf):
    _, queryset, _, data = dados_para_geracao_excel_e_pdf

    subtitulo_canceladas = build_subtitulo(
        data,
        "Canceladas",
        queryset,
        [],
        ["SUSP_ALIMENTACAO", "INC_ALIMENTA"],
        [],
        [],
    )
    assert (
        subtitulo_canceladas
        == "Total de Solicitações Canceladas: 11 | Tipo(s) de solicitação(ões): Suspensão de Alimentação, Inclusão de Alimentação | Data inicial: 01/01/2025 | Data final: 28/02/2025 | Data de Extração do Relatório: 22/01/2025"
    )


def test_nomes_colunas(dados_para_montar_excel):
    (
        linhas,
        colunas,
        _,
        _,
        _,
        worksheet,
        single_cell_format,
        nome_aba,
    ) = dados_para_montar_excel

    nomes_colunas(worksheet, "Canceladas", linhas, colunas, single_cell_format)

    assert len(worksheet.col_sizes) == len(colunas)
    assert len(worksheet.row_sizes) == 2
    assert worksheet.name == nome_aba


def test_aplica_fundo_amarelo_canceladas(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)
    aplica_fundo_amarelo_canceladas(df, worksheet, workbook, queryset, linhas, colunas)

    xlwriter.close()
    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.coordinate == "F5":
                assert cell.fill.start_color.rgb == "FFFFFF00"
    workbook_openpyxl.close()


def test_aplica_fundo_amarelo_tipo1(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)
    indexes_cancelados = []

    for index, solicitacao in enumerate(queryset):
        model_obj = solicitacao.get_raw_model.objects.get(uuid=solicitacao.uuid)
        if model_obj.status == "ESCOLA_CANCELOU":
            indexes_cancelados.append(index)
            aplica_fundo_amarelo_tipo1(
                df, worksheet, workbook, solicitacao, model_obj, linhas, colunas, index
            )

    xlwriter.close()
    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for index, row in enumerate(sheet.iter_rows()):
        for cell in row:
            if cell.coordinate == "F5" and index in indexes_cancelados:
                assert cell.fill.start_color.rgb == "FFFFFF00"

    workbook_openpyxl.close()


def test_aplica_fundo_amarelo_tipo2(
    dados_para_geracao_excel_e_pdf, dados_para_montar_excel
):
    _, queryset, serializer, _ = dados_para_geracao_excel_e_pdf
    (
        linhas,
        colunas,
        output_file,
        xlwriter,
        workbook,
        worksheet,
        _,
        nome_aba,
    ) = dados_para_montar_excel

    df = pd.DataFrame(serializer.data)
    df.to_excel(xlwriter, nome_aba, index=False)

    solicitacao = queryset[0]
    model_obj = solicitacao.get_raw_model.objects.get(uuid=solicitacao.uuid)
    idx = aplica_fundo_amarelo_tipo2(
        df, worksheet, workbook, queryset[0], model_obj, linhas, colunas, 0, 0
    )
    xlwriter.close()
    assert idx == 0

    workbook_openpyxl = load_workbook(output_file)
    sheet = workbook_openpyxl[nome_aba]
    for row in sheet.iter_rows():
        for cell in row:
            assert cell.fill.start_color.rgb == "00000000"
    workbook_openpyxl.close()


def test_build_pdf():
    pdf_cancelados = build_pdf([], "CANCELADOS")
    assert isinstance(pdf_cancelados, bytes)

    texto = extrair_texto_de_pdf(pdf_cancelados).lower()
    assert "Total de Solicitações Cancelados:  0".lower() in texto
    assert "SIGPAE - RELATÓRIO DE solicitações de alimentação".lower() in texto
    assert datetime.now().strftime("%d/%m/%Y") in texto
    assert texto.count("cancelados") == 2

    pdf_autorizados = build_pdf([], "AUTORIZADOS")
    assert isinstance(pdf_autorizados, bytes)

    texto = extrair_texto_de_pdf(pdf_autorizados).lower()
    assert "Total de Solicitações autorizados:  0".lower() in texto
    assert "SIGPAE - RELATÓRIO DE solicitações de alimentação".lower() in texto
    assert datetime.now().strftime("%d/%m/%Y") in texto
    assert texto.count("autorizados") == 2

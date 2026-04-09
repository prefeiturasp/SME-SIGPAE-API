import io
import uuid

import pytest
from model_bakery import baker
from openpyxl import load_workbook

from sme_sigpae_api.dados_comuns.fluxo_status import DietaEspecialWorkflow
from sme_sigpae_api.dados_comuns.models import CentralDeDownload
from sme_sigpae_api.dieta_especial.api.serializers import (
    SolicitacaoDietaEspecialExportXLSXSerializer,
)
from sme_sigpae_api.dieta_especial.models import (
    SolicitacaoDietaEspecial,
)
from sme_sigpae_api.dieta_especial.tasks import (
    gera_pdf_relatorio_dieta_especial_async,
    gera_pdf_relatorio_dietas_especiais_terceirizadas_async,
    gera_xlsx_relatorio_dietas_especiais_terceirizadas_async,
)
from sme_sigpae_api.dieta_especial.tasks.utils.relatorio_terceirizadas_xlsx import (
    build_titulo,
    build_xlsx_relatorio_terceirizadas,
)

pytestmark = pytest.mark.django_db


def test_gera_pdf_relatorio_dieta_especial_async(
    usuario_com_pk,
    escola_cei,
    alergias_intolerancias,
    solicitacoes_processa_dieta_especial,
):
    request_data = {
        "dre": escola_cei.diretoria_regional.uuid,
        "escola": [escola_cei.uuid],
        "diagnostico": [ai.pk for ai in alergias_intolerancias],
        "data_inicial": "2025-01-01",
        "data_final": "2025-01-31",
    }
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]

    resultado = gera_pdf_relatorio_dieta_especial_async.delay(
        user=usuario_com_pk.username,
        nome_arquivo="relatorio_dieta_especial.pdf",
        data=request_data,
        ids_dietas=ids_dietas,
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


def test_gera_pdf_relatorio_dieta_especial(
    usuario_com_pk,
    escola_cei,
    alergias_intolerancias,
    solicitacoes_processa_dieta_especial,
):
    request_data = {
        "dre": escola_cei.diretoria_regional.uuid,
        "escola": [escola_cei.uuid],
        "diagnostico": [ai.pk for ai in alergias_intolerancias],
        "data_inicial": "2025-01-01",
        "data_final": "2025-01-31",
    }
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]
    nome_arquivo = "relatorio_dieta_especial.pdf"

    gera_pdf_relatorio_dieta_especial_async(
        user=usuario_com_pk.username,
        nome_arquivo=nome_arquivo,
        data=request_data,
        ids_dietas=ids_dietas,
    )
    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO


def test_gera_pdf_relatorio_dieta_especial_erro(usuario_com_pk):
    nome_arquivo = "relatorio_dieta_especial.pdf"

    gera_pdf_relatorio_dieta_especial_async(
        user=usuario_com_pk.username,
        nome_arquivo=nome_arquivo,
        data=[],
        ids_dietas=[],
    )
    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_ERRO


def test_gera_pdf_relatorio_dietas_especiais_terceirizadas_async(
    usuario_com_pk, solicitacoes_processa_dieta_especial
):
    request_data = {"status_selecionado": "CANCELADAS"}
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]
    resultado = gera_pdf_relatorio_dietas_especiais_terceirizadas_async.delay(
        user=usuario_com_pk.username,
        data=request_data,
        nome_arquivo="relatorio_dietas_especiais.pdf",
        ids_dietas=ids_dietas,
        filtros="texto a ser enviado",
    )
    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


def test_gera_pdf_relatorio_dietas_especiais_terceirizadas(
    usuario_com_pk, solicitacoes_processa_dieta_especial
):
    request_data = {"status_selecionado": "CANCELADAS"}
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]

    nome_arquivo = "relatorio_dietas_especiais.pdf"
    gera_pdf_relatorio_dietas_especiais_terceirizadas_async(
        user=usuario_com_pk.username,
        data=request_data,
        nome_arquivo=nome_arquivo,
        ids_dietas=ids_dietas,
        filtros="texto a ser enviado",
    )
    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO


def test_gera_pdf_relatorio_dietas_especiais_terceirizadas_com_erro(usuario_com_pk):
    request_data = {"status_selecionado": "CANCELADAS"}
    ultima_solicitacao = baker.make(
        SolicitacaoDietaEspecial, status=DietaEspecialWorkflow.CODAE_AUTORIZADO
    )
    ids_dietas = [ultima_solicitacao.pk]

    nome_arquivo = "relatorio_dietas_especiais.pdf"

    gera_pdf_relatorio_dietas_especiais_terceirizadas_async(
        user=usuario_com_pk.username,
        data=request_data,
        nome_arquivo=nome_arquivo,
        ids_dietas=ids_dietas,
        filtros="texto a ser enviado",
    )
    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_ERRO


def test_build_titulo(
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
    escola,
    classificacoes_dietas,
):
    lotes = [escola.lote.uuid]
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]

    status = "CANCELADAS"
    data_inicial = "2025-01-01"
    data_final = "2025-01-31"
    resultado_canceladas = build_titulo(
        lotes, status, classificacoes, protocolos, data_inicial, data_final
    )
    assert all(cd.nome in resultado_canceladas for cd in classificacoes_dietas)
    assert escola.lote.nome in resultado_canceladas
    assert protocolo_padrao_dieta_especial.nome_protocolo in resultado_canceladas
    assert protocolo_padrao_dieta_especial_2.nome_protocolo in resultado_canceladas
    assert data_inicial in resultado_canceladas
    assert data_final in resultado_canceladas
    assert (
        resultado_canceladas
        == "Dietas Canceladas: | LOTE 07 | Classificação(ões) da dieta: Tipo A,Tipo A Enteral,Tipo B | Protocolo(s) padrão(ões): ALERGIA A ABACAXI, ALERGIA A AVEIA | Data inicial: 2025-01-01 | Data final: 2025-01-31"
    )

    status = "AUTORIZADAS"
    data_inicial = "01/01/2025"
    data_final = "31/01/2015"
    resultado_autorizadas = build_titulo(
        lotes, status, classificacoes, protocolos, data_inicial, data_final
    )
    assert all(cd.nome in resultado_autorizadas for cd in classificacoes_dietas)
    assert escola.lote.nome in resultado_autorizadas
    assert protocolo_padrao_dieta_especial.nome_protocolo in resultado_autorizadas
    assert protocolo_padrao_dieta_especial_2.nome_protocolo in resultado_autorizadas
    assert data_inicial in resultado_autorizadas
    assert data_final in resultado_autorizadas
    assert (
        resultado_autorizadas
        == "Dietas Autorizadas: | LOTE 07 | Classificação(ões) da dieta: Tipo A,Tipo A Enteral,Tipo B | Protocolo(s) padrão(ões): ALERGIA A ABACAXI, ALERGIA A AVEIA | Data inicial: 01/01/2025 | Data final: 31/01/2015"
    )


def test_build_xlsx_status_cancelado(
    solicitacoes_processa_dieta_especial,
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
    classificacoes_dietas,
):
    output = io.BytesIO()
    query_set = SolicitacaoDietaEspecial.objects.all()
    status = "CANCELADAS"
    serializer = SolicitacaoDietaEspecialExportXLSXSerializer(
        query_set, context={"status": status}, many=True
    )
    lotes = []
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos_padrao = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]
    data_inicial = "2025-01-01"
    data_final = "2025-01-31"
    exibir_diagnostico = False

    build_xlsx_relatorio_terceirizadas(
        output,
        serializer,
        query_set,
        status,
        lotes,
        classificacoes,
        protocolos_padrao,
        data_inicial,
        data_final,
        exibir_diagnostico,
    )
    workbook = load_workbook(output)
    nome_aba = "Solicitações de dieta especial"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de dietas especiais",
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "Dietas Canceladas: | Classificação(ões) da dieta: Tipo A,Tipo A Enteral,Tipo B | Protocolo(s) padrão(ões): ALERGIA A ABACAXI, ALERGIA A AVEIA | Data inicial: 2025-01-01 | Data final: 2025-01-31",
        None,
        None,
        None,
        None,
        None,
        "Total de dietas: 4",
    )
    assert rows[2] == (None, None, None, None, None, None, None)
    assert rows[3] == (
        2,
        "COD.EOL do Aluno",
        "Nome do Aluno",
        "Nome da Escola",
        "Classificação da dieta",
        "Protocolo",
        "Data de cancelamento",
    )
    assert rows[4] == (
        3,
        "123456",
        "Roberto Alves da Silva",
        "CEI DIRET JOAO MENDES",
        None,
        None,
        None,
    )
    assert rows[5] == (
        4,
        "123456",
        "Roberto Alves da Silva",
        "CEI DIRET JOAO MENDES",
        None,
        None,
        None,
    )
    assert rows[6] == (
        5,
        "123456",
        "Roberto Alves da Silva",
        "CEI DIRET JOAO MENDES",
        None,
        None,
        None,
    )
    assert rows[7] == (
        6,
        "123456",
        "Roberto Alves da Silva",
        "CEI DIRET JOAO MENDES",
        None,
        None,
        None,
    )


def test_build_xlsx_status_autorizado(
    solicitacoes_processa_dieta_especial,
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
    classificacoes_dietas,
):
    output = io.BytesIO()
    query_set = SolicitacaoDietaEspecial.objects.all()
    status = "AUTORIZADAS"
    serializer = SolicitacaoDietaEspecialExportXLSXSerializer(
        query_set, context={"status": status}, many=True
    )
    lotes = []
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos_padrao = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]
    data_inicial = "01/01/2025"
    data_final = "31/01/2025"
    exibir_diagnostico = False

    build_xlsx_relatorio_terceirizadas(
        output,
        serializer,
        query_set,
        status,
        lotes,
        classificacoes,
        protocolos_padrao,
        data_inicial,
        data_final,
        exibir_diagnostico,
    )
    workbook = load_workbook(output)
    nome_aba = "Solicitações de dieta especial"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][0] == "Relatório de dietas especiais"

    assert "Dietas Autorizadas:" in rows[1][0]
    assert "Classificação(ões) da dieta: Tipo A,Tipo A Enteral,Tipo B" in rows[1][0]
    assert "Protocolo(s) padrão(ões): ALERGIA A ABACAXI, ALERGIA A AVEIA" in rows[1][0]
    assert "Data inicial: 01/01/2025" in rows[1][0]
    assert "Data final: 31/01/2025" in rows[1][0]
    assert rows[1][-1] == "Total de dietas: 4"

    assert all(cell is None for cell in rows[2])

    assert rows[3][1] == "COD.EOL do Aluno"
    assert rows[3][2] == "Nome do Aluno"
    assert rows[3][3] == "Data de Nascimento"
    assert rows[3][4] == "Nome da Escola"
    assert rows[3][5] == "Classificação da dieta"
    assert rows[3][6] == "Protocolo"

    for row_idx in range(4, 8):
        assert rows[row_idx][1] == "123456"
        assert rows[row_idx][2] == "Roberto Alves da Silva"
        assert rows[row_idx][4] == "CEI DIRET JOAO MENDES"


def test_gera_xlsx_relatorio_dietas_especiais_terceirizadas_async(
    usuario_com_pk,
    solicitacoes_processa_dieta_especial,
    classificacoes_dietas,
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
):
    request_data = {"status_selecionado": "CANCELADAS"}
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]
    lotes = []
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos_padrao = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]

    nome_arquivo = "relatorio_dietas_especiais.xlsx"
    resultado = gera_xlsx_relatorio_dietas_especiais_terceirizadas_async.delay(
        user=usuario_com_pk.username,
        nome_arquivo=nome_arquivo,
        ids_dietas=ids_dietas,
        data=request_data,
        lotes=lotes,
        classificacoes=classificacoes,
        protocolos_padrao=protocolos_padrao,
    )

    assert resultado.status == "SUCCESS"
    assert resultado.id == resultado.task_id
    assert isinstance(resultado.task_id, str)
    assert isinstance(uuid.UUID(resultado.task_id), uuid.UUID)
    assert isinstance(resultado.id, str)
    assert isinstance(uuid.UUID(resultado.id), uuid.UUID)
    assert resultado.ready() is True
    assert resultado.successful() is True


def test_gera_xlsx_relatorio_dietas_especiais_terceirizadas(
    usuario_com_pk,
    solicitacoes_processa_dieta_especial,
    classificacoes_dietas,
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
):
    request_data = {"status_selecionado": "CANCELADAS"}
    ids_dietas = [dieta.pk for dieta in SolicitacaoDietaEspecial.objects.all()]
    lotes = []
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos_padrao = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]

    nome_arquivo = "relatorio_dietas_especiais.xlsx"
    gera_xlsx_relatorio_dietas_especiais_terceirizadas_async(
        user=usuario_com_pk.username,
        nome_arquivo=nome_arquivo,
        ids_dietas=ids_dietas,
        data=request_data,
        lotes=lotes,
        classificacoes=classificacoes,
        protocolos_padrao=protocolos_padrao,
    )

    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO


def test_gera_xlsx_relatorio_dietas_especiais_terceirizadas_erro(
    usuario_com_pk,
    classificacoes_dietas,
    protocolo_padrao_dieta_especial,
    protocolo_padrao_dieta_especial_2,
):
    request_data = {"status_selecionado": "CANCELADAS"}
    ultima_solicitacao = baker.make(
        SolicitacaoDietaEspecial, status=DietaEspecialWorkflow.CODAE_AUTORIZADO
    )
    ids_dietas = [ultima_solicitacao.pk]

    lotes = []
    classificacoes = [cd.id for cd in classificacoes_dietas]
    protocolos_padrao = [
        protocolo_padrao_dieta_especial.uuid,
        protocolo_padrao_dieta_especial_2.uuid,
    ]

    nome_arquivo = "relatorio_dietas_especiais.xlsx"
    gera_xlsx_relatorio_dietas_especiais_terceirizadas_async(
        user=usuario_com_pk.username,
        nome_arquivo=nome_arquivo,
        ids_dietas=ids_dietas,
        data=request_data,
        lotes=lotes,
        classificacoes=classificacoes,
        protocolos_padrao=protocolos_padrao,
    )

    central_download = CentralDeDownload.objects.get(identificador=nome_arquivo)
    assert central_download.arquivo is not None
    assert central_download.status == CentralDeDownload.STATUS_ERRO

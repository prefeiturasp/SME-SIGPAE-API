import json

import pytest
from freezegun.api import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.dados_comuns.models import CentralDeDownload
from sme_sigpae_api.dieta_especial.tasks import (
    gera_xlsx_relatorio_historico_dietas_especiais_async,
)

from .test_setup_relatorio_historico import BaseSetupHistoricoDietas

pytestmark = pytest.mark.django_db


@pytest.mark.usefixtures("client_autenticado_vinculo_codae_gestao_alimentacao_dieta")
@freeze_time("2025-05-09")
class TestGeraXlsxRelatorioHistoricoDietasEspeciaisAsync(BaseSetupHistoricoDietas):
    def test_gera_xlsx_historico_dietas_especiais(
        self, client_autenticado_vinculo_codae_gestao_alimentacao_dieta
    ):
        self.setup()
        client, user = client_autenticado_vinculo_codae_gestao_alimentacao_dieta
        nome_arquivo = "relatorio_historico_dietas_especiais.xlsx"
        data = json.dumps({"lote": str(self.lote.uuid), "data": "09/05/2025"})
        gera_xlsx_relatorio_historico_dietas_especiais_async(user, nome_arquivo, data)

        central_download = CentralDeDownload.objects.get()
        assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO

        with central_download.arquivo.open("rb") as f:
            loaded_wb = load_workbook(f)
            sheet = loaded_wb["Histórico de Dietas Autorizadas"]
            assert (
                sheet["A2"].value
                == "Total de Dietas Autorizadas em 09/05/2025 para as unidades da DRE IPIRANGA: 120 | Data de extração do relatório: 09/05/2025"
            )

            assert sheet["C5"].value == "CEI DIRET HAROLDO"
            assert sheet["F5"].value == "07 a 11 meses"
            assert sheet["F6"].value == "07 a 11 meses"
            assert sheet["F7"].value == "01 ano a 03 anos e 11 meses"
            assert sheet["F8"].value == "01 ano a 03 anos e 11 meses"

            assert sheet["C9"].value == "CEMEI ALZIRA"
            assert sheet["F9"].value == "07 a 11 meses"
            assert sheet["F10"].value == "07 a 11 meses"
            assert sheet["F11"].value == "Infantil"
            assert sheet["F12"].value == "Infantil"
            assert sheet["F13"].value == "Infantil"
            assert sheet["F14"].value == "Infantil"

            assert sheet["C15"].value == "CEU GESTAO 9 DE JULHO"

            assert sheet["C17"].value == "EMEBS HELEN KELLER"
            assert sheet["F17"].value == "Infantil (4 a 6 anos)"
            assert sheet["F18"].value == "Fundamental (acima de 6 anos)"
            assert sheet["F19"].value == "Infantil (4 a 6 anos)"
            assert sheet["F20"].value == "Fundamental (acima de 6 anos)"
            assert sheet["F21"].value == "Infantil (4 a 6 anos)"
            assert sheet["F22"].value == "Fundamental (acima de 6 anos)"
            assert sheet["F23"].value == "Infantil (4 a 6 anos)"
            assert sheet["F24"].value == "Fundamental (acima de 6 anos)"

            assert sheet["C25"].value == "EMEF PERICLES"

    def test_gera_xlsx_historico_dietas_especiais_periodo_param(
        self, client_autenticado_vinculo_codae_gestao_alimentacao_dieta
    ):
        self.setup()
        client, user = client_autenticado_vinculo_codae_gestao_alimentacao_dieta
        nome_arquivo = "relatorio_historico_dietas_especiais.xlsx"
        data = json.dumps(
            {
                "lote": str(self.lote.uuid),
                "data": "09/05/2025",
                "periodos_escolares_selecionadas[]": str(self.periodo_integral.uuid),
            }
        )
        gera_xlsx_relatorio_historico_dietas_especiais_async(user, nome_arquivo, data)

        central_download = CentralDeDownload.objects.get()
        assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO

        with central_download.arquivo.open("rb") as f:
            loaded_wb = load_workbook(f)
            sheet = loaded_wb["Histórico de Dietas Autorizadas"]
            assert "Períodos: INTEGRAL" in sheet["A2"].value

import pytest
from freezegun.api import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.dados_comuns.models import CentralDeDownload
from sme_sigpae_api.dieta_especial.tasks import (
    gera_xlsx_relatorio_recreio_nas_ferias_async,
)

from .test_setup_relatorio_recreio import BaseSetupRecreioNasFerias

pytestmark = pytest.mark.django_db


def linha(sheet, row):
    return (
        sheet[f"A{row}"].value,
        sheet[f"B{row}"].value,
        sheet[f"C{row}"].value,
        sheet[f"D{row}"].value,
        sheet[f"E{row}"].value,
        sheet[f"G{row}"].value,
    )


@pytest.mark.usefixtures("client_autenticado_vinculo_codae_gestao_alimentacao_dieta")
@freeze_time("2025-09-04")
class TestGeraXlsxRelatorioRecreioNasFeriasAsync(BaseSetupRecreioNasFerias):

    def test_gera_xlsx_recreio_nas_ferias(
        self, client_autenticado_vinculo_codae_gestao_alimentacao_dieta
    ):
        self.setup()
        client, user = client_autenticado_vinculo_codae_gestao_alimentacao_dieta
        gera_xlsx_relatorio_recreio_nas_ferias_async(
            user, "relatorio_recreio_nas_ferias.pdf", {"lote": self.lote.uuid}
        )

        central_download = CentralDeDownload.objects.get()
        assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO

        with central_download.arquivo.open("rb") as f:
            loaded_wb = load_workbook(f)
            sheet = loaded_wb["Relatório de Recreio nas Férias"]
            assert (
                sheet["A1"].value
                == "Relatório de Dietas Autorizadas para Recreio nas Férias"
            )
            assert (
                sheet["A2"].value
                == "Total de Dietas Autorizadas: 3 | Para as Unidades da DRE/LOTE: IP - LOTE 01 | Data de extração do relatório: 04/09/2025"
            )

            linhas_obtidas = {
                linha(sheet, 5),
                linha(sheet, 6),
                linha(sheet, 7),
            }

            linhas_esperadas = {
                (
                    1,
                    "7654321 - JOÃO COSTA",
                    "EMEF PERICLES",
                    "EMEBS HELEN KELLER",
                    "Tipo B",
                    "DE 01/09/2025 ATÉ 29/09/2025",
                ),
                (
                    2,
                    "1234567 - MARIA SILVA",
                    "EMEBS HELEN KELLER",
                    "EMEF PERICLES",
                    "Tipo A",
                    "DE 01/08/2025 ATÉ 31/08/2025",
                ),
                (
                    3,
                    "Aluno não matriculado - GOHAN MENESES",
                    "-",
                    "EMEF PERICLES",
                    "Tipo A",
                    "DE 01/08/2025 ATÉ 31/08/2025",
                ),
            }

            assert linhas_obtidas == linhas_esperadas

    def test_gera_pdf_historico_dietas_especiais_periodo_param(
        self, client_autenticado_vinculo_codae_gestao_alimentacao_dieta
    ):
        self.setup()
        client, user = client_autenticado_vinculo_codae_gestao_alimentacao_dieta
        gera_xlsx_relatorio_recreio_nas_ferias_async(
            user,
            "relatorio_recreio_nas_ferias.pdf",
            {"lote": self.lote.uuid, "data_inicio": "01/09/2025"},
        )
        central_download = CentralDeDownload.objects.get()
        assert central_download.status == CentralDeDownload.STATUS_CONCLUIDO

        with central_download.arquivo.open("rb") as f:
            loaded_wb = load_workbook(f)
            sheet = loaded_wb["Relatório de Recreio nas Férias"]
            assert (
                sheet["A1"].value
                is not "Relatório de Dietas Autorizadas para Recreio nas Férias"
            )

            assert sheet["B6"].value is not "1234567 - MARIA SILVA"
            assert sheet["C6"].value is not "EMEBS HELEN KELLER"

            assert sheet["B5"].value == "7654321 - JOÃO COSTA"
            assert sheet["C5"].value == "EMEF PERICLES"

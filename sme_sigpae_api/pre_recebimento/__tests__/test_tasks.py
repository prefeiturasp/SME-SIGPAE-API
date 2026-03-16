import io
from datetime import date
from unittest.mock import MagicMock, call, patch

import pandas as pd
import pytest
from freezegun import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.pre_recebimento.cronograma_entrega.models import (
    InterrupcaoProgramadaEntrega,
)
from sme_sigpae_api.pre_recebimento.tasks import (
    _aplicar_estilos_leve_leite,
    _criar_linha_base_excel,
    _deve_mostrar_linha_a_receber,
    _processar_fichas_recebimento,
    importa_feriados_para_interrupcoes_programadas,
)


def test_criar_linha_base_excel():
    cronograma_mock = {
        "numero": "CR-001",
        "produto": "Arroz",
        "empresa": "Empresa X",
        "marca": "Marca Y",
        "qtd_total_programada": 1000,
        "custo_unitario_produto": "5,50",
        "armazem": "Armazém Central",
        "status": "ATIVO",
    }

    etapa_mock = {
        "etapa": "1",
        "parte": "1",
        "data_programada": "15/01/2024",
        "quantidade": "500,00 kg",
        "total_embalagens": "50",
        "unidade_medida": "kg",
    }

    result = _criar_linha_base_excel(cronograma_mock, etapa_mock)

    assert result["cronograma_numero"] == "CR-001"
    assert result["etapa"] == "1"
    assert result["parte"] == "1"
    assert "quantidade" in result


def test_processar_fichas_recebimento():
    linha_base = {"etapa": "1", "parte": "1", "situacao": ""}
    etapa = {"etapa": "1", "parte": "1"}
    fichas = [
        {"situacao": "Recebido", "houve_reposicao": False},
        {"situacao": "Reposição", "houve_reposicao": True},
    ]

    result = _processar_fichas_recebimento(linha_base, etapa, fichas)

    assert len(result) == 2
    assert result[0]["situacao"] == "Recebido"
    assert result[1]["situacao"] == "Reposição"
    assert "Reposição / Pagamento" in result[1]["etapa"]


def test_deve_mostrar_linha_a_receber(parametros_deve_mostrar_linha_a_receber):
    params = parametros_deve_mostrar_linha_a_receber
    etapa = {"foi_recebida": params["foi_recebida"]}
    filtros = (
        {"situacao": params["filtros_situacao"]} if params["filtros_situacao"] else {}
    )

    result = _deve_mostrar_linha_a_receber(etapa, params["fichas_recebimento"], filtros)
    assert result == params["expected"]


def test_aplicar_estilos_leve_leite_escreve_na_linha_correta():
    worksheet = MagicMock()
    workbook = MagicMock()
    workbook.add_format.return_value = "formato-azul"

    dados = [
        {"produto_nome": "ABACAXI"},
        {"produto_nome": "BANANA"},
        {"produto_nome": "LEITE"},
    ]

    _aplicar_estilos_leve_leite(worksheet, workbook, [0, 2], dados)

    worksheet.write.assert_has_calls(
        [
            call(1, 1, "ABACAXI", "formato-azul"),
            call(3, 1, "LEITE", "formato-azul"),
        ]
    )


def test_aplicar_estilos_leve_leite_nao_sobrescreve_coluna_produto_no_xlsx():
    produtos = [
        "ABACAXI DESIDRATADO PICADO 30G",
        "DOCE DE BANANA ORGANICO SEM ACUCAR",
        "FORMULA INFANTIL DE SEGUIMENTO 6 AO 12 MES",
        "FORMULA INFANTIL DE PARTIDA 0 AO 6 MES",
        "MACA DESIDRATADA CROCANTE",
    ]

    df = pd.DataFrame(
        {
            "cronograma_numero": [
                "001/2026A",
                "002/2026A",
                "003/2026A",
                "004/2026A",
                "005/2026A",
            ],
            "produto_nome": produtos,
        }
    )

    dados = [{"produto_nome": produto} for produto in produtos]
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(
            writer,
            "Relatório de Cronogramas",
            index=False,
            header=False,
            startrow=1,
        )
        workbook = writer.book
        worksheet = writer.sheets["Relatório de Cronogramas"]
        _aplicar_estilos_leve_leite(worksheet, workbook, [2, 3], dados)

    output.seek(0)
    workbook_lido = load_workbook(output)
    worksheet_lida = workbook_lido["Relatório de Cronogramas"]

    for indice, produto_esperado in enumerate(produtos):
        # startrow=1 => primeira linha de dados no Excel e a linha 2.
        assert worksheet_lida.cell(row=indice + 2, column=2).value == produto_esperado


@pytest.mark.django_db
class TestImportaFeriadosTask:

    @freeze_time("2024-05-20")
    @patch("sme_sigpae_api.pre_recebimento.tasks.BrazilSaoPauloCity")
    def test_importa_feriados_cria_registros_corretamente(self, mock_calendario_class):
        mock_calendario = MagicMock()
        mock_calendario_class.return_value = mock_calendario

        mock_calendario.holidays.side_effect = [
            [(date(2024, 1, 1), "New year")],
            [(date(2025, 1, 1), "New year")],
        ]

        importa_feriados_para_interrupcoes_programadas()

        assert InterrupcaoProgramadaEntrega.objects.count() == 4

        qs_2024 = InterrupcaoProgramadaEntrega.objects.filter(data=date(2024, 1, 1))
        assert qs_2024.count() == 2

        assert qs_2024.filter(descricao_motivo="Ano novo").count() == 2

        assert qs_2024.filter(
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL
        ).exists()
        assert qs_2024.filter(
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_PONTO_A_PONTO
        ).exists()

    @freeze_time("2024-05-20")
    @patch("sme_sigpae_api.pre_recebimento.tasks.BrazilSaoPauloCity")
    def test_importa_feriados_nao_cria_duplicados(self, mock_calendario_class):
        mock_calendario = MagicMock()
        mock_calendario_class.return_value = mock_calendario
        mock_calendario.holidays.side_effect = [
            [(date(2024, 1, 1), "New year")],
            [(date(2025, 1, 1), "New year")],
        ]

        InterrupcaoProgramadaEntrega.objects.create(
            data=date(2024, 1, 1),
            motivo=InterrupcaoProgramadaEntrega.MOTIVO_OUTROS,
            descricao_motivo="Outro motivo",
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL,
        )

        importa_feriados_para_interrupcoes_programadas()

        assert InterrupcaoProgramadaEntrega.objects.count() == 4

        existente = InterrupcaoProgramadaEntrega.objects.get(
            data=date(2024, 1, 1),
            tipo_calendario=InterrupcaoProgramadaEntrega.TIPO_CALENDARIO_ARMAZENAVEL,
        )
        assert existente.motivo == InterrupcaoProgramadaEntrega.MOTIVO_OUTROS

        novos_feriados = InterrupcaoProgramadaEntrega.objects.filter(
            motivo=InterrupcaoProgramadaEntrega.MOTIVO_FERIADO
        )
        assert novos_feriados.count() == 3

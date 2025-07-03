from io import BytesIO

import openpyxl
import pandas as pd
import pytest
from openpyxl import load_workbook

from sme_sigpae_api.medicao_inicial.services.relatorio_consolidado_excel import (
    _formata_filtros,
    _formata_total_geral,
    _preenche_linha_dos_filtros_selecionados,
    _preenche_titulo,
    gera_relatorio_consolidado_xlsx,
)

pytestmark = pytest.mark.django_db


def test_gera_relatorio_consolidado_xlsx_emef(
    relatorio_consolidado_xlsx_emef, mock_query_params_excel_emef
):
    solicitacoes = [relatorio_consolidado_xlsx_emef.uuid]
    tipos_unidade = ["EMEF"]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emef
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emef.mes}-{ relatorio_consolidado_xlsx_emef.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Lanche",
        "Lanche 4h",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        "EMEF",
        "123456",
        "EMEF TESTE",
        10,
        10,
        125,
        125,
        125,
        125,
        125,
        125,
        20,
        20,
        10,
        10,
        10,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        10,
        10,
        125,
        125,
        125,
        125,
        125,
        125,
        20,
        20,
        10,
        10,
        10,
    )


def test_gera_relatorio_consolidado_xlsx_emei(
    relatorio_consolidado_xlsx_emei, mock_query_params_excel_emei
):
    solicitacoes = [relatorio_consolidado_xlsx_emei.uuid]
    tipos_unidade = ["EMEI"]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emei
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emei.mes}-{ relatorio_consolidado_xlsx_emei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Lanche",
        "Lanche 4h",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        "EMEI",
        "987654",
        "EMEI TESTE",
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        40,
        40,
        20,
        20,
        20,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        150,
        150,
        150,
        150,
        150,
        150,
        40,
        40,
        20,
        20,
        20,
    )


def test_gera_relatorio_consolidado_xlsx_cei(
    relatorio_consolidado_xlsx_cei, mock_query_params_excel_cei
):
    solicitacoes = [relatorio_consolidado_xlsx_cei.uuid]
    tipos_unidade = ["CEI DIRET"]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_cei
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_cei.mes}-{ relatorio_consolidado_xlsx_cei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEI DIRET",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "PARCIAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        "TARDE",
        None,
        "DIETA ESPECIAL - TIPO A",
        "DIETA ESPECIAL - TIPO B",
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "04 a 05 meses",
        "08 a 11 meses",
        "06 a 07 meses",
        "02 anos a 03 anos e 11 meses",
        "04 a 05 meses",
        "06 a 07 meses",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        "CEI DIRET",
        "765432",
        "CEI DIRET TESTE",
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        60,
        60,
        80,
        8,
        4,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        80,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        8,
        80,
        60,
        60,
        80,
        8,
        4,
    )


def test_gera_relatorio_consolidado_xlsx_cemei(
    relatorio_consolidado_xlsx_cemei, mock_query_params_excel_cemei
):
    solicitacoes = [relatorio_consolidado_xlsx_cemei.uuid]
    tipos_unidade = ["CEMEI"]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_cemei
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_cemei.mes}-{ relatorio_consolidado_xlsx_cemei.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEMEI",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "PARCIAL",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "INFANTIL INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "INFANTIL MANHA",
        None,
        None,
        None,
        None,
        None,
        "INFANTIL TARDE",
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
    )
    assert rows[3] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Kit Lanche",
        "Lanche Emerg.",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "00 meses",
        "01 a 03 meses",
        "04 a 05 meses",
        "06 a 07 meses",
        "08 a 11 meses",
        "01 ano a 01 ano e 11 meses",
        "02 anos a 03 anos e 11 meses",
        "04 anos a 06 anos",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Lanche",
        "Lanche 4h",
    )
    assert rows[4] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[5] == (
        "CEMEI",
        "543210",
        "CEMEI TESTE",
        5,
        5,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        30,
        30,
        15,
        15,
        15,
    )
    assert rows[6] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        100,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        10,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        15,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        150,
        30,
        30,
        15,
        15,
        15,
    )


def test_gera_relatorio_consolidado_xlsx_emebs(
    relatorio_consolidado_xlsx_emebs, mock_query_params_excel_emebs
):
    solicitacoes = [relatorio_consolidado_xlsx_emebs.uuid]
    tipos_unidade = ["EMEBS"]
    arquivo = gera_relatorio_consolidado_xlsx(
        solicitacoes, tipos_unidade, mock_query_params_excel_emebs
    )
    assert isinstance(arquivo, bytes)
    excel_buffer = BytesIO(arquivo)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = f"Relatório Consolidado { relatorio_consolidado_xlsx_emebs.mes}-{ relatorio_consolidado_xlsx_emebs.ano}"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        "INFANTIL (4 a 6 anos)",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "FUNDAMENTAL (acima de 6 anos)",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        None,
        None,
        None,
        None,
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "TARDE",
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
        "MANHA",
        None,
        None,
        None,
        None,
        None,
        "TARDE",
        None,
        None,
        None,
        None,
        None,
        "INTEGRAL",
        None,
        None,
        None,
        None,
        None,
        "NOITE",
        None,
        None,
        None,
        None,
        None,
        "PROGRAMAS E PROJETOS",
        None,
        None,
        None,
        None,
        None,
        "DIETA ESPECIAL - TIPO A",
        None,
        None,
        "DIETA ESPECIAL - TIPO B",
        None,
    )
    assert rows[4] == (
        "Tipo",
        "Cód. EOL",
        "Unidade Escolar",
        "Lanche Emerg.",
        "Kit Lanche",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Lanche",
        "Lanche 4h",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Refeições p/ Pagamento",
        "Sobremesa",
        "Sobremesas p/ Pagamento",
        "Lanche",
        "Lanche 4h",
        "Refeição",
        "Lanche",
        "Lanche 4h",
    )
    assert rows[5] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[6] == (
        "EMEBS",
        "000329",
        "EMEBS TESTE",
        5,
        5,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        40,
        40,
        20,
        20,
        20,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        50,
        50,
        25,
        25,
        25,
    )
    assert rows[7] == (
        "TOTAL",
        None,
        None,
        5,
        5,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        40,
        40,
        20,
        20,
        20,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        350,
        50,
        50,
        25,
        25,
        25,
    )


def test_preenche_titulo(informacoes_excel_writer_emef):
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _preenche_titulo(workbook, worksheet, df.columns)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A1:P1"}
    assert {str(r) for r in merged_ranges} == esperados

    assert (
        sheet["A1"].value
        == "Relatório de Totalização da Medição Inicial do Serviço de Fornecimento da Alimentação Escolar"
    )
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A1"].alignment.vertical == "center"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.color.rgb == "FF42474A"
    assert sheet["A1"].fill.fgColor.rgb == "FFD6F2E7"
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emef(
    mock_query_params_excel_emef, informacoes_excel_writer_emef
):
    tipos_unidades = ["EMEF"]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _preenche_linha_dos_filtros_selecionados(
        workbook, worksheet, mock_query_params_excel_emef, df.columns, tipos_unidades
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A2:P2"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emei(
    mock_query_params_excel_emei, informacoes_excel_writer_emei
):
    tipos_unidades = ["EMEI"]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emei
    _preenche_linha_dos_filtros_selecionados(
        workbook, worksheet, mock_query_params_excel_emei, df.columns, tipos_unidades
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A2:P2"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_cei(
    mock_query_params_excel_cei, informacoes_excel_writer_cei
):
    tipos_unidades = ["CEI"]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_cei
    _preenche_linha_dos_filtros_selecionados(
        workbook, worksheet, mock_query_params_excel_cei, df.columns, tipos_unidades
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 10

    assert "A2:BE2" in str(merged_ranges)
    assert "A3:C3" in str(merged_ranges)
    assert "D3:K3" in str(merged_ranges)
    assert "L3:S3" in str(merged_ranges)
    assert "T3:AA3" in str(merged_ranges)
    assert "AB3:AI3" in str(merged_ranges)
    assert "AJ3:AQ3" in str(merged_ranges)
    assert "AR3:AY3" in str(merged_ranges)
    assert "AZ3:BA3" in str(merged_ranges)
    assert "BB3:BC3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_cemei(
    mock_query_params_excel_cemei, informacoes_excel_writer_cemei
):
    tipos_unidades = ["CEMEI"]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_cemei
    _preenche_linha_dos_filtros_selecionados(
        workbook, worksheet, mock_query_params_excel_cemei, df.columns, tipos_unidades
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 13

    assert "A2:BX2" in str(merged_ranges)
    assert "A3:E3" in str(merged_ranges)
    assert "F3:M3" in str(merged_ranges)
    assert "N3:U3" in str(merged_ranges)
    assert "V3:AC3" in str(merged_ranges)
    assert "AD3:AK3" in str(merged_ranges)
    assert "AL3:AS3" in str(merged_ranges)
    assert "AT3:BA3" in str(merged_ranges)
    assert "BB3:BG3" in str(merged_ranges)
    assert "BH3:BM3" in str(merged_ranges)
    assert "BN3:BS3" in str(merged_ranges)
    assert "BT3:BV3" in str(merged_ranges)
    assert "BW3:BX3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - CEMEI"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados_unidade_emebs(
    mock_query_params_excel_emebs, informacoes_excel_writer_emebs
):
    tipos_unidades = ["EMEBS"]
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emebs
    _preenche_linha_dos_filtros_selecionados(
        workbook, worksheet, mock_query_params_excel_emebs, df.columns, tipos_unidades
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 18

    assert "A2:BO2" in str(merged_ranges)
    assert "AF4:AG4" in str(merged_ranges)
    assert "AH4:AM4" in str(merged_ranges)
    assert "A4:E4" in str(merged_ranges)
    assert "BF4:BJ4" in str(merged_ranges)
    assert "BN4:BO4" in str(merged_ranges)
    assert "L4:Q4" in str(merged_ranges)
    assert "R4:W4" in str(merged_ranges)
    assert "AZ4:BE4" in str(merged_ranges)
    assert "F4:K4" in str(merged_ranges)
    assert "BK4:BM4" in str(merged_ranges)
    assert "AC4:AE4" in str(merged_ranges)
    assert "AN4:AS4" in str(merged_ranges)
    assert "AT4:AY4" in str(merged_ranges)
    assert "F3:AG3" in str(merged_ranges)
    assert "X4:AB4" in str(merged_ranges)
    assert "A3:E3" in str(merged_ranges)
    assert "AH3:BO3" in str(merged_ranges)

    assert sheet["A2"].value == "ABRIL/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FF0C6B45"
    assert sheet["A2"].fill.fgColor.rgb == "FFEAFFF6"

    rows = list(sheet.iter_rows(values_only=True))
    assert tipos_unidades[0] in rows[1][0]
    workbook_openpyxl.close()


def test_formata_total_geral(informacoes_excel_writer_emef):
    aba, writer, workbook, worksheet, df, arquivo = informacoes_excel_writer_emef
    _formata_total_geral(workbook, worksheet, df)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]
    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 5
    esperados = {"A3:E3", "F3:K3", "L3:N3", "O3:P3", "A7:C7"}
    assert {str(r) for r in merged_ranges} == esperados

    assert sheet["A7"].value == "TOTAL"
    assert sheet["A7"].alignment.horizontal == "center"
    assert sheet["A7"].alignment.vertical == "center"
    assert sheet["A7"].font.bold is True
    workbook_openpyxl.close()


def test_formata_filtros_unidade_emef(mock_query_params_excel_emef):
    tipos_unidades = ["EMEF"]
    filtros = _formata_filtros(mock_query_params_excel_emef, tipos_unidades)
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL IPIRANGA - 1 - EMEF"


def test_formata_filtros_unidade_emei(mock_query_params_excel_emei):
    tipos_unidades = ["EMEI"]
    filtros = _formata_filtros(mock_query_params_excel_emei, tipos_unidades)
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - EMEI"


def test_formata_filtros_unidade_cei(mock_query_params_excel_cei):
    tipos_unidades = ["CEI"]
    filtros = _formata_filtros(mock_query_params_excel_cei, tipos_unidades)
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - CEI"


def test_formata_filtros_unidade_cemei(mock_query_params_excel_cemei):
    tipos_unidades = ["CEMEI"]
    filtros = _formata_filtros(mock_query_params_excel_cemei, tipos_unidades)
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - CEMEI"


def test_formata_filtros_unidade_emebs(mock_query_params_excel_emebs):
    tipos_unidades = ["EMEBS"]
    filtros = _formata_filtros(mock_query_params_excel_emebs, tipos_unidades)
    assert isinstance(filtros, str)
    assert filtros == "Abril/2025 - DIRETORIA REGIONAL TESTE -  - EMEBS"


def test_gera_relatorio_consolidado_xlsx_tipo_unidade_invalida():
    tipos_de_unidade = ["aaa", "bbb"]
    with pytest.raises(ValueError, match=f"Unidades inválidas"):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {})


def test_gera_relatorio_consolidado_xlsx_retorna_exception():
    tipos_de_unidade = ["CEI"]
    with pytest.raises(Exception):
        gera_relatorio_consolidado_xlsx([], tipos_de_unidade, {})

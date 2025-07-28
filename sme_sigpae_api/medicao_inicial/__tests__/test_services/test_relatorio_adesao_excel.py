from io import BytesIO

import openpyxl
import pandas as pd
import pytest
from freezegun import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.medicao_inicial.services.relatorio_adesao_excel import (
    _ajusta_layout_header,
    _formata_filtros,
    _formata_numeros_linha_total,
    _insere_tabela_periodo_na_planilha,
    _preenche_data_do_relatorio,
    _preenche_linha_do_periodo,
    _preenche_linha_dos_filtros_selecionados,
    _preenche_titulo,
    gera_relatorio_adesao_xlsx,
)

pytestmark = pytest.mark.django_db


@freeze_time("2025-07-20")
def test_gera_relatorio_adesao_xlsx(mock_exportacao_relatorio_adesao):
    resultados, query_params = mock_exportacao_relatorio_adesao
    excel = gera_relatorio_adesao_xlsx(resultados, query_params)
    assert isinstance(excel, bytes)
    excel_buffer = BytesIO(excel)

    workbook = load_workbook(filename=excel_buffer)
    nome_aba = "Relatório de Adesão"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Adesão das Alimentações Servidas",
        None,
        None,
        None,
    )
    assert rows[1] == (
        "MARÇO - 2025 | DIRETORIA REGIONAL IPIRANGA | LOTE 01, LOTE 02, LOTE 03 | EMEF TESTE | PERÍODO DE LANÇAMENTO: 05/03/2025 ATÉ 15/03/2025",
        None,
        None,
        None,
    )
    assert rows[2] == ("Data: 20/07/2025", None, None, None)

    assert rows[3] == ("MANHA", None, None, None)
    assert rows[4] == (
        "Tipo de Alimentação",
        "Total de Alimentações Servidas",
        "Número Total de Frequência",
        "% de Adesão",
    )
    assert rows[5] == ("LANCHE", 140, 755, 0.1854)
    assert rows[6] == ("SOBREMESA", 140, 755, 0.1854)
    assert rows[7] == ("TOTAL", 280, 1510, 0.1854)
    assert rows[8] == (None, None, None, None)

    assert rows[9] == ("TARDE", None, None, None)
    assert rows[10] == (
        "Tipo de Alimentação",
        "Total de Alimentações Servidas",
        "Número Total de Frequência",
        "% de Adesão",
    )
    assert rows[11] == ("LANCHE", 130, 745, 0.1745)
    assert rows[12] == ("SOBREMESA", 250, 745, 0.3356)
    assert rows[13] == ("TOTAL", 380, 1490, 0.255)


def test_preenche_titulo(mock_exportacao_informacoes_excel_writer):
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    _preenche_titulo(workbook, worksheet, colunas)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 1
    assert "A1:D1" in str(merged_ranges)

    assert sheet["A1"].value == "Relatório de Adesão das Alimentações Servidas"
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A1"].alignment.vertical == "center"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb == "FFC1F2B0"
    assert len(sheet._images) == 1

    workbook_openpyxl.close()


def test_preenche_linha_dos_filtros_selecionados(
    mock_exportacao_informacoes_excel_writer, mock_exportacao_relatorio_adesao
):
    _, query_params = mock_exportacao_relatorio_adesao
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    _preenche_linha_dos_filtros_selecionados(workbook, worksheet, query_params, colunas)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 1
    assert "A2:D2" in str(merged_ranges)
    assert (
        sheet["A2"].value
        == "MARÇO - 2025 | DIRETORIA REGIONAL IPIRANGA | LOTE 01, LOTE 02, LOTE 03 | EMEF TESTE | PERÍODO DE LANÇAMENTO: 05/03/2025 ATÉ 15/03/2025"
    )
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["A2"].font.bold is False

    workbook_openpyxl.close()


def test_formata_filtros(mock_exportacao_relatorio_adesao):
    _, query_params = mock_exportacao_relatorio_adesao
    filtros = _formata_filtros(query_params)
    assert filtros == (
        "Março - 2025 | DIRETORIA REGIONAL IPIRANGA | Lote 01, Lote 02, Lote 03 | EMEF TESTE | "
        "Período de lançamento: 05/03/2025 até 15/03/2025"
    )


@freeze_time("2025-07-20")
def test_preenche_data_do_relatorio(mock_exportacao_informacoes_excel_writer):
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    _preenche_data_do_relatorio(workbook, worksheet, colunas)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 1
    assert "A3:D3" in str(merged_ranges)
    assert sheet["A3"].value == "Data: 20/07/2025"
    assert sheet["A3"].alignment.vertical == "center"
    assert sheet["A3"].font.bold is False

    workbook_openpyxl.close()


def test_insere_tabela_periodo_na_planilha(
    mock_exportacao_informacoes_excel_writer, mock_exportacao_relatorio_adesao
):
    colunas, aba, writer, _, _, _ = mock_exportacao_informacoes_excel_writer
    resultados, _ = mock_exportacao_relatorio_adesao
    refeicoes = resultados["MANHA"]
    proxima_linha = 4
    df = _insere_tabela_periodo_na_planilha(
        aba, refeicoes, colunas, proxima_linha, writer
    )
    assert isinstance(df, pd.DataFrame)
    colunas_df = df.columns.tolist()
    assert len(colunas_df) == 4
    colunas_df == [
        "Tipo de Alimentação",
        "Total de Alimentações Servidas",
        "Número Total de Frequência",
        "% de Adesão",
    ]
    df.iloc[0].tolist() == ["LANCHE", 140, 755, 0.1854]
    df.iloc[1].tolist() == ["SOBREMESA", 140, 755, 0.1854]
    df.iloc[2].tolist() == ["TOTAL", 280, 1510, 0.1854]


def test_preenche_linha_do_periodo(mock_exportacao_informacoes_excel_writer):
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    proxima_linha = 4
    periodo = "MANHA"
    _preenche_linha_do_periodo(
        workbook,
        worksheet,
        proxima_linha,
        periodo,
        colunas,
    )
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]
    merged_ranges = sheet.merged_cells.ranges
    assert len(merged_ranges) == 1
    assert "A4:D4" in str(merged_ranges)
    assert sheet["A4"].value == "MANHA"
    assert sheet["A4"].alignment.vertical == "center"
    assert sheet["A4"].font.bold is True
    assert sheet["A4"].font.color.rgb == "FF006400"

    workbook_openpyxl.close()


def test_ajusta_layout_header(
    mock_exportacao_informacoes_excel_writer, mock_exportacao_relatorio_adesao
):
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    resultados, _ = mock_exportacao_relatorio_adesao
    refeicoes = resultados["MANHA"]
    proxima_linha = 4
    df = _insere_tabela_periodo_na_planilha(
        aba, refeicoes, colunas, proxima_linha, writer
    )
    proxima_linha += len(df.index) + 1
    _ajusta_layout_header(workbook, worksheet, proxima_linha, df)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    assert sheet[f"A{proxima_linha}"].value == "TOTAL"

    workbook_openpyxl.close()


def test_formata_numeros_linha_total(
    mock_exportacao_informacoes_excel_writer, mock_exportacao_relatorio_adesao
):
    colunas, aba, writer, workbook, worksheet, arquivo = (
        mock_exportacao_informacoes_excel_writer
    )
    resultados, _ = mock_exportacao_relatorio_adesao
    refeicoes = resultados["MANHA"]
    proxima_linha = 4
    df = _insere_tabela_periodo_na_planilha(
        aba, refeicoes, colunas, proxima_linha, writer
    )
    proxima_linha += len(df.index) + 1
    _formata_numeros_linha_total(workbook, worksheet, proxima_linha, colunas, df)
    writer.close()
    workbook_openpyxl = openpyxl.load_workbook(arquivo)
    sheet = workbook_openpyxl[aba]

    columns = sheet[proxima_linha]
    assert columns[0].number_format == "#,##0"
    assert columns[1].number_format == "#,##0"
    assert columns[2].number_format == "#,##0"
    assert columns[3].number_format == "0.00%"

    workbook_openpyxl.close()

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from sme_sigpae_api.produto.api.serializers.serializers import (
    ReclamacaoDeProdutoExcelSerializer,
)
from sme_sigpae_api.produto.utils.relatorio_reclamacao_produto import (
    _extrair_dados_reclamacao,
    _gerar_subtitulo,
    build_xlsx_reclamacao,
    gerar_relatorio_reclamacao_produto_excel,
)


def test_extrair_dados_reclamacao(reclamacao_produto_query_excel):
    reclamacoes = ReclamacaoDeProdutoExcelSerializer(
        reclamacao_produto_query_excel, many=True
    ).data
    dado = _extrair_dados_reclamacao(reclamacoes[0], index=1)
    assert dado["Nº"] == 1
    assert dado["Edital"] == "Edital de Pregão nº 41/sme/2017"
    assert dado["Nome do Produto"] == "Produto1"
    assert dado["Marca"] == "Marca1"
    assert dado["Fabricante"] == "Fabricante1"
    assert dado["Status do Produto"] == "RASCUNHO"
    assert dado["Status da Reclamação"] == "Aguardando resposta da terceirizada"
    assert "Nº da Reclamação" in dado
    assert "RF e Nome do Reclamante" in dado
    assert "Data da Reclamação" in dado
    assert "DRE/LOTE" in dado
    assert "Empresa" in dado
    assert "Cód EOL e Nome da Escola" in dado


def test_gerar_subtitulo_com_datas():
    filtros = {
        "data_inicial_reclamacao": "01/01/2022",
        "data_final_reclamacao": "01/09/2025",
    }
    resultado = _gerar_subtitulo(filtros, 3)
    assert (
        "Total de Reclamações de produtos para os editais selecionados: 3" in resultado
    )
    assert "Período das Reclamações: De 01/01/2022 a 01/09/2025" in resultado
    assert (
        f"Data de Extração do Relatório: {datetime.now().date().strftime('%d/%m/%Y')}"
        in resultado
    )


def test_gerar_subtitulo_somente_data_inicial():
    filtros = {
        "data_inicial_reclamacao": "01/01/2022",
    }
    resultado = _gerar_subtitulo(filtros, 3)
    assert (
        "Total de Reclamações de produtos para os editais selecionados: 3" in resultado
    )
    assert "Período das Reclamações: A partir de 01/01/2022" in resultado
    assert (
        f"Data de Extração do Relatório: {datetime.now().date().strftime('%d/%m/%Y')}"
        in resultado
    )


def test_gerar_subtitulo_somente_data_final():
    filtros = {"data_final_reclamacao": "01/09/2025"}
    resultado = _gerar_subtitulo(filtros, 3)
    assert (
        "Total de Reclamações de produtos para os editais selecionados: 3" in resultado
    )
    assert "Período das Reclamações: Até 01/09/2025" in resultado
    assert (
        f"Data de Extração do Relatório: {datetime.now().date().strftime('%d/%m/%Y')}"
        in resultado
    )


def test_gerar_subtitulo_sem_datas():
    resultado = _gerar_subtitulo({}, 3)
    assert (
        "Total de Reclamações de produtos para os editais selecionados: 3" in resultado
    )
    assert (
        f"Data de Extração do Relatório: {datetime.now().date().strftime('%d/%m/%Y')}"
        in resultado
    )
    assert "Período das Reclamações" not in resultado


def test_gerar_relatorio_reclamacao_produto_excel(reclamacao_produto_query_excel):
    reclamacoes = ReclamacaoDeProdutoExcelSerializer(
        reclamacao_produto_query_excel, many=True
    ).data
    quantidade_reclamacoes = reclamacao_produto_query_excel.count()
    filtros = {
        "data_inicial_reclamacao": "01/01/2022",
        "data_final_reclamacao": "01/09/2025",
    }
    output = gerar_relatorio_reclamacao_produto_excel(
        reclamacoes, quantidade_reclamacoes, filtros
    )

    assert isinstance(output, io.BytesIO)
    workbook = load_workbook(filename=output)
    nome_aba = f"Relatório Reclamação Produto"
    assert nome_aba in workbook.sheetnames
    sheet = workbook[nome_aba]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório de Acompanhamento de Reclamação de Produtos",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[1] == (
        f"Total de Reclamações de produtos para os editais selecionados: 3 | Período das Reclamações: De 01/01/2022 a 01/09/2025 | Data de Extração do Relatório: {datetime.now().date().strftime('%d/%m/%Y')}",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Nº",
        "Edital",
        "DRE/LOTE",
        "Empresa",
        "Nome do Produto",
        "Marca",
        "Fabricante",
        "Status do Produto",
        "Nº da Reclamação",
        "RF e Nome do Reclamante",
        "Cód EOL e Nome da Escola",
        "Status da Reclamação",
        "Data da Reclamação",
    )


def test_build_xlsx_reclamacao(reclamacao_produto_query_excel):
    titulo = "Tilulo 001"
    subtitulo = "Subtitulo 001"
    colunas = [
        "Nº",
        "Edital",
        "DRE/LOTE",
        "Empresa",
        "Nome do Produto",
        "Marca",
        "Fabricante",
        "Status do Produto",
        "Nº da Reclamação",
        "RF e Nome do Reclamante",
        "Cód EOL e Nome da Escola",
        "Status da Reclamação",
        "Data da Reclamação",
    ]
    reclamacoes = ReclamacaoDeProdutoExcelSerializer(
        reclamacao_produto_query_excel, many=True
    ).data

    output = io.BytesIO()
    build_xlsx_reclamacao(output, reclamacoes, titulo, subtitulo, colunas)
    output.seek(0)

    df = pd.read_excel(output, sheet_name="Relatório Reclamação Produto", skiprows=3)
    assert len(df) == 3
    assert list(df.columns) == [
        "Nº",
        "Edital",
        "DRE/LOTE",
        "Empresa",
        "Nome do Produto",
        "Marca",
        "Fabricante",
        "Status do Produto",
        "Nº da Reclamação",
        "RF e Nome do Reclamante",
        "Cód EOL e Nome da Escola",
        "Status da Reclamação",
        "Data da Reclamação",
    ]
    df["Edital"].iloc[0] == "Edital de Pregão nº 41/sme/2017"
    df["Edital"].iloc[1] == "Edital de Pregão nº 78/sme/2016"
    df["Edital"].iloc[2] == "Edital de Pregão nº 78/sme/2022"

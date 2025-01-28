import io

import pytest
from freezegun import freeze_time
from openpyxl import Workbook, load_workbook

from sme_sigpae_api.escola.utils_montar_planilha_matriculados import (
    build_xlsx_alunos_matriculados,
    monta_celulas_faixas_periodos,
    montar_cabecalho_padrao,
    montar_faixas_etarias,
    style_range,
)

pytestmark = pytest.mark.django_db


@freeze_time("2025-01-28")
def test_build_xlsx_alunos_matriculados(dados_planilha_alunos_matriculados):
    dados = dados_planilha_alunos_matriculados

    output = io.BytesIO()
    build_xlsx_alunos_matriculados(dados, None, output)

    workbook = load_workbook(output)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório SIGPAE - Alunos Matriculados",
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Solicitado por Faker usuario, impresso em 28/01/2025",
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[4] == (
        "DRE",
        "Lote",
        "Tipo de unid.",
        "Unid. educacional",
        "Tipo de Turma",
        "Período",
        "Matriculados",
    )

    queryset = dados["queryset"][0]
    if queryset["tipo_unidade"] == "":
        tipo_unidade = None

    assert rows[5] == (
        queryset["dre"],
        queryset["lote"],
        tipo_unidade,
        queryset["escola"],
        queryset["tipo_turma"],
        queryset["periodo_escolar"],
        queryset["matriculados"],
    )


@freeze_time("2025-01-28")
def test_build_xlsx_alunos_matriculados_cei_cemei(
    dados_planilha_alunos_matriculados_cei_cemei,
):
    dados = dados_planilha_alunos_matriculados_cei_cemei

    output = io.BytesIO()
    build_xlsx_alunos_matriculados(dados, None, output)

    workbook = load_workbook(output)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Relatório SIGPAE - Alunos Matriculados",
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[3] == (
        "Solicitado por Faker usuario, impresso em 28/01/2025",
        None,
        None,
        None,
        None,
        None,
        None,
    )
    assert rows[4] == (
        "DRE",
        "Lote",
        "Tipo de unid.",
        "Unid. educacional",
        "Tipo de Turma",
        "Período",
        "Matriculados",
    )

    queryset = dados["queryset"][0]
    if queryset["tipo_unidade"] == "":
        tipo_unidade = None

    assert rows[5] == (
        queryset["dre"],
        queryset["lote"],
        tipo_unidade,
        queryset["escola"],
        queryset["tipo_turma"],
        queryset["periodo_escolar"],
        queryset["matriculados"],
    )


def test_montar_cabecalho_padrao():
    wb = Workbook()
    ws = wb.active
    montar_cabecalho_padrao(2, ws)

    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (None, None, None, None, None, None, None)
    assert rows[1] == (
        "DRE",
        "Lote",
        "Tipo de unid.",
        "Unid. educacional",
        "Tipo de Turma",
        "Período",
        "Matriculados",
    )
    with pytest.raises(IndexError):
        assert rows[2] == (None, None, None, None, None, None, None)


def test_montar_faixas_etarias(dados_planilha_alunos_matriculados):
    dados = dados_planilha_alunos_matriculados
    dado = dados["queryset"][0]
    faixas_etarias = dados["faixas_etarias"]

    wb = Workbook()
    ws = wb.active
    montar_faixas_etarias(ws, 0, dado, faixas_etarias)

    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "Faixa Etária",
        None,
        None,
        None,
        None,
        "Alunos Matriculados",
        None,
    )
    assert rows[1] == (faixas_etarias[0]["nome"], None, None, None, None, 0, None)
    for indice in range(2, 10):
        assert rows[indice] == (None, None, None, None, None, None, None)
    assert rows[10] == ("Total", None, None, None, None, dado["matriculados"], None)


def test_monta_celulas_faixas_periodos(dados_planilha_alunos_matriculados):
    dados = dados_planilha_alunos_matriculados
    dado = dados["queryset"][0]
    faixas_etarias = dados["faixas_etarias"]

    wb = Workbook()
    ws = wb.active

    monta_celulas_faixas_periodos(faixas_etarias, 0, ws, dado)
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (None, None, None, None, None, None)
    assert rows[1] == (faixas_etarias[0]["nome"], None, None, None, None, 0)
    with pytest.raises(IndexError):
        assert rows[2] == (None, None, None, None, None, None, None)


def test_style_range():
    wb = Workbook()
    ws = wb.active
    for cell_range in ws.merged_cells.ranges:
        style_range(ws, str(cell_range))

    rows = list(ws.iter_rows(values_only=True))
    assert isinstance(rows, list)
    assert len(rows) == 0

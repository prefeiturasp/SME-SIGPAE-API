import datetime
import os
from pathlib import Path
from unittest import mock
from unittest.mock import patch

import pandas as pd
import pytest
from django.core.exceptions import ObjectDoesNotExist
from freezegun import freeze_time
from openpyxl import load_workbook

from sme_sigpae_api.escola.models import (
    AlunoPeriodoParcial,
    AlunosMatriculadosPeriodoEscola,
    Codae,
    DiaCalendario,
    EscolaPeriodoEscolar,
    LogAlunosMatriculadosPeriodoEscola,
    Lote,
)
from sme_sigpae_api.escola.services import NovoSGPServico
from sme_sigpae_api.escola.utils import calendario_sgp

from ..utils import (
    alunos_por_faixa_append,
    analise_alunos_dietas_somente_uma_data,
    create_update_objeto_escola_periodo_escolar,
    cria_arquivo_excel,
    deletar_alunos_periodo_parcial_outras_escolas,
    dias_append,
    duplica_dia_anterior,
    eh_dia_sem_atividade_escolar,
    eh_mes_atual,
    faixa_to_string,
    get_alunos_com_dietas_autorizadas,
    lotes_endpoint_filtrar_relatorio_alunos_matriculados,
    meses_to_mes_e_ano_string,
    ordenar_alunos_matriculados,
    processa_dias_letivos,
    registra_quantidade_matriculados,
    registro_quantidade_alunos_matriculados_por_escola_periodo,
    remove_acentos,
    string_to_faixa,
    string_to_meses,
    trata_filtro_data_relatorio_controle_frequencia_pdf,
    update_datetime_LogAlunosMatriculadosPeriodoEscola,
)

pytestmark = pytest.mark.django_db


def test_meses_para_mes_e_ano_string():
    assert meses_to_mes_e_ano_string(0) == "00 meses"
    assert meses_to_mes_e_ano_string(1) == "01 mês"
    assert meses_to_mes_e_ano_string(2) == "02 meses"
    assert meses_to_mes_e_ano_string(3) == "03 meses"
    assert meses_to_mes_e_ano_string(11) == "11 meses"
    assert meses_to_mes_e_ano_string(12) == "01 ano"
    assert meses_to_mes_e_ano_string(13) == "01 ano e 01 mês"
    assert meses_to_mes_e_ano_string(14) == "01 ano e 02 meses"
    assert meses_to_mes_e_ano_string(15) == "01 ano e 03 meses"
    assert meses_to_mes_e_ano_string(23) == "01 ano e 11 meses"
    assert meses_to_mes_e_ano_string(24) == "02 anos"
    assert meses_to_mes_e_ano_string(25) == "02 anos e 01 mês"
    assert meses_to_mes_e_ano_string(26) == "02 anos e 02 meses"
    assert meses_to_mes_e_ano_string(27) == "02 anos e 03 meses"
    assert meses_to_mes_e_ano_string(35) == "02 anos e 11 meses"
    assert meses_to_mes_e_ano_string(36) == "03 anos"


def test_faixa_to_string():
    assert faixa_to_string(0, 0) == "0 meses a 11 meses"
    assert faixa_to_string(12, 13) == "01 ano"

    assert faixa_to_string(2, 62) == "02 a 05 anos e 01 mês"
    assert faixa_to_string(24, 62) == "02 anos a 05 anos e 01 mês"

    assert faixa_to_string(36, 72) == "03 anos a 06 anos"

    assert faixa_to_string(16, 51) == "01 ano e 04 meses a 04 anos e 02 meses"


def test_string_to_faixa():
    assert string_to_faixa("0 meses a 11 meses") == (0, 12)
    assert string_to_faixa("1") == (1, 2)

    assert string_to_faixa("3 a 5") == (3, 6)
    assert string_to_faixa("3 anos  a 5 anos ") == (36, 61)

    assert string_to_faixa("02 a 05 ") == (2, 6)
    assert string_to_faixa("2 a 5 anos e 1 mês") == (2, 62)
    assert string_to_faixa("02 anos a 05 anos e 01 mês") == (24, 62)


def test_string_to_meses():
    assert string_to_meses("1 ano") == 12
    assert string_to_meses("2 ano") == 24
    assert string_to_meses("3 ano") == 36
    assert string_to_meses("4 ano") == 48
    assert string_to_meses("5 ano") == 60
    assert string_to_meses("6 ano") == 72

    assert string_to_meses("1") == 1

    assert string_to_meses("1 mês") == 1
    assert string_to_meses("2 meses") == 2
    assert string_to_meses("3 meses") == 3
    assert string_to_meses("7 meses") == 7
    assert string_to_meses("11 meses") == 11


def test_remove_acentos():
    assert remove_acentos("àáâãäå") == "aaaaaa"
    assert remove_acentos("èéêë") == "eeee"
    assert remove_acentos("ìíîï") == "iiii"
    assert remove_acentos("abc") == "abc"
    assert remove_acentos("") == ""
    assert remove_acentos("Olá, você está bem?") == "Ola, voce esta bem?"
    assert remove_acentos("voô") == "voô"


@freeze_time("2025-02-05")
def test_update_datetime_log_alunos_matriculados_periodo_escola(
    update_log_alunos_matriculados,
):
    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 2
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 5)
        ).count()
        == 1
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 4)
        ).count()
        == 0
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 1)
        ).count()
        == 1
    )

    update_datetime_LogAlunosMatriculadosPeriodoEscola()
    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 2
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 5)
        ).count()
        == 0
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 4)
        ).count()
        == 1
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 1)
        ).count()
        == 1
    )


def test_registra_quantidade_matriculados(dicionario_de_alunos_matriculados):
    ontem = datetime.date(2025, 2, 5)
    tipo_turma = "REGULAR"

    assert AlunosMatriculadosPeriodoEscola.objects.count() == 3
    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 0

    registra_quantidade_matriculados(
        dicionario_de_alunos_matriculados, ontem, tipo_turma
    )
    assert AlunosMatriculadosPeriodoEscola.objects.count() == 9
    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 8


def test_create_update_objeto_escola_periodo_escolar(escola_cemei, periodo_escolar):
    quantidade_alunos = 100

    with pytest.raises(ObjectDoesNotExist):
        EscolaPeriodoEscolar.objects.get(
            periodo_escolar=periodo_escolar, escola=escola_cemei
        )

    create_update_objeto_escola_periodo_escolar(
        escola_cemei, periodo_escolar, quantidade_alunos
    )

    epe = EscolaPeriodoEscolar.objects.get(
        periodo_escolar=periodo_escolar, escola=escola_cemei
    )
    assert epe.quantidade_alunos == quantidade_alunos
    assert epe.escola == escola_cemei
    assert epe.periodo_escolar == periodo_escolar


@freeze_time("2025-02-07")
def test_duplica_dia_anterior(update_log_alunos_matriculados):
    periodo_escolar_regular, periodo_escolar_programas = update_log_alunos_matriculados

    dre = periodo_escolar_regular.escola.diretoria_regional
    tipo_turma_name = periodo_escolar_regular.tipo_turma
    dois_dias_atras = datetime.date(2025, 2, 5)
    ontem = datetime.date(2025, 2, 6)
    hoje = datetime.date(2025, 2, 7)

    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 2
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=dois_dias_atras
        ).count()
        == 1
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(criado_em__date=ontem).count()
        == 0
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(criado_em__date=hoje).count()
        == 0
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 1)
        ).count()
        == 1
    )
    duplica_dia_anterior(dre, dois_dias_atras, ontem, tipo_turma_name)

    assert LogAlunosMatriculadosPeriodoEscola.objects.count() == 3
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=dois_dias_atras
        ).count()
        == 1
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(criado_em__date=ontem).count()
        == 1
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(criado_em__date=hoje).count()
        == 0
    )
    assert (
        LogAlunosMatriculadosPeriodoEscola.objects.filter(
            criado_em__date=datetime.date(2025, 2, 1)
        ).count()
        == 1
    )


@patch("sme_sigpae_api.escola.models.DiretoriaRegional.objects.all")
@patch("sme_sigpae_api.eol_servico.utils.EOLServicoSGP.matricula_por_escola")
@patch("sme_sigpae_api.escola.utils.registra_quantidade_matriculados")
@patch("sme_sigpae_api.escola.utils.duplica_dia_anterior")
@patch("sme_sigpae_api.escola.utils.logger")
def test_registro_quantidade_alunos_matriculados_por_escola_periodo_sucesso(
    mock_logger,
    mock_duplica_dia_anterior,
    mock_registra_quantidade_matriculados,
    mock_matricula_por_escola,
    mock_diretoria_all,
    mock_diretoria_regional,
    mock_tipo_turma,
):
    mock_diretoria_all.return_value = mock_diretoria_regional
    mock_matricula_por_escola.return_value = {
        "escola": "Escola Teste",
        "matriculados": 50,
    }

    registro_quantidade_alunos_matriculados_por_escola_periodo(mock_tipo_turma)

    mock_matricula_por_escola.assert_called_once()
    mock_registra_quantidade_matriculados.assert_called_once()
    mock_duplica_dia_anterior.assert_not_called()
    mock_logger.error.assert_not_called()


@patch("sme_sigpae_api.escola.models.DiretoriaRegional.objects.all")
@patch(
    "sme_sigpae_api.eol_servico.utils.EOLServicoSGP.matricula_por_escola",
    side_effect=Exception(
        "Erro no serviço; as quantidades de alunos foram duplicadas do dia anterio"
    ),
)
@patch("sme_sigpae_api.escola.utils.registra_quantidade_matriculados")
@patch("sme_sigpae_api.escola.utils.duplica_dia_anterior")
@patch("sme_sigpae_api.escola.utils.logger")
def test_registro_quantidade_alunos_matriculados_por_escola_periodo_exception(
    mock_logger,
    mock_duplica_dia_anterior,
    mock_registra_quantidade_matriculados,
    mock_matricula_por_escola,
    mock_diretoria_all,
    mock_diretoria_regional,
    mock_tipo_turma,
):
    mock_diretoria_all.return_value = mock_diretoria_regional
    registro_quantidade_alunos_matriculados_por_escola_periodo(mock_tipo_turma)

    mock_matricula_por_escola.assert_called_once()
    mock_duplica_dia_anterior.assert_called_once()
    mock_registra_quantidade_matriculados.assert_not_called()
    mock_logger.error.assert_called_once()


def test_processa_dias_letivos(lista_dias_letivos):
    escola, dias_letivos = lista_dias_letivos
    assert DiaCalendario.objects.count() == 2
    assert DiaCalendario.objects.filter(dia_letivo=True).count() == 1
    assert DiaCalendario.objects.filter(dia_letivo=False).count() == 1
    processa_dias_letivos(dias_letivos, escola)
    assert DiaCalendario.objects.count() == 3
    assert DiaCalendario.objects.filter(dia_letivo=True).count() == 1
    assert DiaCalendario.objects.filter(dia_letivo=False).count() == 2


def test_calendario_sgp(mock_escolas):
    with patch(
        "sme_sigpae_api.escola.models.Escola.objects.all", return_value=mock_escolas
    ):
        with patch.object(
            NovoSGPServico, "dias_letivos", return_value={"dias": 180}
        ) as mock_dias_letivos:
            with patch(
                "sme_sigpae_api.escola.utils.processa_dias_letivos"
            ) as mock_processa_dias_letivos:
                calendario_sgp()
                escola_mock = mock_escolas[0]
                hoje = datetime.date.today()
                data_inicio = hoje.strftime("%Y-%m-%d")
                data_final = (hoje + pd.DateOffset(months=3)).date()
                data_fim = data_final.strftime("%Y-%m-%d")
                mock_dias_letivos.assert_any_call(
                    codigo_eol=escola_mock.codigo_eol,
                    data_inicio=data_inicio,
                    data_fim=data_fim,
                )
                mock_processa_dias_letivos.assert_called_once_with(
                    {"dias": 180}, escola_mock
                )


def test_calendario_sgp_exception(mock_escolas):
    with patch(
        "sme_sigpae_api.escola.models.Escola.objects.all", return_value=mock_escolas
    ):
        with patch.object(NovoSGPServico, "dias_letivos") as mock_dias_letivos:
            mock_dias_letivos.side_effect = Exception("Erro ao buscar dias letivos")
            with patch(
                "sme_sigpae_api.escola.utils.processa_dias_letivos"
            ) as mock_processa_dias_letivos:
                calendario_sgp()
                mock_dias_letivos.assert_any_call(
                    codigo_eol="12345", data_inicio=mock.ANY, data_fim=mock.ANY
                )
                mock_dias_letivos.side_effect = Exception(
                    "Erro ao buscar dias letivos no turno da noite"
                )
                calendario_sgp()
                mock_dias_letivos.assert_any_call(
                    codigo_eol="12345",
                    data_inicio=mock.ANY,
                    data_fim=mock.ANY,
                    tipo_turno=3,
                )
                mock_processa_dias_letivos.assert_not_called()


def test_lotes_endpoint_filtrar_relatorio_alunos_matriculados(
    usuario_coordenador_codae, lote
):
    instituicao = usuario_coordenador_codae.vinculo_atual.instituicao
    lotes = lotes_endpoint_filtrar_relatorio_alunos_matriculados(
        instituicao, Codae, Lote
    )
    assert lotes.count() == 2


def test_deletar_alunos_periodo_parcial_outras_escolas(excluir_alunos_periodo_parcial):
    escola, data_referencia = excluir_alunos_periodo_parcial

    assert AlunoPeriodoParcial.objects.count() == 3
    deletar_alunos_periodo_parcial_outras_escolas(escola, data_referencia)
    assert AlunoPeriodoParcial.objects.count() == 2


def test_eh_dia_sem_atividade_escolar(
    dia_calendario_letivo, dia_calendario_nao_letivo, escola, alteracao_cardapio
):
    resposta = eh_dia_sem_atividade_escolar(
        escola, dia_calendario_letivo.data, alteracao_cardapio
    )
    assert resposta is False

    resposta = eh_dia_sem_atividade_escolar(
        escola, dia_calendario_nao_letivo.data, alteracao_cardapio
    )
    assert resposta is True

    resposta = eh_dia_sem_atividade_escolar(
        escola, datetime.date(2025, 2, 10), alteracao_cardapio
    )
    assert resposta is False


def test_analise_alunos_dietas_somente_uma_data(dieta_codae_autorizou):
    alunos_com_dietas_autorizadas = []

    datetime_autorizacao = datetime.datetime(2025, 1, 1)
    data_inicial = "2025-01-10"
    data_final = "2025-03-01"
    alunos = analise_alunos_dietas_somente_uma_data(
        datetime_autorizacao,
        data_inicial,
        data_final,
        dieta_codae_autorizou,
        alunos_com_dietas_autorizadas,
    )
    assert isinstance(alunos, list)
    assert len(alunos) == 0


def test_analise_alunos_dietas_somente_uma_data_inicio_fim_iguais(
    dieta_codae_autorizou,
):
    alunos_com_dietas_autorizadas = []
    datetime_autorizacao = datetime.datetime(2025, 1, 1)
    data_inicial = "2025-03-01"
    data_final = "2025-03-01"
    alunos = analise_alunos_dietas_somente_uma_data(
        datetime_autorizacao,
        data_inicial,
        data_final,
        dieta_codae_autorizou,
        alunos_com_dietas_autorizadas,
    )
    assert isinstance(alunos, list)
    assert len(alunos) == 1
    assert alunos[0]["aluno"] == dieta_codae_autorizou.aluno.nome
    assert alunos[0]["tipo_dieta"] == dieta_codae_autorizou.classificacao.nome
    assert alunos[0]["data_autorizacao"] == dieta_codae_autorizou.data_autorizacao


def test_analise_alunos_dietas_somente_uma_data_dieta_cancelada(dieta_cancelada):
    alunos_com_dietas_autorizadas = []
    datetime_autorizacao = datetime.datetime(2025, 1, 1)
    data_inicial = "2025-02-01"
    data_final = "2025-02-01"
    alunos = analise_alunos_dietas_somente_uma_data(
        datetime_autorizacao,
        data_inicial,
        data_final,
        dieta_cancelada,
        alunos_com_dietas_autorizadas,
    )
    assert isinstance(alunos, list)
    assert len(alunos) == 1
    assert alunos[0]["aluno"] == dieta_cancelada.aluno.nome
    assert alunos[0]["tipo_dieta"] == dieta_cancelada.classificacao.nome
    assert alunos[0]["data_autorizacao"] == dieta_cancelada.data_autorizacao


@freeze_time("2025-01-01")
def test_get_alunos_com_dietas_autorizadas_com_datas(dieta_codae_autorizou, escola):
    data_inicial = "2025-01-01"
    data_final = "2025-01-01"
    query_params = {"data_inicial": data_inicial, "data_final": data_final}
    alunos = get_alunos_com_dietas_autorizadas(query_params, escola)
    assert len(alunos) == 1
    assert alunos[0]["aluno"] == dieta_codae_autorizou.aluno.nome
    assert alunos[0]["tipo_dieta"] == dieta_codae_autorizou.classificacao.nome
    assert alunos[0]["data_autorizacao"] == dieta_codae_autorizou.data_autorizacao


@freeze_time("2025-01-01")
def test_get_alunos_com_dietas_autorizadas_com_mes_ano(dieta_codae_autorizou, escola):
    query_params = {"mes_ano": "02_2025"}
    alunos = get_alunos_com_dietas_autorizadas(query_params, escola)
    assert isinstance(alunos, list)
    assert len(alunos) == 1
    assert alunos[0]["aluno"] == dieta_codae_autorizou.aluno.nome
    assert alunos[0]["tipo_dieta"] == dieta_codae_autorizou.classificacao.nome
    assert alunos[0]["data_autorizacao"] == dieta_codae_autorizou.data_autorizacao


@freeze_time("2024-06-15")
def test_trata_filtro_data_relatorio_mes_futuro():
    filtros = {}
    query_params = {}

    mes_seguinte = trata_filtro_data_relatorio_controle_frequencia_pdf(
        filtros, query_params, "2024", "07", 31
    )
    assert filtros["data"] == "2024-6-14"
    assert mes_seguinte is True


@freeze_time("2024-06-15")
def test_trata_filtro_data_relatorio_mes_atual_data_unica():
    filtros = {}
    query_params = {"data_inicial": "2024-06-15", "data_final": "2024-06-15"}
    mes_seguinte = trata_filtro_data_relatorio_controle_frequencia_pdf(
        filtros, query_params, "2024", "06", 30
    )
    assert filtros["data"] == datetime.date(2024, 6, 14)
    assert mes_seguinte is False


@freeze_time("2024-06-15")
def test_trata_filtro_data_relatorio_mes_atual_intervalo():
    filtros = {}
    query_params = {"data_inicial": "2024-06-16", "data_final": "2024-06-20"}
    mes_seguinte = trata_filtro_data_relatorio_controle_frequencia_pdf(
        filtros, query_params, "2024", "06", 30
    )
    assert filtros["data"] == datetime.date(2024, 6, 14)
    assert mes_seguinte is False


@freeze_time("2024-06-15")
def test_trata_filtro_data_relatorio_mes_anterior():
    filtros = {}
    query_params = {"data_inicial": "2024-05-10", "data_final": "2024-05-20"}
    mes_seguinte = trata_filtro_data_relatorio_controle_frequencia_pdf(
        filtros, query_params, "2024", "05", 31
    )
    assert filtros["data__gte"] == "2024-05-10"
    assert filtros["data__lte"] == "2024-05-20"
    assert mes_seguinte is False


@freeze_time("2025-02-12")
def test_mes_atual_com_mes_correto():
    query_params = {"mes_ano": "02_2025"}
    assert eh_mes_atual(query_params) is True


@freeze_time("2025-02-12")
def test_mes_atual_com_mes_errado():
    query_params = {"mes_ano": "03_2025"}
    assert eh_mes_atual(query_params) is False


@freeze_time("2025-02-12")
def test_mes_atual_com_mes_ausente():
    query_params = {}
    with pytest.raises(AttributeError):
        assert eh_mes_atual(query_params) is False


def test_alunos_por_faixa():
    alunos_por_faixa = ["João", "Maria"]
    aluno = "Carlos"

    resultado = alunos_por_faixa_append(alunos_por_faixa, aluno)

    assert aluno in resultado
    assert len(resultado) == 3


def test_alunos_por_faixa_nao_adiciona_aluno_duplicado():
    alunos_por_faixa = ["João", "Maria"]
    aluno = "Maria"

    resultado = alunos_por_faixa_append(alunos_por_faixa, aluno)

    assert resultado == alunos_por_faixa
    assert len(resultado) == 2


def test_alunos_por_faixa_lista_vazia_adiciona_aluno():
    alunos_por_faixa = []
    aluno = "Carlos"

    resultado = alunos_por_faixa_append(alunos_por_faixa, aluno)

    assert aluno in resultado
    assert len(resultado) == 1


def test_dias_append_dia_com_alunos():
    dias = []
    dia = 5
    alunos_por_dia = ["João", "Maria"]
    dias_append(dias, dia, alunos_por_dia)
    assert len(dias) == 1
    assert dias[0]["dia"] == "05"
    assert dias[0]["alunos_por_dia"] == alunos_por_dia


def test_dias_append_dia_com_alunos_vazio():
    dias = []
    dia = 10
    alunos_por_dia = []
    dias_append(dias, dia, alunos_por_dia)
    assert len(dias) == 1
    assert dias[0]["dia"] == "10"
    assert dias[0]["alunos_por_dia"] == alunos_por_dia


def test_dias_append_em_lista_existente():
    dias = [
        {"dia": "01", "alunos_por_dia": ["Ana"]},
        {"dia": "02", "alunos_por_dia": ["Carlos"]},
    ]
    dia = 3
    alunos_por_dia = ["Mariana"]
    dias_append(dias, dia, alunos_por_dia)
    assert len(dias) == 3
    assert dias[2]["dia"] == "03"
    assert dias[2]["alunos_por_dia"] == alunos_por_dia


def test_ordenar_alunos_matriculados(dicionario_de_alunos_matriculados):
    alunos_matriculados = AlunosMatriculadosPeriodoEscola.objects.all()
    assert alunos_matriculados.count() == 3
    assert alunos_matriculados[0].periodo_escolar.nome == "INTEGRAL"
    assert alunos_matriculados[1].periodo_escolar.nome == "MANHA"
    assert alunos_matriculados[2].periodo_escolar.nome == ""

    queryset = ordenar_alunos_matriculados(alunos_matriculados)
    assert queryset[0].periodo_escolar.nome == "MANHA"
    assert queryset[1].periodo_escolar.nome == "INTEGRAL"
    assert queryset[2].periodo_escolar.nome == ""


def test_cria_arquivo_excel():
    dados = [
        {"Nome": "Alice", "Idade": "25", "Cidade": "São Paulo"},
        {"Nome": "Bob", "Idade": "30", "Cidade": "Rio de Janeiro"},
    ]
    caminho_arquivo = Path("/tmp/teste.xlsx")
    cria_arquivo_excel(caminho_arquivo, dados)
    assert caminho_arquivo.exists(), "O arquivo Excel não foi criado."

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    assert [cell.value for cell in ws[1]] == list(dados[0].keys())

    for idx, row in enumerate(dados, start=2):
        assert [cell.value for cell in ws[idx]] == list(row.values())

    if os.path.exists(caminho_arquivo):
        os.remove(caminho_arquivo)

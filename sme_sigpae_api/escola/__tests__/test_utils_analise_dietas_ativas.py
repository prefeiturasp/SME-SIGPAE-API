import json
import os
import tempfile
import uuid
from datetime import date

import openpyxl
import pytest
from openpyxl import load_workbook

from sme_sigpae_api.escola.utils_analise_dietas_ativas import (
    dict_codigos_escolas,
    escreve_xlsx,
    escreve_xlsx_alunos_com_nascimento_diferente,
    escreve_xlsx_alunos_com_nome_diferente,
    escreve_xlsx_alunos_nao_matriculados_na_escola,
    escreve_xlsx_cod_diagnostico_inexistentes,
    escreve_xlsx_codescola_nao_existentes,
    escreve_xlsx_dados_sigpae,
    escreve_xlsx_primeira_aba,
    escreve_xlsx_protocolo_dieta_inexistentes,
    gera_dict_codigo_aluno_por_codigo_escola,
    gera_dict_codigos_escolas,
    get_codigo_eol_aluno,
    get_codigo_eol_escola,
    get_escolas_json,
    get_escolas_unicas,
    retorna_alunos_com_nascimento_diferente,
    retorna_alunos_com_nome_diferente,
    retorna_alunos_nao_matriculados_na_escola,
    retorna_cod_diagnostico_inexistentes,
    retorna_codigo_eol_escolas_nao_identificadas,
    retorna_dados_sigpae,
    retorna_protocolo_dieta_inexistentes,
    string_to_date,
)

pytestmark = pytest.mark.django_db


def test_get_codigo_eol_escola():
    assert get_codigo_eol_escola("123") == "000123"
    assert get_codigo_eol_escola(" 45 ") == "000045"
    assert get_codigo_eol_escola("123456") == "123456"
    assert get_codigo_eol_escola("1234567") == "1234567"
    assert get_codigo_eol_escola("") == "000000"


def test_get_codigo_eol_aluno():
    assert get_codigo_eol_aluno(1234) == "0001234"
    assert get_codigo_eol_aluno(" 789 ") == "0000789"
    assert get_codigo_eol_aluno("1234567") == "1234567"
    assert get_codigo_eol_aluno("12345678") == "12345678"
    assert get_codigo_eol_aluno("") == "0000000"


def test_gera_dict_codigos_escolas():
    items_codigos_escolas = [
        {"CÓDIGO UNIDADE": 123, "CODIGO EOL": 456},
        {"CÓDIGO UNIDADE": 789, "CODIGO EOL": 101112},
        {"CÓDIGO UNIDADE": 345, "CODIGO EOL": 6789},
    ]
    dict_codigos_escolas.clear()

    gera_dict_codigos_escolas(items_codigos_escolas)
    assert dict_codigos_escolas == {
        "123": "456",
        "789": "101112",
        "345": "6789",
    }


def test_gera_dict_codigo_aluno_por_codigo_escola(variaveis_globais_dieta):
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": "1002", "CodEOLAluno": "20002"},
    ]
    gera_dict_codigo_aluno_por_codigo_escola(items)
    esperado = {"0020001": "123456789", "0020002": "987654321"}
    assert esperado.items() <= variaveis_globais_dieta[1].items()


def test_gera_dict_codigo_aluno_por_codigo_escola_exception():
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": "1002", "CodEOLAluno": "20002"},
        {"CodEscola": "9999", "CodEOLAluno": "20003"},
    ]
    with pytest.raises(Exception):
        gera_dict_codigo_aluno_por_codigo_escola(items)


def test_get_escolas_unicas():
    items = [
        {"CodEscola": "101", "CodEOLAluno": "55"},
        {"CodEscola": "102", "CodEOLAluno": "997"},
        {"CodEscola": "102", "CodEOLAluno": "998"},
        {"CodEscola": "102", "CodEOLAluno": "999"},
    ]

    escolas = get_escolas_unicas(items)
    assert isinstance(escolas, set)
    assert len(escolas) == 2


def test_escreve_xlsx():
    codigo = set([111, 215, 433])
    output, nome_excel = escreve_xlsx(codigo)
    assert isinstance(nome_excel, str)
    assert nome_excel.lower().endswith(".xlsx")
    assert len(nome_excel.strip()) > 0  # Verifica se não é vazio

    workbook = load_workbook(output)
    nome_aba = "Código EOL das Escolas não identificadas no SIGPAE"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("codigo_eol_escola",)
    assert rows[1] == ("433",)
    assert rows[2] == ("215",)
    assert rows[3] == ("111",)

    workbook.close()
    if os.path.exists(nome_excel):
        os.remove(nome_excel)


def test_escreve_xlsx_primeira_aba():
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    r, excel = escreve_xlsx_primeira_aba("OK", arquivo_saida)
    workbook = load_workbook(excel)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("Este arquivo contém as planilhas:",)
    assert rows[1] == ("Código EOL das Escolas não identificadas no SIGPAE",)
    assert rows[2] == ("Código EOL dos Alunos não matriculados na escola",)
    assert rows[3] == ("CodEscola não existentes em unidades_da_rede...",)
    assert rows[4] == ("Alunos com nome diferente do EOL",)
    assert rows[5] == ("Alunos com data nascimento diferente do EOL",)
    assert rows[6] == ("Dados do SIGPAE para as escolas da planilha",)
    assert rows[7] == ("CodDiagnostico inexistentes",)
    assert rows[8] == ("ProtocoloDieta inexistentes",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_alunos_nao_matriculados_na_escola():
    alunos = {
        ("12345", "João", "789654"),
        ("12344", "Pedro", "789654"),
        ("12346", "Ana", "789664"),
    }
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    excel = escreve_xlsx_alunos_nao_matriculados_na_escola(alunos, arquivo_saida)

    workbook = load_workbook(excel)
    nome_aba = "Código EOL dos Alunos não matriculados na escola"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    # Nesse caso, os dados não vem com a mesma ordenação sempre
    esperado = {
        ("codigo_eol_aluno", "nome_aluno", "codigo_eol_escola"),
        ("12345", "João", "789654"),
        ("12344", "Pedro", "789654"),
        ("12346", "Ana", "789664"),
    }
    rows_set = set(rows)
    assert rows_set == esperado

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_codescola_nao_existentes():
    codigo = set([111, 215, 433])
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_codescola_nao_existentes(codigo, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "CodEscola não existentes em unidades_da_rede..."
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0] == ("CodEscola",)
    assert rows[1] == ("433",)
    assert rows[2] == ("215",)
    assert rows[3] == ("111",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_codigo_eol_escolas_nao_identificadas(variaveis_globais_dieta):
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": "1002", "CodEOLAluno": "20002"},
    ]
    excel, nome_excel = retorna_codigo_eol_escolas_nao_identificadas(items)

    assert isinstance(nome_excel, str)
    assert nome_excel.lower().endswith(".xlsx")
    assert len(nome_excel.strip()) > 0  # Verifica se não é vazio

    workbook = load_workbook(excel)
    nome_aba = "Código EOL das Escolas não identificadas no SIGPAE"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    esperado = {
        ("codigo_eol_escola",),
        ("123456789",),
        ("987654321",),
    }

    rows_set = set(rows)
    assert rows_set == esperado

    workbook.close()
    if os.path.exists(nome_excel):
        os.remove(nome_excel)


def test_retorna_codigo_eol_escolas_nao_identificadas_exception():
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": "1002", "CodEOLAluno": "20002"},
    ]

    with pytest.raises(Exception):
        excel, nome_excel = retorna_codigo_eol_escolas_nao_identificadas(items)


def test_get_escolas_json():
    escolas_data = {
        "escolas": [{"id": 1, "nome": "Escola A"}, {"id": 2, "nome": "Escola B"}]
    }

    with tempfile.NamedTemporaryFile(
        mode="w+", suffix=".json", delete=True
    ) as temp_file:
        json.dump(escolas_data, temp_file)
        temp_file.seek(0)
        result = get_escolas_json(temp_file.name)
        assert result == escolas_data


def test_retorna_alunos_nao_matriculados_na_escola(variaveis_globais_dieta):
    items = [
        {"CodEOLAluno": "20001", "NomeAluno": "Ana Silva"},
        {"CodEOLAluno": "20002", "NomeAluno": "Carlos Souza"},
    ]

    escolas = {
        "Escola A": [{"cd_aluno": "20001"}],
        "Escola B": [{"cd_aluno": "99999"}],
    }

    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)
    retorna_alunos_nao_matriculados_na_escola(items, escolas, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Código EOL dos Alunos não matriculados na escola"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    esperado = {
        ("codigo_eol_aluno", "nome_aluno", "codigo_eol_escola"),
        ("20001", "Ana Silva", "Escola A"),
        ("20002", "Carlos Souza", "Escola B"),
    }
    rows_set = set(rows)
    assert rows_set == esperado

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_dados_sigpae(escola, escola_cei):
    codigo = [escola.codigo_eol, escola_cei.codigo_eol]
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_dados_sigpae(codigo, arquivo_saida)
    workbook = load_workbook(arquivo_saida)
    nome_aba = "Dados do SIGPAE para as escolas da planilha"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "codigo_eol_escola",
        "nome_da_escola",
        "nome_dre",
        "lote",
        "tipo_gestao",
        "contato_email",
        "contato_telefone",
        "contato_telefone2",
        "contato_celular",
    )
    assert rows[1] == (
        escola.codigo_eol,
        escola.nome,
        escola.diretoria_regional.nome,
        escola.lote.nome,
        escola.tipo_gestao.nome,
        None,
        None,
        None,
        None,
    )
    assert rows[2] == (
        escola_cei.codigo_eol,
        escola_cei.nome,
        escola_cei.diretoria_regional.nome,
        None,
        escola_cei.tipo_gestao.nome,
        None,
        None,
        None,
        None,
    )

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_dados_sigpae(variaveis_globais_dieta, escola):
    dict_codigos_escolas.update({escola.codigo_eol: "123456789"})
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": escola.codigo_eol, "CodEOLAluno": "20002"},
    ]

    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    retorna_dados_sigpae(items, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Dados do SIGPAE para as escolas da planilha"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "codigo_eol_escola",
        "nome_da_escola",
        "nome_dre",
        "lote",
        "tipo_gestao",
        "contato_email",
        "contato_telefone",
        "contato_telefone2",
        "contato_celular",
    )

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_dados_sigpae_exception(escola):
    dict_codigos_escolas.update({escola.codigo_eol: "123456789"})
    items = [
        {"CodEscola": "1001", "CodEOLAluno": "20001"},
        {"CodEscola": escola.codigo_eol, "CodEOLAluno": "20002"},
    ]

    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    with pytest.raises(Exception):
        retorna_dados_sigpae(items, arquivo_saida)

    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_cod_diagnostico_inexistentes():
    codigo = [1254, 7869]
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_cod_diagnostico_inexistentes(codigo, arquivo_saida)
    workbook = load_workbook(arquivo_saida)
    nome_aba = "CodDiagnostico inexistentes"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("cod_diagnostico",)
    assert rows[1] == ("1254",)
    assert rows[2] == ("7869",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_protocolo_dieta_inexistentes():
    codigo = [1254, 7869]
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_protocolo_dieta_inexistentes(codigo, arquivo_saida)
    workbook = load_workbook(arquivo_saida)
    nome_aba = "ProtocoloDieta inexistentes"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("protocolo_dieta",)
    assert rows[1] == ("1254",)
    assert rows[2] == ("7869",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_cod_diagnostico_inexistentes(protocolos):
    items = [
        {"CodDiagnostico": "Protocolo1", "CODIGO EOL": 456},
        {"CodDiagnostico": "Protocolo2", "CODIGO EOL": 101112},
        {"CodDiagnostico": "Protocolo4", "CODIGO EOL": 6789},
    ]
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    retorna_cod_diagnostico_inexistentes(items, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "CodDiagnostico inexistentes"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("cod_diagnostico",)
    assert rows[1] == ("Protocolo4",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_protocolo_dieta_inexistentes(protocolos):
    items = [
        {"ProtocoloDieta": "Protocolo1", "CODIGO EOL": 456},
        {"ProtocoloDieta": "Protocolo2", "CODIGO EOL": 101112},
        {"ProtocoloDieta": "Protocolo4", "CODIGO EOL": 6789},
    ]
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    retorna_protocolo_dieta_inexistentes(items, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "ProtocoloDieta inexistentes"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0] == ("protocolo_dieta",)
    assert rows[1] == ("Protocolo4",)

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_string_to_date():
    assert string_to_date("2025-01-29", "%Y-%m-%d") == date(2025, 1, 29)
    assert string_to_date("29/01/2025", "%d/%m/%Y") == date(2025, 1, 29)
    assert string_to_date("29-01-2025 15:30:00", "%d-%m-%Y %H:%M:%S") == date(
        2025, 1, 29
    )
    with pytest.raises(ValueError):
        string_to_date("2025/01/29", "%Y-%m-%d")
    with pytest.raises(ValueError):
        string_to_date("29-01-2025", "%Y-%m-%d")
    with pytest.raises(ValueError):
        string_to_date("", "%Y-%m-%d")
    with pytest.raises(ValueError):
        string_to_date("abcdef", "%Y-%m-%d")


def test_escreve_xlsx_alunos_com_nome_diferente():
    alunos = {
        ("Joao Antonio", "João Antônio"),
        ("Maria A. Lopes", "Maria Almeida Lopes"),
    }
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_alunos_com_nome_diferente(alunos, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Alunos com nome diferente do EOL"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    esperado = {
        ("nome_aluno_planilha", "nome_aluno_eol"),
        ("Joao Antonio", "João Antônio"),
        ("Maria A. Lopes", "Maria Almeida Lopes"),
    }
    rows_set = set(rows)
    assert rows_set == esperado

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_escreve_xlsx_alunos_com_nascimento_diferente():
    alunos = {
        ("2022-05-01", "01/05/2022"),
        ("01/05/2022", "2021-05-01"),
    }
    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    escreve_xlsx_alunos_com_nascimento_diferente(alunos, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Alunos com nascimento diferente do EOL"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    esperado = {
        ("nascimento_planilha", "nascimento_eol"),
        ("2022-05-01", "01/05/2022"),
        ("01/05/2022", "2021-05-01"),
    }
    rows_set = set(rows)
    assert rows_set == esperado

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_alunos_com_nome_diferente(variaveis_globais_dieta):
    escolas = {
        "Escola A": [{"cd_aluno": "20001", "nm_aluno": "João Antônio"}],
        "Escola B": [{"cd_aluno": "20002", "nm_aluno": "Maria Lopes"}],
    }

    items = [
        {"CodEOLAluno": "20001", "NomeAluno": "Joao Antonio"},
        {"CodEOLAluno": "20002", "NomeAluno": "Maria Lopes"},
        {"CodEOLAluno": "20003", "NomeAluno": "Carlos Silva"},
    ]

    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    retorna_alunos_com_nome_diferente(items, escolas, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Alunos com nome diferente do EOL"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0] == ("nome_aluno_planilha", "nome_aluno_eol")
    assert rows[1] == ("Joao Antonio", "João Antônio")

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)


def test_retorna_alunos_com_nascimento_diferente(variaveis_globais_dieta):
    escolas = {
        "Escola A": [
            {
                "cd_aluno": "20001",
                "nm_aluno": "João Antônio",
                "dt_nascimento_aluno": "2020-04-11T00:00:00",
            }
        ],
        "Escola B": [
            {
                "cd_aluno": "20002",
                "nm_aluno": "Maria Lopes",
                "dt_nascimento_aluno": "2021-01-31T00:00:00",
            }
        ],
    }

    items = [
        {
            "CodEOLAluno": "20001",
            "NomeAluno": "Joao Antonio",
            "DataNascimento": "12/04/2020",
        },
        {
            "CodEOLAluno": "20002",
            "NomeAluno": "Maria Lopes",
            "DataNascimento": "31/01/2021",
        },
        {
            "CodEOLAluno": "20003",
            "NomeAluno": "Carlos Silva",
            "DataNascimento": "06/11/2019",
        },
    ]

    arquivo_saida = f"/tmp/{uuid.uuid4()}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(arquivo_saida)

    retorna_alunos_com_nascimento_diferente(items, escolas, arquivo_saida)

    workbook = load_workbook(arquivo_saida)
    nome_aba = "Alunos com nascimento diferente do EOL"
    assert nome_aba in workbook.sheetnames

    worksheet = workbook[nome_aba]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0] == ("nascimento_planilha", "nascimento_eol")
    assert rows[1] == ("2020-04-12", "2020-04-11")

    workbook.close()
    if os.path.exists(arquivo_saida):
        os.remove(arquivo_saida)

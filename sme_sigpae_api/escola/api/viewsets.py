import datetime
import json

from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.db.models import F, Max, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django_filters import rest_framework as filters
from openpyxl import Workbook, styles
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import permissions, status
from rest_framework.decorators import action
from rest_framework.mixins import (
    CreateModelMixin,
    ListModelMixin,
    RetrieveModelMixin,
    UpdateModelMixin,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet, ModelViewSet, ReadOnlyModelViewSet

from sme_sigpae_api.medicao_inicial.tasks import (
    exporta_relatorio_controle_frequencia_para_pdf,
)

from ...dados_comuns.permissions import (
    UsuarioCODAEDietaEspecial,
    UsuarioCODAEGabinete,
    UsuarioCODAEGestaoAlimentacao,
    UsuarioCODAENutriManifestacao,
    UsuarioDiretoriaRegional,
    UsuarioEscolaTercTotal,
    ViewSetActionPermissionMixin,
)
from ...dados_comuns.utils import get_ultimo_dia_mes, obter_primeiro_e_ultimo_dia_mes
from ...eol_servico.utils import EOLException
from ...escola.api.serializers import (
    AlunoSerializer,
    AlunoSimplesSerializer,
    AlunosMatriculadosPeriodoEscolaCompletoSerializer,
    CODAESerializer,
    DiretoriaRegionalParaFiltroSerializer,
    EscolaEolSimplesSerializer,
    EscolaParaFiltroSerializer,
    EscolaPeriodoEscolarSerializer,
    LogAlunosMatriculadosFaixaEtariaDiaSerializer,
    LoteNomeSerializer,
    LoteParaFiltroSerializer,
    LoteSerializer,
    LoteSimplesSerializer,
    PeriodoEscolarParaFiltroSerializer,
    TipoUnidadeParaFiltroSerializer,
)
from ...escola.api.serializers_create import (
    EscolaPeriodoEscolarCreateSerializer,
    FaixaEtariaSerializer,
    LoteCreateSerializer,
    MudancaFaixasEtariasCreateSerializer,
)
from ...inclusao_alimentacao.models import InclusaoAlimentacaoContinua
from ...paineis_consolidados.api.constants import FILTRO_DRE_UUID
from ..forms import AlunosPorFaixaEtariaForm
from ..models import (
    Aluno,
    AlunosMatriculadosPeriodoEscola,
    Codae,
    DiaCalendario,
    DiaSuspensaoAtividades,
    DiretoriaRegional,
    Escola,
    EscolaPeriodoEscolar,
    FaixaEtaria,
    GrupoUnidadeEscolar,
    LogAlteracaoQuantidadeAlunosPorEscolaEPeriodoEscolar,
    LogAlunosMatriculadosFaixaEtariaDia,
    LogAlunosMatriculadosPeriodoEscola,
    Lote,
    PeriodoEscolar,
    Subprefeitura,
    TipoGestao,
    TipoTurma,
    TipoUnidadeEscolar,
)
from ..services import NovoSGPServicoLogado, NovoSGPServicoLogadoException
from ..tasks import (
    gera_pdf_relatorio_alunos_matriculados_async,
    gera_xlsx_relatorio_alunos_matriculados_async,
)
from ..utils import (
    EscolaSimplissimaPagination,
    lotes_endpoint_filtrar_relatorio_alunos_matriculados,
    ordenar_alunos_matriculados,
)
from .filters import (
    AlunoFilter,
    DiretoriaRegionalFilter,
    LogAlunosMatriculadosFaixaEtariaDiaFilter,
)
from .permissions import PodeVerEditarFotoAlunoNoSGP
from .serializers import (
    AlunosMatriculadosPeriodoEscolaSerializer,
    DiaCalendarioSerializer,
    DiaSuspensaoAtividadesSerializer,
    DiretoriaRegionalCompletaSerializer,
    DiretoriaRegionalLookUpSerializer,
    DiretoriaRegionalSimplissimaSerializer,
    EscolaListagemSimplissimaComDRESelializer,
    EscolaParaFiltrosPeriodoEscolarReadOnlySerializer,
    EscolaParaFiltrosReadOnlySerializer,
    EscolaSimplesSerializer,
    EscolaSimplissimaSerializer,
    GrupoUnidadeEscolarSerializer,
    LogAlunosMatriculadosPeriodoEscolaSerializer,
    PeriodoEscolarSerializer,
    SubprefeituraSerializer,
    SubprefeituraSerializerSimples,
    TipoAlimentacaoSerializer,
    TipoGestaoSerializer,
    TipoUnidadeEscolarSerializer,
)
from .serializers_create import DiaSuspensaoAtividadesCreateManySerializer


class EscolaSimplesViewSet(
    ListModelMixin, RetrieveModelMixin, UpdateModelMixin, GenericViewSet
):
    lookup_field = "uuid"
    serializer_class = EscolaSimplesSerializer
    queryset = Escola.objects.all()


class EscolaSimplissimaViewSet(ListModelMixin, RetrieveModelMixin, GenericViewSet):
    lookup_field = "uuid"
    queryset = Escola.objects.all()
    serializer_class = EscolaSimplissimaSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    pagination_class = EscolaSimplissimaPagination
    filterset_fields = ["codigo_eol", "nome", "diretoria_regional__uuid"]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if "page" in request.query_params:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data})

    @action(detail=False, methods=["GET"], url_path=f"{FILTRO_DRE_UUID}")
    def filtro_por_diretoria_regional(self, request, dre_uuid=None):
        escolas = Escola.objects.filter(diretoria_regional__uuid=dre_uuid)
        return Response(self.get_serializer(escolas, many=True).data)


class EscolaParaFiltrosViewSet(ListModelMixin, GenericViewSet):
    lookup_field = "uuid"
    queryset = (
        Escola.objects.select_related("diretoria_regional", "lote", "tipo_unidade")
        .all()
        .order_by()
    )
    serializer_class = EscolaParaFiltrosReadOnlySerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ["diretoria_regional__uuid"]
    pagination_class = None

    @action(detail=True, url_path="periodos-escolares", url_name="periodos-escolares")
    def periodos_escolares(self, _, uuid: str):
        escola = get_object_or_404(Escola, uuid=uuid)
        serializer = EscolaParaFiltrosPeriodoEscolarReadOnlySerializer(
            instance=escola.periodos_escolares(), many=True
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, url_path="tipos-alimentacao", url_name="tipos-alimentacao")
    def tipos_alimentacao(self, _, uuid: str):
        escola = get_object_or_404(Escola, uuid=uuid)
        serializer = TipoAlimentacaoSerializer(
            instance=escola.tipos_alimentacao, many=True
        )
        return Response(serializer.data, status=status.HTTP_200_OK)


class EscolaSimplissimaComEolViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    queryset = Escola.objects.all()
    serializer_class = EscolaEolSimplesSerializer

    @action(detail=False, methods=["POST"], url_path="escolas-com-cod-eol")
    def escolas_com_cod_eol(self, request):
        escolas = self.get_queryset()
        lote = request.data.get("lote", None)
        if lote:
            escolas = escolas.filter(lote__uuid=lote)
        serializer = self.serializer_class(escolas, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["POST"], url_path="terc-total")
    def terc_total(self, request):
        escolas = self.get_queryset().filter(tipo_gestao__nome="TERC TOTAL")
        lotes = request.data.get("lotes", None)
        if lotes:
            escolas = escolas.filter(lote__uuid__in=lotes)
        serializer = self.serializer_class(escolas, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EscolaSimplissimaComDREViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    queryset = Escola.objects.all().prefetch_related("diretoria_regional")
    serializer_class = EscolaListagemSimplissimaComDRESelializer


class EscolaSimplissimaComDREUnpaginatedViewSet(EscolaSimplissimaComDREViewSet):
    pagination_class = None
    filterset_class = DiretoriaRegionalFilter

    @action(detail=False, methods=["GET"], url_path="terc-total")
    def terc_total(self, request):
        escolas = self.get_queryset().filter(tipo_gestao__nome="TERC TOTAL")
        escola = request.query_params.get("escola", None)
        dre = request.query_params.get("dre", None)
        terceirizada = request.query_params.get("terceirizada", None)
        nome_edital = request.query_params.get("nome_edital", None)
        if escola:
            escolas = escolas.filter(uuid=escola)
        if dre:
            escolas = escolas.filter(diretoria_regional__uuid=dre)
        if terceirizada:
            escolas = escolas.filter(lote__terceirizada__uuid=terceirizada)
        if nome_edital:
            escolas = escolas.filter(
                lote__contratos_do_lote__edital__numero=nome_edital
            )
        return Response(self.get_serializer(escolas, many=True).data)


class EscolaQuantidadeAlunosPorPeriodoEFaixaViewSet(GenericViewSet):
    lookup_field = "uuid"
    queryset = Escola.objects.all()

    def formata_resultado_faixa_etaria(self, counter_faixas_etarias, resultado):
        for periodo in counter_faixas_etarias:
            for faixa in counter_faixas_etarias[periodo]:
                for faixa_resultado in resultado:
                    if faixa == faixa_resultado["faixa_etaria"]["uuid"]:
                        faixa_resultado["count"] += counter_faixas_etarias[periodo][
                            faixa
                        ]
        return resultado

    @action(
        detail=True, url_path="somatorio-faixas-etarias/(?P<data_referencia_str>[^/.]+)"
    )
    def somatorio_faixas_etarias(self, request, uuid, data_referencia_str):
        try:
            form = AlunosPorFaixaEtariaForm({"data_referencia": data_referencia_str})

            if not form.is_valid():
                return Response(form.errors)

            escola = self.get_object()
            data_referencia = form.cleaned_data["data_referencia"]
            counter_faixas_etarias = escola.alunos_por_periodo_e_faixa_etaria(
                data_referencia
            )
            faixas_uuids = []

            for k in counter_faixas_etarias:
                faixas_uuids += counter_faixas_etarias[k].keys()
            faixas_uuids = list(set(faixas_uuids))

            resultado = []
            for faixa_etaria in FaixaEtaria.objects.filter(uuid__in=faixas_uuids):
                resultado.append(
                    {
                        "faixa_etaria": FaixaEtariaSerializer(faixa_etaria).data,
                        "count": 0,
                    }
                )
            resultado = self.formata_resultado_faixa_etaria(
                counter_faixas_etarias, resultado
            )
            return Response({"count": len(resultado), "results": resultado})
        except EOLException as error:
            return Response(
                data={"detail": str(error)}, status=status.HTTP_400_BAD_REQUEST
            )


class PeriodoEscolarViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    queryset = PeriodoEscolar.objects.all()
    serializer_class = PeriodoEscolarSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ("nome",)

    #  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(
        detail=True, url_path="alunos-por-faixa-etaria/(?P<data_referencia_str>[^/.]+)"
    )
    def alunos_por_faixa_etaria(self, request, uuid, data_referencia_str):  # noqa C901
        form = AlunosPorFaixaEtariaForm({"data_referencia": data_referencia_str})

        if not form.is_valid():
            return Response(form.errors)

        periodo_escolar = self.get_object()
        if periodo_escolar.nome == "PARCIAL":
            periodo_escolar = PeriodoEscolar.objects.get(nome="INTEGRAL")
        escola = self.request.user.vinculos.get(ativo=True).instituicao
        escola_periodo, created = EscolaPeriodoEscolar.objects.get_or_create(
            escola=escola, periodo_escolar=periodo_escolar
        )
        data_referencia = form.cleaned_data["data_referencia"]

        try:
            faixa_alunos = escola_periodo.alunos_por_faixa_etaria(
                data_referencia
            )  # noqa
        except ObjectDoesNotExist:
            return Response(
                {
                    "detail": "Não há faixas etárias cadastradas. Contate a coordenadoria CODAE."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except EOLException:
            return Response(
                {
                    "detail": "API EOL indisponível para carregar as faixas etárias. Tente novamente mais tarde."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []
        for uuid_faixa_etaria in faixa_alunos:
            results.append(
                {
                    "faixa_etaria": FaixaEtariaSerializer(
                        FaixaEtaria.objects.get(uuid=uuid_faixa_etaria)
                    ).data,
                    "count": faixa_alunos[uuid_faixa_etaria],
                }
            )
        results = sorted(results, key=lambda x: x["faixa_etaria"]["inicio"])

        return Response({"count": len(results), "results": results})

    @action(
        detail=False,
        methods=["GET"],
        url_path="inclusao-continua-por-mes",
        permission_classes=[
            UsuarioEscolaTercTotal
            | UsuarioDiretoriaRegional
            | UsuarioCODAEGestaoAlimentacao
            | UsuarioCODAEDietaEspecial
            | UsuarioCODAENutriManifestacao
            | UsuarioCODAEGabinete
        ],
    )
    def inclusao_continua_por_mes(self, request):
        try:
            for param in ["mes", "ano"]:
                if param not in request.query_params:
                    raise ValidationError(f"{param} é obrigatório via query_params")
            mes = request.query_params.get("mes")
            ano = request.query_params.get("ano")
            escola = request.query_params.get("escola")
            primeiro_dia_mes = datetime.date(int(ano), int(mes), 1)
            ultimo_dia_mes = get_ultimo_dia_mes(primeiro_dia_mes)
            instituicao = request.user.vinculo_atual.instituicao
            if (
                isinstance(instituicao, DiretoriaRegional)
                or isinstance(instituicao, Codae)
            ) and escola:
                instituicao = Escola.objects.get(uuid=escola)
            periodos = dict(
                InclusaoAlimentacaoContinua.objects.filter(
                    status="CODAE_AUTORIZADO", rastro_escola=instituicao
                )
                .filter(
                    data_inicial__lte=ultimo_dia_mes,
                    data_final__gte=primeiro_dia_mes,
                )
                .exclude(motivo__nome="ETEC")
                .values_list(
                    "quantidades_por_periodo__periodo_escolar__nome",
                    "quantidades_por_periodo__periodo_escolar__uuid",
                )
            )
            return Response({"periodos": periodos if len(periodos) else None})
        except ValidationError as e:
            return Response({"detail": e}, status=status.HTTP_400_BAD_REQUEST)


class PeriodosComMatriculadosPorUEViewSet(ReadOnlyModelViewSet):
    queryset = Escola.objects.all()

    def list(self, request, uuid=None):
        escola = self.get_object()
        periodos = escola.alunos_matriculados_por_periodo.filter(
            tipo_turma=TipoTurma.REGULAR.name, quantidade_alunos__gt=0
        ).values_list("periodo_escolar__nome", flat=True)
        return Response(periodos)

    def get_object(self):
        uuid = self.request.query_params.get("escola_uuid", None)
        return get_object_or_404(self.get_queryset(), uuid=uuid.rstrip("/"))

    def get_serializer_class(self):
        return AlunosMatriculadosPeriodoEscolaSerializer


class DiretoriaRegionalViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    queryset = DiretoriaRegional.objects.all()
    serializer_class = DiretoriaRegionalCompletaSerializer


class DiretoriaRegionalSimplissimaViewSet(ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    lookup_field = "uuid"
    queryset = DiretoriaRegional.objects.all()
    serializer_class = DiretoriaRegionalSimplissimaSerializer

    @action(detail=False, methods=["GET"], url_path="lista-completa")
    def lista_completa(self, request):
        response = {
            "results": DiretoriaRegionalLookUpSerializer(
                self.get_queryset(), many=True
            ).data
        }
        return Response(response)


class TipoGestaoViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    queryset = TipoGestao.objects.all()
    serializer_class = TipoGestaoSerializer


class SubprefeituraViewSet(ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    lookup_field = "uuid"
    queryset = Subprefeitura.objects.all()
    serializer_class = SubprefeituraSerializer

    @action(detail=False, methods=["get"], url_path="lista-completa")
    def lista_completa(self, request):
        response = {
            "results": SubprefeituraSerializerSimples(
                self.get_queryset(), many=True
            ).data
        }
        return Response(response)


class LoteViewSet(ModelViewSet):
    lookup_field = "uuid"
    serializer_class = LoteSimplesSerializer
    queryset = Lote.objects.all()

    @action(detail=False, methods=["GET"], url_path="meus-lotes-vinculados")
    def meus_lotes_vinculados(self, request):
        if request.user.tipo_usuario == "diretoriaregional":
            lotes = self.queryset.filter(
                diretoria_regional=request.user.vinculo_atual.instituicao
            )
        else:
            lotes = self.queryset.filter(
                terceirizada=request.user.vinculo_atual.instituicao
            )
        return Response({"results": LoteSerializer(lotes, many=True).data})

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return LoteCreateSerializer
        return LoteSimplesSerializer

    def destroy(self, request, *args, **kwargs):
        return Response(
            {"detail": "Não é permitido excluir um Lote com escolas associadas"},
            status=status.HTTP_401_UNAUTHORIZED,
        )


class LoteSimplesViewSet(ModelViewSet):
    lookup_field = "uuid"
    serializer_class = LoteNomeSerializer
    queryset = Lote.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ("uuid", "diretoria_regional__uuid", "terceirizada__uuid")


class CODAESimplesViewSet(ModelViewSet):
    lookup_field = "uuid"
    queryset = Codae.objects.all()
    serializer_class = CODAESerializer


class TipoUnidadeEscolarViewSet(ReadOnlyModelViewSet):
    lookup_field = "uuid"
    serializer_class = TipoUnidadeEscolarSerializer
    queryset = TipoUnidadeEscolar.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ("pertence_relatorio_solicitacoes_alimentacao",)


class LogAlunosMatriculadosPeriodoEscolaViewSet(ModelViewSet):
    serializer_class = LogAlunosMatriculadosPeriodoEscolaSerializer
    queryset = LogAlunosMatriculadosPeriodoEscola.objects.all()
    pagination_class = None

    def get_queryset(self):
        queryset = LogAlunosMatriculadosPeriodoEscola.objects.all()

        escola_uuid = self.request.query_params.get("escola_uuid", "")
        mes = self.request.query_params.get("mes", "")
        ano = self.request.query_params.get("ano", "")
        tipo_turma = self.request.query_params.get("tipo_turma", "")
        periodo_escolar = self.request.query_params.get("periodo_escolar", "")

        queryset = queryset.filter(
            escola__uuid=escola_uuid,
            criado_em__month=mes,
            criado_em__year=ano,
            tipo_turma=tipo_turma,
        )
        if (
            Escola.objects.get(uuid=escola_uuid).eh_cemei
            and "Infantil" in periodo_escolar
        ):
            periodo = periodo_escolar.replace("Infantil ", "")
            queryset = queryset.filter(periodo_escolar__nome=periodo)
            if "INTEGRAL" in periodo_escolar:
                queryset = queryset.filter(cei_ou_emei="EMEI")
            else:
                queryset = queryset.filter(cei_ou_emei="N/A")
        else:
            queryset = queryset.filter(
                escola__uuid=escola_uuid,
                criado_em__month=mes,
                criado_em__year=ano,
                tipo_turma=tipo_turma,
                periodo_escolar__uuid=periodo_escolar,
            )

        return queryset

    @action(detail=False, url_path="quantidade-por-data")
    def quantidade_por_data(self, request):
        try:
            data = self.request.query_params.get("data", "")
            cei_ou_emei = self.request.query_params.get("cei_ou_emei", "N/A")
            infantil_ou_fundamental = self.request.query_params.get(
                "infantil_ou_fundamental", "N/A"
            )
            escola_uuid = self.request.query_params.get("escola_uuid", "")

            if not data:
                raise ValidationError("campo data é obrigatório")

            if not escola_uuid:
                raise ValidationError("campo escola_uuid é obrigatório")

            today = datetime.date.today()
            if data == today.strftime("%Y-%m-%d"):
                data = today - datetime.timedelta(days=1)

            resultado = LogAlunosMatriculadosPeriodoEscola.objects.filter(
                escola__uuid=escola_uuid,
                criado_em=data,
                cei_ou_emei=cei_ou_emei,
                infantil_ou_fundamental=infantil_ou_fundamental,
            ).aggregate(soma=Sum("quantidade_alunos"))

            soma_quantidade_alunos = resultado["soma"] or 0

            return Response(soma_quantidade_alunos, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response({"detail": e}, status=status.HTTP_400_BAD_REQUEST)


class EscolaPeriodoEscolarViewSet(ModelViewSet):
    lookup_field = "uuid"

    def get_queryset(self):
        if self.request.user.tipo_usuario == "escola":
            escola = self.request.user.vinculos.get(ativo=True).instituicao
            return EscolaPeriodoEscolar.objects.filter(escola=escola)
        return EscolaPeriodoEscolar.objects.all()

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return EscolaPeriodoEscolarCreateSerializer
        return EscolaPeriodoEscolarSerializer

    def destroy(self, request, *args, **kwargs):
        return Response(
            {"detail": "Não é permitido excluir um periodo já existente"},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    @action(detail=False, url_path="escola/(?P<escola_uuid>[^/.]+)")
    def filtro_por_escola(self, request, escola_uuid=None):
        periodos = EscolaPeriodoEscolar.objects.filter(escola__uuid=escola_uuid)
        page = self.paginate_queryset(periodos)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    #  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(
        detail=True, url_path="alunos-por-faixa-etaria/(?P<data_referencia_str>[^/.]+)"
    )
    def alunos_por_faixa_etaria(self, request, uuid, data_referencia_str):  # noqa C901
        """
        EscolaCEI: Deve retornar a quantidade de alunos por faixa etária e período escolar através do uuid do
        PeriodoEscolar

        Outras escolas: Deve retornar a quantidade de alunos por faixa etária e período escolar através do uuid do
        EscolaPeriodoEscolar
        """

        form = AlunosPorFaixaEtariaForm({"data_referencia": data_referencia_str})

        if not form.is_valid():
            return Response(form.errors)

        if request.user.vinculo_atual:
            escola = request.user.vinculo_atual.instituicao
            if escola.eh_cei:
                escola_periodo = EscolaPeriodoEscolar.objects.get(
                    periodo_escolar__uuid=uuid, escola__uuid=escola.uuid
                )
        else:
            escola_periodo = self.get_object()

        data_referencia = form.cleaned_data["data_referencia"]

        try:
            faixa_alunos = escola_periodo.alunos_por_faixa_etaria(
                data_referencia
            )  # noqa
        except ObjectDoesNotExist:
            return Response(
                {
                    "detail": "Não há faixas etárias cadastradas. Contate a coordenadoria CODAE."
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except EOLException as erro:
            return Response(
                data={"detail": str(erro)}, status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        for uuid_faixa_etaria in faixa_alunos:
            results.append(
                {
                    "faixa_etaria": FaixaEtariaSerializer(
                        FaixaEtaria.objects.get(uuid=uuid_faixa_etaria)
                    ).data,
                    "count": faixa_alunos[uuid_faixa_etaria],
                }
            )
        results = sorted(results, key=lambda x: x["faixa_etaria"]["inicio"])

        return Response({"count": len(results), "results": results})

    #  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(
        detail=True, url_path="matriculados-na-data/(?P<data_referencia_str>[^/.]+)"
    )  # noqa C901
    def matriculados_na_data(self, request, uuid, data_referencia_str):
        form = AlunosPorFaixaEtariaForm({"data_referencia": data_referencia_str})

        if not form.is_valid():
            return Response(form.errors)

        escola_periodo = self.get_object()
        data_referencia = form.cleaned_data["data_referencia"]

        log = (
            LogAlteracaoQuantidadeAlunosPorEscolaEPeriodoEscolar.objects.filter(
                criado_em__gte=data_referencia,
                escola=escola_periodo.escola,
                periodo_escolar=escola_periodo.periodo_escolar,
            )
            .order_by("criado_em")
            .first()
        )

        if log:
            quantidade_alunos = log.quantidade_alunos_de
        else:
            quantidade_alunos = escola_periodo.quantidade_alunos

        return Response({"quantidade_alunos": {"convencional": quantidade_alunos}})


class AlunoViewSet(RetrieveModelMixin, ListModelMixin, GenericViewSet):
    lookup_field = "codigo_eol"
    queryset = Aluno.objects.all()
    serializer_class = AlunoSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = AlunoFilter

    @property
    def pagination_class(self):
        self._pagination_class = super(AlunoViewSet, self).pagination_class
        if self.request.query_params.get("sem_paginacao") == "true":
            self._pagination_class = None
        return self._pagination_class

    def get_serializer_class(self):
        if self.action == "list":
            return AlunoSimplesSerializer
        return self.serializer_class

    def get_queryset(self):
        if self.action == "retrieve":
            return self.queryset.select_related("escola__diretoria_regional")
        return self.queryset

    @action(
        detail=True,
        methods=["GET"],
        url_path="aluno-pertence-a-escola/(?P<escola_codigo_eol>[^/.]+)",
    )
    def aluno_pertence_a_escola(self, request, codigo_eol, escola_codigo_eol):
        escola = Escola.objects.filter(codigo_eol=escola_codigo_eol).first()
        aluno = Aluno.objects.filter(codigo_eol=codigo_eol).first()

        resposta = True if aluno and aluno.escola == escola else False
        return Response({"pertence_a_escola": resposta})

    @action(
        detail=True,
        methods=["GET"],
        url_path="ver-foto",
        permission_classes=(PodeVerEditarFotoAlunoNoSGP,),
    )
    def ver_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.pegar_foto_aluno(codigo_eol_)
            if response.status_code == status.HTTP_200_OK:
                return Response({"data": response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=True,
        methods=["POST"],
        url_path="atualizar-foto",
        permission_classes=(PodeVerEditarFotoAlunoNoSGP,),
    )
    def atualizar_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.atualizar_foto_aluno(
                codigo_eol_, request.FILES["file"]
            )
            if response.status_code == status.HTTP_200_OK:
                return Response({"data": response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=True,
        methods=["DELETE"],
        url_path="deletar-foto",
        permission_classes=(PodeVerEditarFotoAlunoNoSGP,),
    )
    def deletar_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.deletar_foto_aluno(codigo_eol_)
            if response.status_code == status.HTTP_200_OK:
                return Response({"data": response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=("GET",),
        url_path="quantidade-cemei-por-cei-emei",  # noqa C901
        permission_classes=(IsAuthenticated,),
    )
    def quantidade_cemei_por_cei_emei(self, request):
        try:
            codigo_eol_escola = request.query_params.get("codigo_eol_escola", None)
            manha_e_tarde_sempre = request.query_params.get(
                "manha_e_tarde_sempre", False
            )
            if not codigo_eol_escola:
                raise ValidationError(
                    "`codigo_eol_escola` como query_param é obrigatório"
                )
            escola = Escola.objects.get(codigo_eol=codigo_eol_escola)
            if not escola.eh_cemei:
                raise ValidationError("escola não é CEMEI")
            return Response(
                escola.quantidade_alunos_por_cei_emei(manha_e_tarde_sempre == "true"),
                status=status.HTTP_200_OK,
            )
        except ValidationError as e:
            return Response(e, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=("GET",),
        url_path="quantidade-alunos-por-periodo-cei-emei",  # noqa C901
        permission_classes=(IsAuthenticated,),
    )
    def quantidade_alunos_por_periodo_cei_emei(self, request):
        try:
            codigo_eol_escola = request.query_params.get("codigo_eol_escola", None)
            if not codigo_eol_escola:
                raise ValidationError(
                    "`codigo_eol_escola` como query_param é obrigatório"
                )
            escola = Escola.objects.get(codigo_eol=codigo_eol_escola)
            if not escola.eh_cemei:
                raise ValidationError("escola não é CEMEI")
            return Response(
                escola.quantidade_alunos_por_periodo_cei_emei, status=status.HTTP_200_OK
            )
        except ValidationError as e:
            return Response(e, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=("GET",),
        url_path="dados-aluno-nao-matriculado-detalhes-dieta",
        permission_classes=(IsAuthenticated,),
    )
    def dados_aluno_nao_matriculado_detalhes_dieta(self, request):
        try:
            codigo_eol_escola = request.query_params.get("codigo_eol_escola", None)
            nome_aluno = request.query_params.get("nome_aluno", False)
            if not codigo_eol_escola:
                raise ValidationError(
                    "`codigo_eol_escola` como query_param é obrigatório"
                )
            if not nome_aluno:
                raise ValidationError("`nome_aluno` como query_param é obrigatório")
            aluno = Aluno.objects.filter(
                nao_matriculado=True,
                escola__codigo_eol=codigo_eol_escola,
                nome=nome_aluno,
            ).first()
            return Response(self.get_serializer(aluno).data, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response(e, status=status.HTTP_400_BAD_REQUEST)


class FaixaEtariaViewSet(CreateModelMixin, ListModelMixin, GenericViewSet):
    queryset = FaixaEtaria.objects.filter(ativo=True)

    def get_serializer_class(self):
        if self.action == "create":
            return MudancaFaixasEtariasCreateSerializer
        return FaixaEtariaSerializer


class DiaCalendarioViewSet(ModelViewSet):
    serializer_class = DiaCalendarioSerializer
    queryset = DiaCalendario.objects.all()
    pagination_class = None

    def get_queryset(self):
        queryset = DiaCalendario.objects.all()

        escola_uuid = self.request.query_params.get("escola_uuid", "")
        mes = self.request.query_params.get("mes", "")
        ano = self.request.query_params.get("ano", "")

        queryset = queryset.filter(
            escola__uuid=escola_uuid, data__month=mes, data__year=ano
        )

        return queryset


def exportar_planilha_importacao_tipo_gestao_escola(request, **kwargs):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=planilha_importacao_tipo_gestao_escolas.xlsx"
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = "UNIDADES COM TIPO DE GESTÃO"
    headers = [
        "CÓDIGO EOL",
        "CÓDIGO CODAE",
        "NOME UNIDADE",
        "TIPO",
    ]
    _font = styles.Font(name="Calibri", sz=11)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill("solid", fgColor="ffff99")
        cabecalho.font = styles.Font(name="Calibri", size=11, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style="thin", color="000000"),
            right=styles.Side(border_style="thin", color="000000"),
            top=styles.Side(border_style="thin", color="000000"),
            bottom=styles.Side(border_style="thin", color="000000"),
        )
    dv = DataValidation(
        type="list",
        formula1='"PARCEIRA, DIRETA, MISTA, TERCEIRIZADA TOTAL"',
        allow_blank=True,
    )
    dv.error = "Tipo Inválido"
    dv.errorTitle = "Tipo não permitido"
    ws.add_data_validation(dv)
    dv.add("D2:H1048576")

    for colunas in ws.columns:
        unmerged_cells = list(
            filter(
                lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells,
                colunas,
            )
        )
        length = max(len(str(cell.value)) for cell in unmerged_cells)
        ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.3
    workbook.save(response)

    return response


class RelatorioAlunosMatriculadosViewSet(ModelViewSet):
    queryset = AlunosMatriculadosPeriodoEscola.objects.all()

    def obter_alunos_matriculados(self, request, lotes):
        query_params = request.query_params
        if query_params.getlist("lotes[]"):
            lotes = lotes.filter(uuid__in=query_params.getlist("lotes[]"))
        if query_params.getlist("diretorias_regionais[]"):
            lotes = lotes.filter(
                diretoria_regional__uuid__in=query_params.getlist(
                    "diretorias_regionais[]"
                )
            )
        escolas_uuids = lotes.values_list("escolas__uuid", flat=True).distinct()
        alunos_matriculados = AlunosMatriculadosPeriodoEscola.objects.filter(
            escola__uuid__in=escolas_uuids, escola__tipo_gestao__nome="TERC TOTAL"
        )
        alunos_matriculados = self.filtra_alunos_matriculados(
            alunos_matriculados, query_params
        )
        return alunos_matriculados

    def filtra_alunos_matriculados(self, queryset, query_params):
        if query_params.getlist("diretorias_regionais[]"):
            queryset = queryset.filter(
                escola__diretoria_regional__uuid__in=query_params.getlist(
                    "diretorias_regionais[]"
                )
            )
        if query_params.getlist("tipos_unidades[]"):
            tipos = query_params.getlist("tipos_unidades[]")
            queryset = queryset.filter(escola__tipo_unidade__uuid__in=tipos)
        if query_params.getlist("unidades_educacionais[]"):
            unidades_eudacionais = query_params.getlist("unidades_educacionais[]")
            queryset = queryset.filter(escola__uuid__in=unidades_eudacionais)
        if query_params.getlist("tipos_turmas[]"):
            tipos_turmas = query_params.getlist("tipos_turmas[]")
            queryset = queryset.filter(tipo_turma__in=tipos_turmas)
        return queryset

    @action(detail=False, methods=["GET"], url_path="filtros")
    def filtros(self, request):
        instituicao = request.user.vinculo_atual.instituicao
        if isinstance(instituicao, Codae):
            lotes = Lote.objects.all()
            diretorias_regionais = DiretoriaRegional.objects.all()
        else:
            lotes = instituicao.lotes.filter(escolas__isnull=False).distinct()
            diretorias_regionais_uuids = lotes.values_list(
                "diretoria_regional__uuid", flat=True
            ).distinct()
            diretorias_regionais = DiretoriaRegional.objects.filter(
                uuid__in=diretorias_regionais_uuids
            )
        escolas_uuids = lotes.values_list("escolas__uuid", flat=True).distinct()
        escolas = Escola.objects.filter(
            uuid__in=escolas_uuids, tipo_gestao__nome="TERC TOTAL"
        )
        tipos_unidade_uuids = escolas.values_list(
            "tipo_unidade__uuid", flat=True
        ).distinct()
        tipos_unidade_escolar = TipoUnidadeEscolar.objects.filter(
            uuid__in=tipos_unidade_uuids
        )
        tipos_turmas = [name for name, _ in TipoTurma.choices()]
        filtros = {
            "lotes": LoteParaFiltroSerializer(lotes, many=True).data,
            "diretorias_regionais": DiretoriaRegionalParaFiltroSerializer(
                diretorias_regionais, many=True
            ).data,
            "tipos_unidade_escolar": TipoUnidadeParaFiltroSerializer(
                tipos_unidade_escolar, many=True
            ).data,
            "escolas": EscolaParaFiltroSerializer(escolas, many=True).data,
            "tipos_turmas": tipos_turmas,
        }
        return Response(filtros, status=status.HTTP_200_OK)

    @action(detail=False, methods=["GET"], url_path="filtrar")
    def filtrar(self, request):
        instituicao = request.user.vinculo_atual.instituicao
        lotes = lotes_endpoint_filtrar_relatorio_alunos_matriculados(
            instituicao, Codae, Lote
        )
        alunos_matriculados = self.obter_alunos_matriculados(request, lotes)
        alunos_matriculados_ordenados = ordenar_alunos_matriculados(alunos_matriculados)
        page = self.paginate_queryset(alunos_matriculados_ordenados)
        serializer = AlunosMatriculadosPeriodoEscolaCompletoSerializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(detail=False, methods=["GET"], url_path="gerar-pdf")
    def gerar_pdf(self, request):
        user = request.user.get_username()
        instituicao = request.user.vinculo_atual.instituicao
        lotes = lotes_endpoint_filtrar_relatorio_alunos_matriculados(
            instituicao, Codae, Lote
        )
        alunos_matriculados = self.obter_alunos_matriculados(request, lotes)
        uuids = [str(matriculados.uuid) for matriculados in alunos_matriculados]
        gera_pdf_relatorio_alunos_matriculados_async.delay(
            user=user, nome_arquivo="relatorio_alunos_matriculados.pdf", uuids=uuids
        )
        return Response(
            dict(detail="Solicitação de geração de arquivo recebida com sucesso."),
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["GET"], url_path="gerar-xlsx")
    def gerar_xlsx(self, request):
        user = request.user.get_username()
        instituicao = request.user.vinculo_atual.instituicao
        lotes = lotes_endpoint_filtrar_relatorio_alunos_matriculados(
            instituicao, Codae, Lote
        )
        alunos_matriculados = self.obter_alunos_matriculados(request, lotes)
        uuids = [str(matriculados.uuid) for matriculados in alunos_matriculados]
        gera_xlsx_relatorio_alunos_matriculados_async.delay(
            user=user, nome_arquivo="relatorio_alunos_matriculados.xlsx", uuids=uuids
        )
        return Response(
            dict(detail="Solicitação de geração de arquivo recebida com sucesso."),
            status=status.HTTP_200_OK,
        )


class LogAlunosMatriculadosFaixaEtariaDiaViewSet(ListModelMixin, GenericViewSet):
    serializer_class = LogAlunosMatriculadosFaixaEtariaDiaSerializer
    queryset = LogAlunosMatriculadosFaixaEtariaDia.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = LogAlunosMatriculadosFaixaEtariaDiaFilter
    pagination_class = None


class DiaSuspensaoAtividadesViewSet(ViewSetActionPermissionMixin, ModelViewSet):
    permission_action_classes = {
        "list": [UsuarioCODAEGestaoAlimentacao],
        "create": [UsuarioCODAEGestaoAlimentacao],
        "delete": [UsuarioCODAEGestaoAlimentacao],
    }
    queryset = (
        DiaSuspensaoAtividades.objects.select_related(
            "tipo_unidade", "criado_por", "edital"
        )
        .annotate(edital_numero=F("edital__numero"))
        .all()
    )
    lookup_field = "uuid"
    pagination_class = None

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return DiaSuspensaoAtividadesCreateManySerializer
        return DiaSuspensaoAtividadesSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        if "mes" in self.request.query_params and "ano" in self.request.query_params:
            queryset = queryset.filter(
                data__month=self.request.query_params.get("mes"),
                data__year=self.request.query_params.get("ano"),
            )
        return queryset

    def create(self, request, *args, **kwargs):
        try:
            return super(DiaSuspensaoAtividadesViewSet, self).create(
                request, *args, **kwargs
            )
        except AssertionError as error:
            if str(error) == "`create()` did not return an object instance.":
                return Response(status=status.HTTP_201_CREATED)
            return Response({"detail": str(error)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        DiaSuspensaoAtividades.objects.filter(
            data=instance.data, tipo_unidade=instance.tipo_unidade
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class GrupoUnidadeEscolarViewSet(ModelViewSet):
    lookup_field = "uuid"
    serializer_class = GrupoUnidadeEscolarSerializer
    queryset = GrupoUnidadeEscolar.objects.all()


class RelatorioControleDeFrequenciaViewSet(ModelViewSet):
    def get_queryset(self):
        escola = self.request.user.vinculo_atual.instituicao
        if escola.eh_cei or escola.eh_cemei:
            return LogAlunosMatriculadosFaixaEtariaDia.objects.all()
        return LogAlunosMatriculadosPeriodoEscola.objects.all()

    def filtrar_alunos_matriculados(
        self, queryset, escola_eh_cei_ou_cemei, periodos_uuids
    ):
        max_quantidades_por_periodo = {}
        param_quantidade = (
            "quantidade" if escola_eh_cei_ou_cemei else "quantidade_alunos"
        )

        soma_quantidades_por_periodo = queryset.values(
            "periodo_escolar__uuid", "criado_em__date"
        ).annotate(soma_quantidade=Sum(param_quantidade))

        for periodo_uuid in json.loads(periodos_uuids):
            periodo = PeriodoEscolar.objects.get(uuid=periodo_uuid)

            if periodo.nome == "INTEGRAL":
                periodo_parcial = PeriodoEscolar.objects.get(nome="PARCIAL")

                total_matriculados_integral = self.get_total_matriculados_por_periodo(
                    soma_quantidades_por_periodo, periodo
                )
                total_matriculados_parcial = self.get_total_matriculados_por_periodo(
                    soma_quantidades_por_periodo, periodo_parcial
                )
                total_matriculados = (
                    total_matriculados_integral + total_matriculados_parcial
                )
            else:
                total_matriculados = self.get_total_matriculados_por_periodo(
                    soma_quantidades_por_periodo, periodo
                )
            max_quantidades_por_periodo[periodo.nome] = (
                total_matriculados if total_matriculados else 0
            )
        response = {
            "periodos": max_quantidades_por_periodo,
            "total_matriculados": sum(max_quantidades_por_periodo.values()),
        }

        return response

    def get_total_matriculados_por_periodo(self, soma_logs_alunos_por_periodo, periodo):
        return (
            soma_logs_alunos_por_periodo.filter(periodo_escolar=periodo).aggregate(
                max_quantidade=Max("soma_quantidade")
            )["max_quantidade"]
            or 0
        )

    def validar_periodos(self, filtros, periodos_uuids):
        periodos = PeriodoEscolar.objects.filter(uuid__in=json.loads(periodos_uuids))
        periodo_integral = PeriodoEscolar.objects.get(nome="INTEGRAL")

        if periodo_integral in periodos:
            periodos_uuids_list = json.loads(periodos_uuids)
            periodo_parcial = PeriodoEscolar.objects.get(nome="PARCIAL")
            periodos_uuids_list.append(str(periodo_parcial.uuid))
            periodos_uuids = json.dumps(periodos_uuids_list)
        filtros["periodo_escolar__uuid__in"] = json.loads(periodos_uuids)
        return filtros

    def validar_datas(self, filtros, data_inicial, data_final, escola_eh_cei_ou_cemei):
        nome_campo = "data" if escola_eh_cei_ou_cemei else "criado_em"

        ano, mes, dia_inicial = data_inicial.split("-")
        datetime_inicial = datetime.date(int(ano), int(mes), int(dia_inicial))
        hoje = datetime.date.today()
        ontem = hoje - datetime.timedelta(days=1)

        if data_inicial == data_final and not datetime_inicial >= hoje:
            filtros[nome_campo] = data_inicial
        elif datetime_inicial >= hoje:
            filtros[nome_campo] = ontem
        else:
            if data_inicial:
                filtros[f"{nome_campo}__gte"] = data_inicial
            if data_final:
                filtros[f"{nome_campo}__lte"] = data_final
        return filtros

    @action(detail=False, methods=["GET"], url_path="meses-anos")
    def meses_anos(self, _):
        data_atual = datetime.date.today()
        mes_ano_atual = {"mes": data_atual.month, "ano": data_atual.year}

        mes_anterior = data_atual.month - 1 if data_atual.month > 1 else 12
        ano_anterior = data_atual.year - 1 if data_atual.month == 1 else data_atual.year
        mes_ano_anterior = {"mes": mes_anterior, "ano": ano_anterior}

        mes_posterior = data_atual.month + 1 if data_atual.month < 12 else 1
        ano_posterior = (
            data_atual.year + 1 if data_atual.month == 12 else data_atual.year
        )
        mes_ano_posterior = {"mes": mes_posterior, "ano": ano_posterior}

        meses_anos = [mes_ano_anterior, mes_ano_atual, mes_ano_posterior]

        return Response({"results": meses_anos}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["GET"], url_path="filtros")
    def filtros(self, request):
        escola = request.user.vinculo_atual.instituicao
        mes = request.query_params.get("mes")
        ano = request.query_params.get("ano")

        if mes is None or ano is None:
            return Response(
                data={
                    "detail": "É necessário informar o mês e o ano para extrair o relatório"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            escola_eh_cei_ou_cemei = escola.eh_cei or escola.eh_cemei

            if escola_eh_cei_ou_cemei:
                log_alunos = escola.logs_alunos_matriculados_por_faixa_etaria.filter(
                    quantidade__gte=1, criado_em__year=ano
                )
            else:
                log_alunos = escola.logs_alunos_matriculados_por_periodo.filter(
                    tipo_turma="REGULAR", quantidade_alunos__gte=1, criado_em__year=ano
                )

            periodos_ids = log_alunos.values_list("periodo_escolar", flat=True)
            periodos = PeriodoEscolar.objects.filter(id__in=periodos_ids).exclude(
                nome="PARCIAL"
            )

            datas = obter_primeiro_e_ultimo_dia_mes(int(ano), int(mes))
            data_inicial = datas[0]
            data_final = datas[1]

            response = {
                "periodos": PeriodoEscolarParaFiltroSerializer(
                    periodos, many=True
                ).data,
                "data_inicial": data_inicial,
                "data_final": data_final,
            }
            return Response(response, status=status.HTTP_200_OK)
        except ObjectDoesNotExist as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET"], url_path="filtrar")
    def filtrar(self, request):
        escola = request.user.vinculo_atual.instituicao
        queryset = self.get_queryset().filter(escola=escola)
        escola_eh_cei_ou_cemei = escola.eh_cei or escola.eh_cemei

        filtros = {}

        data_inicial = request.query_params.get("data_inicial")
        data_final = request.query_params.get("data_final")
        if data_inicial or data_final:
            filtros = self.validar_datas(
                filtros, data_inicial, data_final, escola_eh_cei_ou_cemei
            )

        periodos_uuids = request.query_params.get("periodos")
        if periodos_uuids:
            filtros = self.validar_periodos(filtros, periodos_uuids)

        if filtros:
            queryset = queryset.filter(**filtros)

        response = self.filtrar_alunos_matriculados(
            queryset, escola_eh_cei_ou_cemei, periodos_uuids
        )

        return Response(response, status=status.HTTP_200_OK)

    @action(
        detail=False,
        url_path="imprimir-pdf",
    )
    def relatorio_controle_frequencia_exportar_pdf(self, request: Request):
        escola_uuid = request.user.vinculo_atual.instituicao.uuid
        query_params = request.query_params

        mes_ano = query_params.get("mes_ano")
        if not mes_ano:
            return Response(
                data="É necessário informar o mês/ano de referência",
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            query_params_dict = query_params.dict()
            exporta_relatorio_controle_frequencia_para_pdf.delay(
                user=request.user.get_username(),
                nome_arquivo="controle-frequencia.pdf",
                query_params=query_params_dict,
                escola_uuid=escola_uuid,
            )

            return Response(
                data={
                    "detail": "Solicitação de geração de arquivo recebida com sucesso."
                },
                status=status.HTTP_200_OK,
            )
        except Exception:
            return Response(
                data={"detail": "Verifique os parâmetros e tente novamente"},
                status=status.HTTP_400_BAD_REQUEST,
            )

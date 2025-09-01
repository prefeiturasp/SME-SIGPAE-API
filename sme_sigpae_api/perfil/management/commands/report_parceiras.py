import ast
from difflib import SequenceMatcher

import pdfplumber
from django.core.management.base import BaseCommand
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Lê relatórios e cruza usuários com dados de diretores."

    def get_array_diretores(self):
        dados = []
        with pdfplumber.open("usuarios_parceiras.pdf") as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or "Carimbo" in row[0]:
                            continue
                        carimbo, unidade, diretora, cpf = row
                        dados.append(
                            {
                                "data": carimbo.strip(),
                                "unidade": unidade.strip(),
                                "nome": diretora.strip(),
                                "cpf": cpf.strip()
                                .replace(".", "")
                                .replace("-", "")
                                .replace(" ", ""),
                            }
                        )
        return dados

    def handle(self, *args, **options):
        wb = load_workbook("report_parceiras.xlsx")
        ws = wb.active
        lista_diretores = self.get_array_diretores()

        index_erro = -1
        index_email = 2
        index_cargo = 3

        self.processar_usuarios(
            ws,
            lista_diretores,
            cond=lambda row: row[index_erro],  # tem erro
            titulo="Usuários não encontrados",
        )

        self.processar_usuarios(
            ws,
            lista_diretores,
            cond=lambda row: not row[index_email]
            and not row[index_erro],  # email vazio e sem erro
            titulo="Usuários sem e-mail",
        )

        self.processar_usuarios(
            ws,
            lista_diretores,
            cond=lambda row: row[index_email]
            and row[index_email].endswith("@"),  # email inválido
            titulo="Usuários com e-mail incorreto",
        )

        self.processar_usuarios(
            ws,
            lista_diretores,
            cond=lambda row: not row[index_cargo]
            and not row[index_erro],  # cargo vazio e sem erro
            titulo="Usuários sem cargo",
        )

        self.processar_usuarios(
            ws,
            lista_diretores,
            cond=lambda row: row[index_cargo]
            and ast.literal_eval(row[index_cargo])[0]["codigoCargo"]
            != 1,  # cargo não é DIRETOR
            titulo="Usuários não são diretores",
        )

        self.usuarios_escola_diferente(ws, lista_diretores)

    def processar_usuarios(self, ws, lista_diretores, cond, titulo):
        index_cpf = 1
        self.stdout.write(self.style.SUCCESS(f"\n{titulo}:\n"))

        for row in ws.iter_rows(values_only=True, min_row=2):
            if cond(row):
                try:
                    diretor = next(
                        res for res in lista_diretores if row[index_cpf] in res["cpf"]
                    )
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"{diretor['unidade']} - {diretor['nome']} - {diretor['cpf']}"
                        )
                    )
                except StopIteration:
                    print("nao achou: ", row[index_cpf])

    def normaliza_unidade(self, unidade):
        if unidade.startswith("CEI"):
            unidade = unidade.replace("CEI ", "")
        elif unidade.startswith("Cei"):
            unidade = unidade.replace("Cei ", "")
        elif unidade.startswith("CR.P.CONV."):
            unidade = unidade.replace("CR.P.CONV.", "")
        return unidade

    def usuarios_escola_diferente(self, ws, lista_diretores):
        index_cpf = 1
        index_cargo = 3

        self.stdout.write(self.style.SUCCESS("\nUsuários com escola errada:\n"))

        for row in ws.iter_rows(values_only=True, min_row=2):
            if row[index_cargo]:
                try:
                    diretor = next(
                        res for res in lista_diretores if row[index_cpf] in res["cpf"]
                    )
                    unidade_pdf = self.normaliza_unidade(diretor["unidade"]).upper()
                    unidade_xlsx = (
                        ast.literal_eval(row[index_cargo])[0]["descricaoUnidade"]
                        .replace("CR.P.CONV - ", "")
                        .replace("CEI INDIR - ", "")
                        .upper()
                    )
                    similaridade = SequenceMatcher(
                        None, unidade_pdf, unidade_xlsx
                    ).ratio()
                    if similaridade < 0.6:
                        self.stdout.write(
                            self.style.SUCCESS(
                                f"{diretor['unidade']} - {diretor['nome']} - {diretor['cpf']}"
                            )
                        )
                except StopIteration:
                    print("nao achou: ", row[index_cpf])
